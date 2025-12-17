import pandas as pd
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Set, Dict, Tuple, Optional, Any
import random
import openpyxl
from datetime import datetime, timedelta
import typer
from itertools import combinations
from collections import defaultdict, Counter
from html import escape
import math
import hashlib
import json
import warnings
from io import BytesIO


DEFAULT_START_TIME_HHMM = "12:50"
DEFAULT_ROUND_MINUTES = 13

TEAM_LIST_TEMPLATE_SHEET_NAME = "チーム一覧"
TEAM_LIST_TEMPLATE_HEADERS = [
    "ペア名",
    "氏名",
    "優先対戦",
    "優先対戦相手",
]


TEAM_LIST_SAMPLE_SHEET_NAME = TEAM_LIST_TEMPLATE_SHEET_NAME
TEAM_LIST_SAMPLE_HEADERS = TEAM_LIST_TEMPLATE_HEADERS


def build_team_list_sample_rows() -> List[List[str]]:
    """Return dummy rows with clearly non-personal placeholder data."""

    def row(pair: str, members: str) -> List[str]:
        return [pair, members, "", ""]

    rows: List[List[str]] = []
    # A: 4 teams (unique groups)
    rows += [
        row("サンプルA1", "TEST_A1P1, TEST_A1P2"),
        row("サンプルA2", "TEST_A2P1, TEST_A2P2"),
        row("サンプルA3", "TEST_A3P1, TEST_A3P2"),
        row("サンプルA4", "TEST_A4P1, TEST_A4P2"),
    ]
    # B: 4 teams (unique groups)
    rows += [
        row("サンプルB1", "TEST_B1P1, TEST_B1P2"),
        row("サンプルB2", "TEST_B2P1, TEST_B2P2"),
        row("サンプルB3", "TEST_B3P1, TEST_B3P2"),
        row("サンプルB4", "TEST_B4P1, TEST_B4P2"),
    ]
    # C: 4 teams (unique groups)
    rows += [
        row("サンプルC1", "TEST_C1P1, TEST_C1P2"),
        row("サンプルC2", "TEST_C2P1, TEST_C2P2"),
        row("サンプルC3", "TEST_C3P1, TEST_C3P2"),
        row("サンプルC4", "TEST_C4P1, TEST_C4P2"),
    ]
    return rows


def build_team_list_sample_bytes(
    sheet_name: str = TEAM_LIST_SAMPLE_SHEET_NAME,
    headers: List[str] = TEAM_LIST_SAMPLE_HEADERS,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers))
    for r in build_team_list_sample_rows():
        ws.append(r)
    ws.freeze_panes = "A2"
    with BytesIO() as bio:
        wb.save(bio)
        return bio.getvalue()


def build_team_list_template_bytes(
    sheet_name: str = TEAM_LIST_TEMPLATE_SHEET_NAME,
    headers: List[str] = TEAM_LIST_TEMPLATE_HEADERS,
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(headers))
    ws.freeze_panes = "A2"
    with BytesIO() as bio:
        wb.save(bio)
        return bio.getvalue()


def _parse_hhmm(value: str) -> tuple[int, int]:
    """Parse HH:MM string."""
    try:
        parts = value.strip().split(":")
        if len(parts) != 2:
            raise ValueError
        hour = int(parts[0])
        minute = int(parts[1])
    except Exception as e:
        raise ValueError(f"Invalid time '{value}'. Use HH:MM (e.g. 12:50).") from e
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        raise ValueError(f"Invalid time '{value}'. Use HH:MM with 00:00..23:59.")
    return hour, minute


def _base_datetime_from_hhmm(value: str) -> datetime:
    hour, minute = _parse_hhmm(value)
    return datetime(2000, 1, 1, hour, minute)


def apply_round_times(matches: List["Match"], start_time_hhmm: str, round_minutes: int) -> None:
    if round_minutes <= 0:
        raise ValueError("round_minutes must be positive")
    base = _base_datetime_from_hhmm(start_time_hhmm)
    step = timedelta(minutes=int(round_minutes))
    for m in matches:
        m.start_time = base + (m.round_num - 1) * step


@dataclass
class Team:
    name: str
    members: str
    level: str
    group: str
    preferred_opponents: List[str] = field(default_factory=list)
    matches: int = 0
    opponents: Set[str] = field(default_factory=set)
    groups_faced: Set[str] = field(default_factory=set)
    last_round: int = 0


@dataclass
class Match:
    round_num: int
    court: int
    team1: Team
    team2: Team
    start_time: datetime


GROUP_SOFT_TARGET = 4  # preferred unique groups per team when building level graphs
GROUP_HARD_TARGET = 3   # minimum unique groups required before accepting a graph
GROUP_REPEAT_CAP = 2    # max times the same group pairing is allowed
MAX_GROUP_GRAPH_RESTARTS = 6

# Default target matches per team (historical value). This is now configurable.
DEFAULT_MATCHES_PER_TEAM = 6
TARGET_MATCHES_PER_TEAM = DEFAULT_MATCHES_PER_TEAM


def compute_auto_matches_per_team(num_teams: int, num_rounds: int, courts: int) -> int:
    """Compute the maximum equal matches-per-team that fits capacity.

    Total match slots = num_rounds * courts.
    If every team plays m matches, total matches needed is num_teams*m/2 (must be integer).
    We pick the largest m that fits capacity. When num_teams is odd, m must be even.
    """

    if num_teams <= 0 or num_rounds <= 0 or courts <= 0:
        return 0
    slots = num_rounds * courts
    m = (2 * slots) // num_teams
    # total matches must be integer => num_teams*m must be even
    if (num_teams % 2 == 1) and (m % 2 == 1):
        m -= 1
    # cannot exceed round-robin upper bound
    m = min(m, max(0, num_teams - 1))
    return max(0, m)


def compute_hard_feasible_matches_per_team(teams: List[Team]) -> int:
    """Compute an upper bound for matches-per-team under hard constraints.

    Current hard model:
    - Matches are only scheduled within the same level.
    - Teams cannot play against the same group.

    Therefore, for each team, the maximum number of distinct opponents is the
    count of teams in the same level whose group differs. We take the minimum
    across all teams as a safe global upper bound.

    If group is missing for either side, we treat the pairing as allowed
    (cannot enforce the constraint reliably).
    """

    if not teams:
        return 0
    by_level: Dict[str, List[Team]] = defaultdict(list)
    for t in teams:
        by_level[t.level].append(t)

    caps: List[int] = []
    for level_teams in by_level.values():
        for t in level_teams:
            possible = 0
            for other in level_teams:
                if other.name == t.name:
                    continue
                if not t.group or not other.group or (t.group != other.group):
                    possible += 1
            caps.append(possible)
    return max(0, min(caps) if caps else 0)


def expected_total_matches(num_teams: int, matches_per_team: int) -> int:
    return (num_teams * matches_per_team) // 2
LEVEL_SEGMENTS = {
    'A': (1, 5),
    'B': (6, 10),
    'C': (11, 15),
}


def load_teams(file_path: str) -> List[Team]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active
    data = list(sheet.values)
    wb.close()
    if not data:
        return []
    df = pd.DataFrame(data[1:], columns=data[0])
    teams: List[Team] = []

    def _norm_header(v: Any) -> str:
        if v is None:
            return ""
        return str(v).strip().replace("\u3000", " ")

    def _pick_col(*, include_any: tuple[str, ...], exclude_any: tuple[str, ...] = ()) -> str | None:
        """Pick the first matching column name by keyword containment.

        This supports templates like:
          - 'ペア名' / '氏名'
          - 'ペア名 ↓値ばりで記入' / '氏名 ↓値ばりで記入'
        """

        for c in df.columns:
            h = _norm_header(c)
            if not h:
                continue
            if any(k in h for k in include_any) and not any(k in h for k in exclude_any):
                return c
        return None

    col_pair = _pick_col(include_any=("ペア名", "ペア"), exclude_any=("相手",))
    col_members = _pick_col(include_any=("氏名", "選手名"), exclude_any=("相手",))
    col_level = _pick_col(include_any=("レベル",), exclude_any=("相手",))
    col_group = _pick_col(include_any=("グループ",), exclude_any=("相手",))

    if col_pair is None:
        raise ValueError(
            "チーム一覧の列 'ペア名' が見つかりません（例: 'ペア名', 'ペア名 ↓値ばりで記入'）。"
            f" 実際のヘッダー: {[ _norm_header(c) for c in list(df.columns)[:12] ]}"
        )
    if col_members is None:
        # Members can be empty, but the column should exist in our templates.
        # Still, allow running without it by treating as empty.
        col_members = ""

    def _infer_level(name: str) -> str:
        # Common naming: '上海A1' / '罗湖B8' / etc.
        # Prefer a level letter that appears right before trailing digits.
        import re

        s = str(name)
        m = re.search(r"([ABC])\d+$", s)
        if m:
            return m.group(1)
        for ch in s:
            if ch in "ABC":
                return ch
        return ""

    def _infer_group(name: str) -> str:
        # Common naming: strip trailing digits to get club/root.
        try:
            return str(name).rstrip("0123456789")
        except Exception:
            return ""

    # Collect raw preference cells first; we'll filter after we know all team names.
    raw_pref_by_team: dict[str, list[str]] = {}

    for _, row in df.iterrows():
        name = str(row.get(col_pair, "") or "").strip()
        if not name:
            continue
        members = str(row.get(col_members, "") or "").strip() if col_members else ""
        level = str(row.get(col_level, "") or "").strip() if col_level else ""
        group = str(row.get(col_group, "") or "").strip() if col_group else ""
        if not level:
            level = _infer_level(name)
        if not group:
            group = _infer_group(name)

        pref_candidates: List[str] = []
        pref_cols = [c for c in df.columns if isinstance(c, str) and any(k in c for k in ("優先", "希望", "対戦", "相手"))]
        for col in pref_cols:
            val = row.get(col)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            sval = str(val).strip()
            if not sval:
                continue
            pref_candidates.append(sval)
        raw_pref_by_team[name] = pref_candidates

        teams.append(
            Team(
                name=name,
                members=members,
                level=level,
                group=group,
                preferred_opponents=[],
            )
        )

    # Second pass: keep only values that look like actual opponent team names.
    team_names: set[str] = {t.name for t in teams}

    def _split_candidates(s: str) -> list[str]:
        # Split on common separators; keep order.
        import re

        raw = str(s).replace("\r", "\n")
        parts: list[str] = []
        for chunk in re.split(r"[\n,、/|]+", raw):
            v = chunk.strip()
            if v:
                parts.append(v)
        return parts

    def _extract_opponent_from_match_expr(expr: str, self_name: str) -> list[str]:
        # Accept strings like 'A1 vs B1', 'A1-B1', 'A1対B1'.
        import re

        s = expr
        for sep in (" vs ", " VS ", " Vs ", "-", "―", "−", "–", "—", "対", "ｖｓ", "vs"):
            if sep in s:
                left, right = s.split(sep, 1)
                left = left.strip()
                right = right.strip()
                candidates = [left, right]
                out: list[str] = []
                for c in candidates:
                    if c == self_name:
                        continue
                    if c in team_names:
                        out.append(c)
                return out
        # Fallback: if the string contains self name and another known name, pick the other.
        out: list[str] = []
        if self_name and self_name in s:
            for nm in team_names:
                if nm != self_name and nm in s:
                    out.append(nm)
        # De-dupe while preserving order.
        seen: set[str] = set()
        uniq: list[str] = []
        for x in out:
            if x not in seen:
                seen.add(x)
                uniq.append(x)
        return uniq

    def _looks_like_marker(s: str) -> bool:
        v = s.strip().lower()
        if not v:
            return True
        if v in {"○", "●", "◎", "◯", "x", "×", "yes", "y", "true", "t", "1", "0", "ok", "ng"}:
            return True
        if v.isdigit() and len(v) <= 2:
            return True
        return False

    by_name: dict[str, Team] = {t.name: t for t in teams}
    for t in teams:
        preferred: list[str] = []
        for raw in raw_pref_by_team.get(t.name, []):
            for token in _split_candidates(raw):
                if _looks_like_marker(token):
                    continue
                if token in team_names:
                    preferred.append(token)
                    continue
                preferred.extend(_extract_opponent_from_match_expr(token, t.name))

        # Normalize
        norm: list[str] = []
        seen: set[str] = set()
        for opp in preferred:
            if opp == t.name:
                continue
            if opp not in team_names:
                continue
            if opp in seen:
                continue
            seen.add(opp)
            norm.append(opp)
        by_name[t.name].preferred_opponents = norm
    return teams

###################################################################################################
# 旧スケジューラは 334～336 の揺らぎがあったため、以下でグラフ構築ベースへ再設計
# 要件: 同レベル / 同グループ禁止 / 希望ペアは必ず含める / 各ペアちょうど TARGET_MATCHES_PER_TEAM 試合
# （総試合数は expected_total_matches(len(teams), TARGET_MATCHES_PER_TEAM) を基本に、容量で上限が決まる）
# 方針:
# 1. レベル別に次数6の単純グラフを構築 (希望ペア固定 → 残りを貪欲 + バックオフ)
# 2. 全レベルのエッジ(試合)集合を統合
# 3. ラウンド割当: 早いラウンドから両ペア空き枠/コート空きに順次詰める (num_rounds*courts容量 > 必要試合数)
# 4. コート再配置で A左/B中/C右の帯寄せ
# 5. 分散最大化: エッジ選択時に「新グループ対戦」優先 + 複数シード試行
###################################################################################################

def build_level_graph(level_teams: List[Team], preferred_pairs: Set[frozenset], seed: int) -> List[Tuple[Team, Team]]:
    level_groups = sorted({t.group for t in level_teams})
    max_unique_groups = max(0, len(level_groups) - 1)
    soft_target = min(GROUP_SOFT_TARGET, max_unique_groups) if max_unique_groups else 0
    hard_target = min(GROUP_HARD_TARGET, max_unique_groups) if max_unique_groups else 0
    restarts = MAX_GROUP_GRAPH_RESTARTS if soft_target else 1

    best_edges: Optional[List[Tuple[Team, Team]]] = None
    best_min_diversity = -1
    last_error: Optional[Exception] = None

    for idx in range(restarts):
        attempt_seed = seed + idx * 997  # spread randomness per restart
        effective_soft = max(0, soft_target - idx // 2)
        try:
            edges, diversity_map = _construct_level_graph(level_teams, preferred_pairs, attempt_seed, effective_soft)
        except RuntimeError as err:
            last_error = err
            continue
        min_diversity = min(diversity_map.values()) if diversity_map else 0
        if hard_target == 0 or min_diversity >= hard_target:
            return edges
        if min_diversity > best_min_diversity:
            best_edges = edges
            best_min_diversity = min_diversity

    if best_edges is not None:
        return best_edges
    if last_error:
        raise last_error
    raise RuntimeError(f"レベル '{level_teams[0].level}' グラフ構築失敗: グループ分散要件を満たせませんでした")


def _construct_level_graph(level_teams: List[Team], preferred_pairs: Set[frozenset], seed: int,
                           soft_group_target: int) -> Tuple[List[Tuple[Team, Team]], Dict[str, int]]:
    random.seed(seed)
    required_degree = TARGET_MATCHES_PER_TEAM
    unique_groups = len({t.group for t in level_teams})
    unique_groups_available = max(0, unique_groups - 1)
    other_groups = max(1, unique_groups_available)
    tight_cap = math.ceil(required_degree / other_groups)
    repeat_cap = max(GROUP_REPEAT_CAP + 1, tight_cap)
    if level_teams and level_teams[0].level == 'C':
        repeat_cap = max(repeat_cap, GROUP_REPEAT_CAP + 2)
    pref_count_per_team: Dict[str, int] = {t.name: 0 for t in level_teams}
    for edge in preferred_pairs:
        a, b = list(edge)
        if a in pref_count_per_team:
            pref_count_per_team[a] += 1
        if b in pref_count_per_team:
            pref_count_per_team[b] += 1
    for t in level_teams:
        if pref_count_per_team[t.name] > required_degree:
            raise ValueError(
                f"希望ペア過多: {t.name} が {pref_count_per_team[t.name]} 件 (最大{required_degree})"
            )

    team_by_name = {t.name: t for t in level_teams}
    degree: Dict[str, int] = {t.name: 0 for t in level_teams}
    edges: List[Tuple[Team, Team]] = []
    existing_pairs: Set[frozenset] = set()
    groups_faced: Dict[str, Set[str]] = {t.name: set() for t in level_teams}
    group_counts: Dict[str, Dict[str, int]] = {t.name: defaultdict(int) for t in level_teams}

    def add_edge(t1: Team, t2: Team) -> None:
        edge_key = frozenset({t1.name, t2.name})
        edges.append((t1, t2))
        existing_pairs.add(edge_key)
        degree[t1.name] += 1
        degree[t2.name] += 1
        group_counts[t1.name][t2.group] += 1
        group_counts[t2.name][t1.group] += 1
        groups_faced[t1.name].add(t2.group)
        groups_faced[t2.name].add(t1.group)

    def remove_edge(entry: Tuple[Team, Team]) -> None:
        edges.remove(entry)
        key = frozenset({entry[0].name, entry[1].name})
        existing_pairs.remove(key)
        degree[entry[0].name] -= 1
        degree[entry[1].name] -= 1
        for src, dst in ((entry[0], entry[1]), (entry[1], entry[0])):
            group_counts[src.name][dst.group] -= 1
            if group_counts[src.name][dst.group] <= 0:
                del group_counts[src.name][dst.group]
                groups_faced[src.name].discard(dst.group)

    filtered_pref = [e for e in preferred_pairs if all(name in team_by_name for name in e)]
    for pair in sorted(filtered_pref, key=lambda x: tuple(sorted(list(x)))):
        a_name, b_name = list(pair)
        ta = team_by_name[a_name]
        tb = team_by_name[b_name]
        if ta.group == tb.group:
            continue
        if degree[ta.name] >= required_degree or degree[tb.name] >= required_degree:
            continue
        if group_counts[ta.name].get(tb.group, 0) >= repeat_cap:
            continue
        if group_counts[tb.name].get(ta.group, 0) >= repeat_cap:
            continue
        edge_key = frozenset({ta.name, tb.name})
        if edge_key in existing_pairs:
            continue
        add_edge(ta, tb)

    max_attempts = 4000
    attempts = 0
    order = level_teams[:]
    random.shuffle(order)

    while attempts < max_attempts:
        attempts += 1
        needers = [t for t in order if degree[t.name] < required_degree]
        if not needers:
            break
        needers.sort(key=lambda t: (required_degree - degree[t.name], -len(groups_faced[t.name])), reverse=True)
        progressed = False
        for t1 in needers:
            if degree[t1.name] >= required_degree:
                continue
            candidates = [t2 for t2 in level_teams
                          if t2.name != t1.name
                          and t2.group != t1.group
                          and degree[t2.name] < required_degree
                          and frozenset({t1.name, t2.name}) not in existing_pairs
                          and group_counts[t1.name].get(t2.group, 0) < repeat_cap
                          and group_counts[t2.name].get(t1.group, 0) < repeat_cap]
            if not candidates:
                continue
            needs_new_group = soft_group_target > 0 and len(groups_faced[t1.name]) < soft_group_target
            if needs_new_group:
                new_group_candidates = [t2 for t2 in candidates if t2.group not in groups_faced[t1.name]]
                if new_group_candidates:
                    candidates = new_group_candidates

            def cand_key(t2: Team):
                new_group_flag = 1 if t2.group not in groups_faced[t1.name] else 0
                t2_needs_new = 1 if (soft_group_target > 0 and len(groups_faced[t2.name]) < soft_group_target and t1.group not in groups_faced[t2.name]) else 0
                return (new_group_flag, t2_needs_new, required_degree - degree[t2.name], -len(groups_faced[t2.name]), random.random())

            candidates.sort(key=cand_key, reverse=True)
            t2 = candidates[0]
            add_edge(t1, t2)
            progressed = True
        if not progressed:
            removable = [e for e in edges if frozenset({e[0].name, e[1].name}) not in filtered_pref]
            if not removable:
                break
            rem = random.choice(removable)
            remove_edge(rem)

    if any(degree[t.name] != required_degree for t in level_teams):
        missing = [t.name for t in level_teams if degree[t.name] != required_degree]
        raise RuntimeError(f"レベル '{level_teams[0].level}' グラフ構築失敗: 次数未充足 {missing[:10]}")

    diversity_map = {t.name: len(groups_faced[t.name]) for t in level_teams}
    return edges, diversity_map

def build_all_level_graphs(teams: List[Team], seed: int) -> List[Tuple[Team, Team]]:
    # 希望ペア集合 (同レベル/異グループのみ有効)
    team_by_name = {t.name: t for t in teams}
    preferred_pairs: Set[frozenset] = set()
    for t in teams:
        for opp_name in t.preferred_opponents:
            o = team_by_name.get(opp_name)
            if not o:
                continue
            if o.level != t.level:
                continue
            if o.group == t.group:
                continue
            if o.name == t.name:
                continue
            preferred_pairs.add(frozenset({t.name, o.name}))
    level_map = {'A': [t for t in teams if t.level == 'A'],
                 'B': [t for t in teams if t.level == 'B'],
                 'C': [t for t in teams if t.level == 'C']}
    all_edges: List[Tuple[Team, Team]] = []
    for lvl in ['A','B','C']:
        lvl_pref = {p for p in preferred_pairs if list(p)[0] in {t.name for t in level_map[lvl]}}
        edges = build_level_graph(level_map[lvl], lvl_pref, seed + hash(lvl) % 1000)
        all_edges.extend(edges)
    return all_edges

def pack_rounds(edges: List[Tuple[Team, Team]], num_rounds: int, courts: int,
                shuffle_seed: Optional[int] = None) -> List[Match]:
    """単純早期割当アルゴリズム: 各エッジ(試合)を最も早いラウンドに配置する。
    条件: 1ラウンドで同一チームは1試合のみ / コート数上限遵守。
    容量: num_rounds*courts >= 必要試合数 なので失敗しない前提。
    """
    start_base = datetime(2025, 11, 26, 12, 50)
    # ラウンド別保持
    round_matches_raw: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds+1)}
    # 各ラウンドで使用済みチーム集合
    used_in_round: Dict[int, Set[str]] = {r: set() for r in range(1, num_rounds+1)}
    last_round_played: Dict[str, int] = defaultdict(int)
    played_rounds: Dict[str, Set[int]] = defaultdict(set)

    if shuffle_seed is None:
        edges_sorted = sorted(edges, key=lambda e: (e[0].level, e[0].name, e[1].name))
    else:
        edges_sorted = edges[:]
        random.Random(shuffle_seed).shuffle(edges_sorted)

    for (t1, t2) in edges_sorted:
        placed = False

        def _penalty_for(team_name: str, round_num: int) -> Tuple[int, int]:
            """Return (triple_flag, b2b_flag) if team plays at round_num."""

            rs = played_rounds.get(team_name, set())
            b2b = 1 if (round_num - 1) in rs else 0
            triple = 1 if b2b and (round_num - 2) in rs else 0
            return triple, b2b

        candidates: List[Tuple[Tuple[int, int, int, int, int], int]] = []
        for r in range(1, num_rounds + 1):
            if len(round_matches_raw[r]) >= courts:
                continue  # コート満杯
            if t1.name in used_in_round[r] or t2.name in used_in_round[r]:
                continue  # 同ラウンド重複不可

            p1 = _penalty_for(t1.name, r)
            p2 = _penalty_for(t2.name, r)
            triple = max(p1[0], p2[0])
            b2b = max(p1[1], p2[1])
            load = len(round_matches_raw[r])
            # Prefer: avoid triples > avoid back-to-backs > spread load > earlier rounds.
            key = (triple, b2b, load, r)
            # As a small tie-breaker, prefer not placing immediately after last round played.
            key = (key[0], key[1], key[2], key[3], abs(r - max(last_round_played[t1.name], last_round_played[t2.name])))
            candidates.append((key, r))

        if candidates:
            candidates.sort(key=lambda x: x[0])
            ordered_rounds = [r for _, r in candidates]
        else:
            ordered_rounds = []

        for r in ordered_rounds:
            start_time = start_base + timedelta(minutes=13*(r-1))
            m = Match(r, 0, t1, t2, start_time)
            round_matches_raw[r].append(m)
            used_in_round[r].add(t1.name)
            used_in_round[r].add(t2.name)
            t1.matches += 1
            t2.matches += 1
            t1.opponents.add(t2.name)
            t2.opponents.add(t1.name)
            t1.groups_faced.add(t2.group)
            t2.groups_faced.add(t1.group)
            last_round_played[t1.name] = r
            last_round_played[t2.name] = r
            played_rounds[t1.name].add(r)
            played_rounds[t2.name].add(r)
            t1.last_round = max(t1.last_round, r)
            t2.last_round = max(t2.last_round, r)
            placed = True
            break
        if not placed:
            # 詳細デバッグ情報
            usage_counts = {r: (t1.name in used_in_round[r], t2.name in used_in_round[r], len(round_matches_raw[r])) for r in range(1, num_rounds+1)}
            debug_lines = [f"R{r}:t1={'Y' if u[0] else 'N'} t2={'Y' if u[1] else 'N'} count={u[2]}" for r,u in usage_counts.items()]
            info = ", ".join(debug_lines[:10]) + (" ..." if len(debug_lines)>10 else "")
            raise RuntimeError(f"ラウンド分配失敗: {t1.name}-{t2.name} 未割当 / 状態 {info}")

    # 全試合収集
    matches = [m for r in range(1, num_rounds+1) for m in round_matches_raw[r]]

    # コート割当 (A左 / C右 / B中)
    for r in range(1, num_rounds+1):
        rm = round_matches_raw[r]
        if not rm:
            continue
        a_ms = [m for m in rm if m.team1.level == 'A']
        b_ms = [m for m in rm if m.team1.level == 'B']
        c_ms = [m for m in rm if m.team1.level == 'C']
        # 左からA, 右からC, 残りB
        a_targets = list(range(1, 1+len(a_ms)))
        c_targets = list(range(courts, courts - len(c_ms), -1))
        used = set(a_targets) | set(c_targets)
        b_targets = []
        for c in range(1, courts+1):
            if c in used:
                continue
            b_targets.append(c)
            if len(b_targets) >= len(b_ms):
                break
        for m, c in zip(a_ms, a_targets):
            m.court = c
        for m, c in zip(b_ms, b_targets):
            m.court = c
        for m, c in zip(c_ms, c_targets):
            m.court = c

    return sorted(matches, key=lambda m: (m.round_num, m.court))

def schedule_matches_graph(teams: List[Team], num_rounds: int, courts: int, seed: int = 0) -> List[Match]:
    # 初期化リセット (複数試行用)
    for t in teams:
        t.matches = 0
        t.opponents = set()
        t.groups_faced = set()
        t.last_round = 0
    edges = build_all_level_graphs(teams, seed)
    try:
        return pack_rounds(edges, num_rounds, courts)
    except RuntimeError as primary_error:
        # 再配置の順番を変えて最大3回まで再試行
        for retry in range(1, 4):
            try:
                return pack_rounds(edges, num_rounds, courts, shuffle_seed=seed + retry)
            except RuntimeError:
                continue
        raise primary_error

# ============================= ヒューリスティック フォールバック =============================
def schedule_matches_heuristic(teams: List[Team], num_rounds: int, courts: int, seed: int = 0) -> List[Match]:
    random.seed(seed)
    start_time = datetime(2025, 11, 26, 12, 50)
    round_duration = timedelta(minutes=13)
    matches: List[Match] = []

    # レベル別リスト
    level_team_map = {
        'A': [t for t in teams if t.level == 'A'],
        'B': [t for t in teams if t.level == 'B'],
        'C': [t for t in teams if t.level == 'C'],
    }
    level_group_map = {
        lvl: {t.group for t in level_team_map[lvl]}
        for lvl in ['A', 'B', 'C']
    }
    # 希望ペア集合（必ず入れる）
    team_by_name = {t.name: t for t in teams}
    repeat_cap_per_level: Dict[str, int] = {}
    required_degree = TARGET_MATCHES_PER_TEAM
    for lvl, lst in level_team_map.items():
        unique_groups = len({t.group for t in lst})
        other_groups = max(1, unique_groups - 1)
        # 各レベルで TARGET_MATCHES_PER_TEAM 試合を必達にするため、必要な繰り返し回数を動的に算出
        tight_cap = math.ceil(required_degree / other_groups)
        repeat_cap = max(GROUP_REPEAT_CAP + 1, tight_cap)
        # C はグループ人数の偏りが大きく、平均値ぎりぎりだと組み切れないことがあるため追加の余裕を与える
        if lvl == 'C':
            repeat_cap = max(repeat_cap, GROUP_REPEAT_CAP + 2)
        repeat_cap_per_level[lvl] = repeat_cap
    desired_pairs_set: Set[frozenset] = set()
    for t in teams:
        for opp_name in t.preferred_opponents:
            opp = team_by_name.get(opp_name)
            if not opp:
                continue
            if opp.level != t.level:
                continue
            if opp.group == t.group:
                continue
            if opp.name == t.name:
                continue
            desired_pairs_set.add(frozenset({t.name, opp.name}))

    # チーム状態初期化
    for t in teams:
        t.matches = 0
        t.opponents = set()
        t.groups_faced = set()
        t.last_round = 0

    # 各希望ペアの進捗
    scheduled_desired: Set[frozenset] = set()

    def rest_penalty(team: Team, current_round: int) -> int:
        if team.last_round == 0:
            return 0
        return 0 if current_round - team.last_round > 1 else 1

    def group_repeat_count(team: Team, target_group: str) -> int:
        return sum(1 for opp_name in team.opponents
                   if (opp := team_by_name.get(opp_name)) and opp.group == target_group)

    def allow_pair(t1: Team, t2: Team) -> bool:
        if t1.name == t2.name:
            return False
        if t1.level != t2.level:
            return False
        if t1.group == t2.group:
            return False
        if t2.name in t1.opponents:
            return False
        cap1 = repeat_cap_per_level[t1.level]
        cap2 = repeat_cap_per_level[t2.level]
        max_unique1 = max(0, len(level_group_map.get(t1.level, set())) - 1)
        max_unique2 = max(0, len(level_group_map.get(t2.level, set())) - 1)
        remaining_groups1 = max(0, max_unique1 - len(t1.groups_faced))
        remaining_groups2 = max(0, max_unique2 - len(t2.groups_faced))
        remaining_matches1 = max(0, TARGET_MATCHES_PER_TEAM - t1.matches)
        remaining_matches2 = max(0, TARGET_MATCHES_PER_TEAM - t2.matches)
        if (max_unique1 and len(t1.groups_faced) >= max_unique1) or remaining_matches1 > remaining_groups1:
            cap1 = None
        if (max_unique2 and len(t2.groups_faced) >= max_unique2) or remaining_matches2 > remaining_groups2:
            cap2 = None
        if cap1 is not None and group_repeat_count(t1, t2.group) >= cap1:
            return False
        if cap2 is not None and group_repeat_count(t2, t1.group) >= cap2:
            return False
        return True

    # ラウンドごとに希望ペアを優先配置 → 余りを通常ペア
    for rnd in range(1, num_rounds + 1):
        courts_free = list(range(1, courts + 1))
        used_names: Set[str] = set()
        round_matches: List[Match] = []

        # 希望ペア優先: レベル順 B→A→C (Bが最大で不足しやすいため)
        for lvl in ['B', 'A', 'C']:
            candidates = [p for p in desired_pairs_set if p not in scheduled_desired if all(team_by_name[n].level == lvl for n in p)]
            # 両チーム未使用 & 試合数 < TARGET_MATCHES_PER_TEAM & 未対戦
            for pair in sorted(candidates, key=lambda x: tuple(sorted(list(x)))):
                if not courts_free:
                    break
                a_name, b_name = list(pair)
                ta = team_by_name[a_name]
                tb = team_by_name[b_name]
                if ta.matches >= TARGET_MATCHES_PER_TEAM or tb.matches >= TARGET_MATCHES_PER_TEAM:
                    continue
                if ta.name in used_names or tb.name in used_names:
                    continue
                if tb.name in ta.opponents:
                    continue
                if not allow_pair(ta, tb):
                    continue
                court = courts_free.pop(0)
                m = Match(rnd, court, ta, tb, start_time + (rnd-1)*round_duration)
                round_matches.append(m)
                used_names.add(ta.name); used_names.add(tb.name)
                ta.matches += 1; tb.matches += 1
                ta.opponents.add(tb.name); tb.opponents.add(ta.name)
                ta.groups_faced.add(tb.group); tb.groups_faced.add(ta.group)
                ta.last_round = rnd; tb.last_round = rnd
                scheduled_desired.add(pair)
        # 通常ペア充填
        for lvl in ['B', 'A', 'C']:
            if not courts_free:
                break
            pool = [t for t in level_team_map[lvl] if t.matches < TARGET_MATCHES_PER_TEAM and t.name not in used_names]
            # ソート: 未試合数昇順
            pool.sort(key=lambda t: (rest_penalty(t, rnd), t.matches, t.name))
            i = 0
            while i < len(pool) and courts_free:
                t1 = pool[i]
                if t1.name in used_names or t1.matches >= TARGET_MATCHES_PER_TEAM:
                    i += 1
                    continue
                # パートナー候補
                partners = [t2 for t2 in pool[i+1:]
                            if t2.name not in used_names
                            and t2.matches < TARGET_MATCHES_PER_TEAM
                            and allow_pair(t1, t2)]
                if not partners:
                    i += 1
                    continue
                # 希望ペア優先 / 次に未対戦グループ / 未試合数
                def pk(t2: Team):
                    desired_flag = 0 if frozenset({t1.name, t2.name}) in desired_pairs_set else 1
                    new_group_flag = 0 if t2.group not in t1.groups_faced else 1
                    return (rest_penalty(t2, rnd), desired_flag, new_group_flag, t2.matches, t2.name)
                partners.sort(key=pk)
                t2 = partners[0]
                court = courts_free.pop(0)
                m = Match(rnd, court, t1, t2, start_time + (rnd-1)*round_duration)
                round_matches.append(m)
                used_names.add(t1.name); used_names.add(t2.name)
                t1.matches += 1; t2.matches += 1
                t1.opponents.add(t2.name); t2.opponents.add(t1.name)
                t1.groups_faced.add(t2.group); t2.groups_faced.add(t1.group)
                t1.last_round = rnd; t2.last_round = rnd
                i += 1

        # 残りコートを可能なら追加 (優先 B→A→C)
        for lvl in ['B','A','C']:
            if not courts_free:
                break
            pool = [t for t in level_team_map[lvl] if t.matches < TARGET_MATCHES_PER_TEAM and t.name not in used_names]
            pool.sort(key=lambda t: (rest_penalty(t, rnd), t.matches, t.name))
            for t1 in pool:
                if not courts_free:
                    break
                if t1.name in used_names or t1.matches >= TARGET_MATCHES_PER_TEAM:
                    continue
                partners = [t2 for t2 in level_team_map[lvl]
                            if t2.matches < TARGET_MATCHES_PER_TEAM
                            and t2.name not in used_names
                            and t2.name != t1.name
                            and allow_pair(t1, t2)]
                if not partners:
                    continue
                partners.sort(key=lambda t2: (rest_penalty(t2, rnd), t2.matches, t2.name))
                t2 = partners[0]
                court = courts_free.pop(0)
                m = Match(rnd, court, t1, t2, start_time + (rnd-1)*round_duration)
                round_matches.append(m)
                used_names.add(t1.name); used_names.add(t2.name)
                t1.matches += 1; t2.matches += 1
                t1.opponents.add(t2.name); t2.opponents.add(t1.name)
                t1.groups_faced.add(t2.group); t2.groups_faced.add(t1.group)
                t1.last_round = rnd; t2.last_round = rnd

        # コート帯寄せ再配置 (A左/C右/B中)
        a_ms = [m for m in round_matches if m.team1.level == 'A']
        b_ms = [m for m in round_matches if m.team1.level == 'B']
        c_ms = [m for m in round_matches if m.team1.level == 'C']
        a_targets = list(range(1, 1+len(a_ms)))
        c_targets = list(range(courts, courts - len(c_ms), -1))
        used_ct = set(a_targets) | set(c_targets)
        b_targets = []
        for c in range(1, courts+1):
            if c in used_ct:
                continue
            b_targets.append(c)
            if len(b_targets) >= len(b_ms):
                break
        for m,c in zip(a_ms,a_targets): m.court = c
        for m,c in zip(b_ms,b_targets): m.court = c
        for m,c in zip(c_ms,c_targets): m.court = c

        matches.extend(sorted(round_matches, key=lambda m: m.court))

    # 希望ペア強制挿入 (未達) - 既存試合の置換 / 追加
    missing_pairs = [p for p in desired_pairs_set if p not in scheduled_desired]
    for pair in missing_pairs:
        a_name, b_name = list(pair)
        ta = team_by_name[a_name]; tb = team_by_name[b_name]
        # 既に対戦済みならスキップ (理論上入っていないので不要)
        if tb.name in ta.opponents:
            continue
        # そのまま追加可能か (両者 < TARGET_MATCHES_PER_TEAM) → 空きラウンド検索
        added = False
        if ta.matches < TARGET_MATCHES_PER_TEAM and tb.matches < TARGET_MATCHES_PER_TEAM:
            for rnd in range(1, num_rounds+1):
                # ラウンド内使用チーム
                names_r = {m.team1.name for m in matches if m.round_num == rnd} | {m.team2.name for m in matches if m.round_num == rnd}
                if ta.name in names_r or tb.name in names_r:
                    continue
                used_courts = {m.court for m in matches if m.round_num == rnd}
                free_courts = [c for c in range(1, courts+1) if c not in used_courts]
                if not free_courts:
                    continue
                court = free_courts[0]
                m = Match(rnd, court, ta, tb, start_time + (rnd-1)*round_duration)
                matches.append(m)
                ta.matches += 1; tb.matches += 1
                ta.opponents.add(tb.name); tb.opponents.add(ta.name)
                ta.groups_faced.add(tb.group); tb.groups_faced.add(ta.group)
                added = True
                break
        if added:
            continue
        # 置換: ta または tb が TARGET_MATCHES_PER_TEAM の場合、当該チームの既存試合を破棄し希望ペアに置換
        # 適当なラウンドで ta/tb を含む試合 (非希望ペア) を探す
        def is_desired(m: Match) -> bool:
            return frozenset({m.team1.name, m.team2.name}) in desired_pairs_set
        replace_targets = [m for m in matches if (m.team1.name in (a_name,b_name) or m.team2.name in (a_name,b_name)) and not is_desired(m)]
        if not replace_targets:
            continue  # 置換余地なし
        victim = replace_targets[0]
        rnd = victim.round_num
        court = victim.court
        other = victim.team2 if victim.team1.name in (a_name,b_name) else victim.team1
        # victim 削除
        matches.remove(victim)
        victim.team1.matches -= 1; victim.team2.matches -= 1
        # 希望ペア追加
        m_new = Match(rnd, court, ta, tb, start_time + (rnd-1)*round_duration)
        matches.append(m_new)
        if tb.name not in ta.opponents:
            ta.opponents.add(tb.name); tb.opponents.add(ta.name)
            ta.groups_faced.add(tb.group); tb.groups_faced.add(ta.group)
        ta.matches += 1; tb.matches += 1
        # 減ったチーム other を後でリペア対象に
    # リペア: 試合数 < TARGET_MATCHES_PER_TEAM のチームへ追加
    target_total = expected_total_matches(len(teams), TARGET_MATCHES_PER_TEAM)
    def round_free_courts(r: int) -> List[int]:
        used = {m.court for m in matches if m.round_num == r}
        return [c for c in range(1, courts+1) if c not in used]
    iteration = 0
    while iteration < 200:
        iteration += 1
        under = [t for t in teams if t.matches < TARGET_MATCHES_PER_TEAM]
        if not under:
            break
        progressed = False
        for t1 in sorted(under, key=lambda x: (x.matches, x.name)):
            if t1.matches >= TARGET_MATCHES_PER_TEAM:
                continue
            partners = [t2 for t2 in teams if t2.level == t1.level and allow_pair(t1, t2) and t2.matches < TARGET_MATCHES_PER_TEAM]
            if not partners:
                continue
            partners.sort(key=lambda t2: (t2.matches, t2.name))
            t2 = partners[0]
            # 空きラウンド
            for r in range(1, num_rounds+1):
                names_r = {m.team1.name for m in matches if m.round_num == r} | {m.team2.name for m in matches if m.round_num == r}
                if t1.name in names_r or t2.name in names_r:
                    continue
                free = round_free_courts(r)
                if not free:
                    continue
                court = free[0]
                m = Match(r, court, t1, t2, start_time + (r-1)*round_duration)
                matches.append(m)
                t1.matches += 1; t2.matches += 1
                t1.opponents.add(t2.name); t2.opponents.add(t1.name)
                t1.groups_faced.add(t2.group); t2.groups_faced.add(t1.group)
                progressed = True
                break
        if not progressed:
            break
    # 最終チェック & 整理
    matches = sorted(matches, key=lambda m: (m.round_num, m.court))
    # ================= 追加再配置フェーズ finalize_fill =================
    team_by_name = {t.name: t for t in teams}
    desired_pairs_set: Set[frozenset] = set()
    for t in teams:
        for opp_name in t.preferred_opponents:
            if opp_name in team_by_name:
                o = team_by_name[opp_name]
                if o.level == t.level and o.group != t.group and o.name != t.name:
                    desired_pairs_set.add(frozenset({t.name, o.name}))

    def rebuild_round_state() -> Tuple[Dict[int, Set[str]], Dict[int, Set[int]], Dict[int, List[Match]]]:
        names = {r: set() for r in range(1, num_rounds+1)}
        courts_used = {r: set() for r in range(1, num_rounds+1)}
        match_map = {r: [] for r in range(1, num_rounds+1)}
        for m in matches:
            names[m.round_num].add(m.team1.name)
            names[m.round_num].add(m.team2.name)
            courts_used[m.round_num].add(m.court)
            match_map[m.round_num].append(m)
        return names, courts_used, match_map

    def first_free_court(used: Set[int]) -> Optional[int]:
        for c in range(1, courts+1):
            if c not in used:
                return c
        return None

    def add_match(t1: Team, t2: Team, round_num: int, court: int) -> None:
        start = start_time + (round_num-1) * round_duration
        matches.append(Match(round_num, court, t1, t2, start))
        refresh_team_stats(teams, matches)

    def try_place_pair(t1: Team, t2: Team) -> bool:
        if not allow_pair(t1, t2):
            return False
        if t2.name in t1.opponents:
            return True
        names, courts_used, match_map = rebuild_round_state()
        # direct placement where both teams are free and a court is free
        for rnd in range(1, num_rounds+1):
            if t1.name in names[rnd] or t2.name in names[rnd]:
                continue
            court = first_free_court(courts_used[rnd])
            if court is None:
                continue
            add_match(t1, t2, rnd, court)
            return True
        # displacement: move a saturated match to open a slot
        names, courts_used, match_map = rebuild_round_state()
        for rnd in range(1, num_rounds+1):
            if t1.name in names[rnd] or t2.name in names[rnd]:
                continue
            if len(match_map[rnd]) < courts:
                continue
            victims = sorted(match_map[rnd], key=lambda m: (m.team1.matches + m.team2.matches, m.court))
            for victim in victims:
                for target in range(1, num_rounds+1):
                    if target == rnd:
                        continue
                    if victim.team1.name in names[target] or victim.team2.name in names[target]:
                        continue
                    court2 = first_free_court(courts_used[target])
                    if court2 is None:
                        continue
                    freed_court = victim.court
                    victim.round_num = target
                    victim.court = court2
                    victim.start_time = start_time + (target-1) * round_duration
                    add_match(t1, t2, rnd, freed_court)
                    return True
        return False

    def release_for_team(target: Team) -> bool:
        donors = [t for t in teams if t.level == target.level and t.name != target.name and allow_pair(target, t)]
        donors.sort(key=lambda t: (t.matches, len(t.groups_faced), t.name))
        for donor in donors:
            donor_matches = [m for m in matches
                             if (m.team1 == donor or m.team2 == donor)
                             and frozenset({m.team1.name, m.team2.name}) not in desired_pairs_set]
            donor_matches.sort(key=lambda m: (-m.round_num, m.court))
            for victim in donor_matches:
                other = victim.team2 if victim.team1 == donor else victim.team1
                if other.name == target.name:
                    continue
                try:
                    matches.remove(victim)
                except ValueError:
                    continue
                refresh_team_stats(teams, matches)
                if try_place_pair(target, donor):
                    return True
                matches.append(victim)
                refresh_team_stats(teams, matches)
        return False

    # 未達希望ペアの挿入
    remaining_pairs = [pair for pair in desired_pairs_set]
    for pair in remaining_pairs:
        a_name, b_name = list(pair)
        ta = team_by_name.get(a_name)
        tb = team_by_name.get(b_name)
        if not ta or not tb:
            continue
        if tb.name in ta.opponents:
            continue
        try_place_pair(ta, tb)

    # 未達チームの補填
    max_repairs = 300
    repair_iter = 0
    while repair_iter < max_repairs:
        refresh_team_stats(teams, matches)
        under = [t for t in teams if t.matches < TARGET_MATCHES_PER_TEAM]
        if not under:
            break
        under.sort(key=lambda t: (t.matches, len(t.groups_faced), t.level, t.name))
        progressed = False
        for t1 in under:
            partners = [t2 for t2 in teams if t2.level == t1.level and t2.matches < TARGET_MATCHES_PER_TEAM and allow_pair(t1, t2)]
            partners.sort(key=lambda t: (t.matches, len(t.groups_faced), t.name))
            for t2 in partners:
                if try_place_pair(t1, t2):
                    progressed = True
                    break
            if progressed:
                break
        if not progressed:
            for t1 in under:
                if release_for_team(t1):
                    progressed = True
                    break
        if not progressed:
            break
        repair_iter += 1

    matches = sorted(matches, key=lambda m: (m.round_num, m.court))

    # ================= 最終コート帯寄せ再調整 =================
    def band_reassign(all_matches: List[Match]):
        # 動的に B を中央寄せの連続ブロックへ、A 左 / C 右。
        for r in range(1, num_rounds+1):
            rm = [m for m in all_matches if m.round_num == r]
            if not rm:
                continue
            a_ms = [m for m in rm if m.team1.level == 'A']
            b_ms = [m for m in rm if m.team1.level == 'B']
            c_ms = [m for m in rm if m.team1.level == 'C']
            # 既存サイズ
            nA, nB, nC = len(a_ms), len(b_ms), len(c_ms)
            # 左側 A, 右側 C 割当
            a_targets = ([1, 2, 3, 4] * ((nA + 3) // 4))[:nA]
            c_targets = ([15, 14, 13] * ((nC + 2) // 3))[:nC]
            used = set(a_targets) | set(c_targets)
            # B の中央ブロック探索
            b_targets = []
            if nB > 0:
                center = (courts + 1)/2
                # 可能な連続ブロック候補生成 (重複除去済み)
                candidates = []
                for start in range(1, courts - nB + 2):
                    block = list(range(start, start + nB))
                    if all(c not in used for c in block):
                        b_center = (block[0] + block[-1])/2
                        score = abs(b_center - center)
                        candidates.append((score, block))
                if candidates:
                    candidates.sort(key=lambda x: x[0])
                    b_targets = candidates[0][1]
                else:
                    # フォールバック: 未使用コート昇順
                    free = [c for c in range(1, courts+1) if c not in used]
                    b_targets = free[:nB]
            # コート再割当
            for m,c in zip(a_ms, a_targets):
                m.court = c
            for m,c in zip(b_ms, b_targets):
                m.court = c
            for m,c in zip(c_ms, c_targets):
                m.court = c
        return all_matches

    matches = band_reassign(matches)
    refresh_team_stats(teams, matches)
    return sorted(matches, key=lambda m: (m.round_num, m.court))

    # (以下の旧コードは到達不能のため削除済み)

    # Preferred court ranges per level (soft preference)
    preferred_courts = {
        'A': list(range(1, 4)),    # 1-3 main
        'B': list(range(5, 14)),   # 5-13 main (middle)
        'C': list(range(13, 16)),  # 13-15 main
    }
    # Allowed courts relaxed: permit any court for any level (adjacency becomes best-effort)
    allowed_courts = {
        'A': list(range(1, courts + 1)),
        'B': list(range(1, courts + 1)),
        'C': list(range(1, courts + 1)),
    }

    # Target TARGET_MATCHES_PER_TEAM matches per team -> desired totals per level
    level_team_map = {
        'A': [t for t in teams if t.level == 'A'],
        'B': [t for t in teams if t.level == 'B'],
        'C': [t for t in teams if t.level == 'C'],
    }
    desired_total_matches = {lvl: expected_total_matches(len(level_team_map[lvl]), TARGET_MATCHES_PER_TEAM) for lvl in ['A', 'B', 'C']}

    remaining_desired = desired_total_matches.copy()
    # Build desired pairs set (undirected) from preferred_opponents
    team_by_name = {t.name: t for t in teams}
    desired_pairs_set: Set[frozenset] = set()
    for t in teams:
        for opp_name in (t.preferred_opponents or []):
            opp = team_by_name.get(opp_name)
            if not opp:
                continue
            if opp.level != t.level:
                continue
            if opp.group == t.group:
                continue
            if opp.name == t.name:
                continue
            desired_pairs_set.add(frozenset({t.name, opp.name}))

    def group_repeat_count(team: Team, target_group: str) -> int:
        return sum(1 for opp_name in team.opponents
                   if (opp := team_by_name.get(opp_name)) and opp.group == target_group)

    def prioritize_diversity(t1: Team, candidates: List[Team]) -> List[Team]:
        filtered = [t2 for t2 in candidates
                    if group_repeat_count(t1, t2.group) < GROUP_REPEAT_CAP
                    and group_repeat_count(t2, t1.group) < GROUP_REPEAT_CAP]
        return filtered if filtered else candidates

    # Helper to form pairs greedily prioritizing teams with fewer matches
    def form_pairs(level: str, max_pairs: int, used_names: Set[str], current_round: int) -> List[tuple[Team, Team]]:
        # Allow consecutive if必要（優先度は下げるが禁止しない）
        candidates = [t for t in level_team_map[level] if t.matches < TARGET_MATCHES_PER_TEAM and t.name not in used_names]
        if len(candidates) < 2 or max_pairs <= 0:
            return []
        # Deterministic ordering (base). For diversity attempts, incorporate seeded permutation for tie-breaks.
        if diversity_mode:
            # seeded stable pseudo-random permutation via hash
            candidates.sort(key=lambda t: (t.matches, (hash(t.name) + seed) % 10_000))
        else:
            candidates.sort(key=lambda t: (t.matches, t.name))

        pairs: List[tuple[Team, Team]] = []
        taken: Set[str] = set()
        for i, t1 in enumerate(candidates):
            if t1.name in taken:
                continue
            # find partner
            # Only cross-group partners（同グループは禁止）
            partners = [t2 for t2 in candidates[i+1:] if t2.name not in taken and t1.group != t2.group and t2.name not in t1.opponents]
            partners = prioritize_diversity(t1, partners)
            # Rank partners: 分散を強く優先（新グループ最優先）→ 希望ペア → 未消化
            def partner_key(t2: Team):
                desired_flag = 0 if frozenset({t1.name, t2.name}) in desired_pairs_set else 1
                # In diversity mode, prefer partner offering a new group for t1
                new_group_flag = 0 if t2.group not in t1.groups_faced else 1
                repeat_penalty = group_repeat_count(t1, t2.group)
                if diversity_mode:
                    return (new_group_flag, repeat_penalty, desired_flag, t2.matches, (hash(t2.name) + seed) % 10_000)
                return (repeat_penalty, desired_flag, t2.matches, t2.name)
            partners.sort(key=partner_key)
            if not partners:
                continue
            t2 = partners[0]
            pairs.append((t1, t2))
            taken.add(t1.name)
            taken.add(t2.name)
            if len(pairs) >= max_pairs:
                break
        return pairs

    # remember last contiguous block per level to keep placement stable
    last_block: dict[str, List[int] | None] = { 'A': None, 'B': None, 'C': None }

    def choose_block(level: str, k: int, avail_courts: List[int]) -> List[int]:
        if k <= 0:
            return []
        avail = sorted([c for c in avail_courts if c in allowed_courts[level]])
        if not avail:
            return []
        # preferred center as average of preferred courts; also consider last block center if exists
        pref_center = sum(preferred_courts[level]) / len(preferred_courts[level])
        last = last_block.get(level)
        last_center = (sum(last)/len(last)) if last else pref_center
        target_center = 0.6 * last_center + 0.4 * pref_center
        # try to find contiguous window of size k; if not possible, shrink k
        avail_set = set(avail)
        def best_window(size: int) -> List[int]:
            best = None
            # generate all contiguous windows within allowed that are fully available
            min_c = min(avail)
            max_c = max(avail)
            for start in range(min_c, max_c - size + 2):
                window = list(range(start, start + size))
                if all(c in avail_set for c in window):
                    center = (window[0] + window[-1]) / 2
                    score = abs(center - target_center)
                    if (best is None) or (score < best[0]):
                        best = (score, window)
            return best[1] if best else []
        size = k
        while size > 0:
            win = best_window(size)
            if win:
                return win
            size -= 1
        return []

    for round_num in range(1, num_rounds + 1):
        courts_this_round = list(range(1, courts + 1))
        used_names: Set[str] = set()

        rounds_left = num_rounds - round_num + 1
        # Initial quotas per level based on remaining desired, capped by allowed courts free to keep adjacency
        quotas = {}
        for lvl in ['A', 'B', 'C']:
            per_round_need = max(0, (remaining_desired[lvl] + rounds_left - 1) // rounds_left)  # ceil division
            allowed_free = len([c for c in allowed_courts[lvl] if c in courts_this_round])
            quotas[lvl] = min(per_round_need, allowed_free)

        # Ensure sum quotas <= courts; if not, reduce largest first
        while sum(quotas.values()) > len(courts_this_round):
            # reduce from level with largest quota and smallest remaining need ratio
            lvl_to_reduce = max(quotas, key=lambda l: (quotas[l], remaining_desired[l]))
            quotas[lvl_to_reduce] -= 1

        # Step 1: select contiguous court blocks per level and schedule into them
        round_matches: List[Match] = []
        # decide order by remaining need so larger levels place blocks first
        order = sorted(['A', 'B', 'C'], key=lambda l: (-quotas[l], l))
        blocks: dict[str, List[int]] = { 'A': [], 'B': [], 'C': [] }
        for lvl in order:
            k = min(quotas[lvl], len(courts_this_round))
            if k <= 0:
                continue
            # form candidate pairs first (so we don't reserve unused courts)
            pairs = form_pairs(lvl, k, used_names, round_num)
            if not pairs:
                continue
            # choose a contiguous block exactly for the number of pairs
            block = choose_block(lvl, len(pairs), courts_this_round)
            if not block or len(block) < len(pairs):
                # fallback: pick allowed courts up to pairs count
                allowed_free = sorted([c for c in allowed_courts[lvl] if c in courts_this_round])
                block = allowed_free[:len(pairs)]
            if not block:
                continue
            blocks[lvl] = block
            # reserve courts
            for c in block:
                if c in courts_this_round:
                    courts_this_round.remove(c)
            # place pairs into chosen block (keep adjacency)
            for (t1, t2), court in zip(pairs, block):
                match = Match(round_num, court, t1, t2, start_time + (round_num-1) * round_duration)
                round_matches.append(match)
                used_names.add(t1.name)
                used_names.add(t2.name)
                t1.matches += 1
                t2.matches += 1
                t1.opponents.add(t2.name)
                t2.opponents.add(t1.name)
                t1.last_round = round_num
                t2.last_round = round_num
                t1.groups_faced.add(t2.group)
                t2.groups_faced.add(t1.group)
                remaining_desired[lvl] = max(0, remaining_desired[lvl] - 1)
            if blocks[lvl]:
                last_block[lvl] = blocks[lvl]

        # Step 2: fill remaining courts expanding blocks contiguously where possible
        levels_by_need = sorted(['A', 'B', 'C'], key=lambda l: (-remaining_desired[l], l))
        for lvl in levels_by_need:
            if not courts_this_round:
                break
            # maximum additional pairs equals remaining courts
            extra_pairs = form_pairs(lvl, len(courts_this_round), used_names, round_num)
            for t1, t2 in extra_pairs:
                if not courts_this_round:
                    break
                # try to attach to existing block by nearest adjacency
                allowed_free_list = sorted([c for c in allowed_courts[lvl] if c in courts_this_round])
                if not allowed_free_list:
                    continue
                if last_block[lvl]:
                    # choose allowed court with minimal distance to current block
                    block = last_block[lvl]
                    center = (block[0] + block[-1]) / 2
                    court = min(allowed_free_list, key=lambda c: abs(c - center))
                    # also extend block record
                    block.append(court)
                    block.sort()
                    last_block[lvl] = block
                else:
                    court = allowed_free_list[0]
                match = Match(round_num, court, t1, t2, start_time + (round_num-1) * round_duration)
                round_matches.append(match)
                used_names.add(t1.name)
                used_names.add(t2.name)
                courts_this_round.remove(court)
                t1.matches += 1
                t2.matches += 1
                t1.opponents.add(t2.name)
                t2.opponents.add(t1.name)
                t1.last_round = round_num
                t2.last_round = round_num
                t1.groups_faced.add(t2.group)
                t2.groups_faced.add(t1.group)
                remaining_desired[lvl] = max(0, remaining_desired[lvl] - 1)

        # Reassign courts to cluster by level per round while keeping within allowed ranges
        # Build level buckets for this round
        a_ms = [m for m in round_matches if m.team1.level == 'A']
        b_ms = [m for m in round_matches if m.team1.level == 'B']
        c_ms = [m for m in round_matches if m.team1.level == 'C']
        # Compose target court sequence: A from left, C from right, B fills middle
        used = set()
        # assign A
        a_targets = []
        for c in range(1, courts + 1):
            if c in allowed_courts['A'] and c not in used:
                a_targets.append(c)
            if len(a_targets) >= len(a_ms):
                break
        # assign C from rightmost down
        c_targets = []
        for c in range(courts, 0, -1):
            if c in allowed_courts['C'] and c not in used and c not in set(a_targets):
                c_targets.append(c)
            if len(c_targets) >= len(c_ms):
                break
        # assign B in the middle preference
        b_targets = []
        for c in range(1, courts + 1):
            if c in allowed_courts['B'] and c not in set(a_targets) and c not in set(c_targets):
                b_targets.append(c)
            if len(b_targets) >= len(b_ms):
                break
        # sort targets to be increasing for zip stability except C which can be any order
        a_targets.sort()
        b_targets.sort()
        # reassign courts
        for m, c in zip(a_ms, a_targets):
            m.court = c
        for m, c in zip(b_ms, b_targets):
            m.court = c
        for m, c in zip(c_ms, c_targets):
            m.court = c
        # Append scheduled matches for this round
        matches.extend(sorted(round_matches, key=lambda m: m.court))

    # Post-fill pass: ensure total matches reaches target (each team up to TARGET_MATCHES_PER_TEAM matches)
    target_total = expected_total_matches(len(teams), TARGET_MATCHES_PER_TEAM)
    if target_total > num_rounds * courts:
        target_total = num_rounds * courts

    def round_free_courts(rn: int) -> List[int]:
        used = {m.court for m in matches if m.round_num == rn}
        return [c for c in range(1, courts + 1) if c not in used]

    def try_add_one(rn: int) -> bool:
        free = round_free_courts(rn)
        if not free:
            return False
        # levels ordered by remaining need
        needs = {lvl: sum(1 for t in level_team_map[lvl] if t.matches < TARGET_MATCHES_PER_TEAM) for lvl in ['A', 'B', 'C']}
        for lvl in sorted(['A', 'B', 'C'], key=lambda l: -needs[l]):
            # candidate teams not playing this round and with < TARGET_MATCHES_PER_TEAM matches
            names_in_rn = {m.team1.name for m in matches if m.round_num == rn} | {m.team2.name for m in matches if m.round_num == rn}
            cand = [t for t in level_team_map[lvl] if t.matches < TARGET_MATCHES_PER_TEAM and t.name not in names_in_rn]
            if len(cand) < 2:
                continue
            # pick best pair prioritizing desired first, then underplayed (cross-group only)
            best = None
            for i in range(len(cand)):
                t1 = cand[i]
                for j in range(i+1, len(cand)):
                    t2 = cand[j]
                    if t1.group == t2.group:
                        continue
                    if t2.name in t1.opponents:
                        continue
                    # scoring: 強く分散優先（新グループを最優先）→ 希望ペア → 未消化（連戦は許容）
                    desired_flag = 0 if frozenset({t1.name, t2.name}) in desired_pairs_set else 1
                    new_group_flag = 0 if (t2.group not in t1.groups_faced and t1.group not in t2.groups_faced) else 1
                    total_matches = t1.matches + t2.matches
                    # Seeded deterministic tie-breaker to allow multi-attempt variation
                    tie_val = (hash(t1.name + t2.name) + seed) % 10_000
                    if diversity_mode:
                        cand_tuple = (new_group_flag, desired_flag, total_matches, tie_val, t1, t2)
                    else:
                        cand_tuple = (desired_flag, total_matches, tie_val, t1, t2)
                    if best is None or cand_tuple < best:
                        best = cand_tuple
            if best is None:
                # Attempt a single swap move to free a candidate and create a valid pair
                needers = sorted([t for t in level_team_map[lvl] if t.matches < TARGET_MATCHES_PER_TEAM and t.name not in names_in_rn], key=lambda x: x.matches)
                swapped = False
                for t1 in needers:
                    # find a partner t2 that is busy this round but valid otherwise
                    for t2 in level_team_map[lvl]:
                        if t2.name in names_in_rn and t2.matches < TARGET_MATCHES_PER_TEAM and t1.name != t2.name and t1.group != t2.group and (t2.name not in t1.opponents):
                            # find t2's match in this round
                            bm = next((m for m in matches if m.round_num == rn and (m.team1.name == t2.name or m.team2.name == t2.name)), None)
                            if not bm:
                                continue
                            t3 = bm.team2 if bm.team1.name == t2.name else bm.team1
                            # find another round r2 where both t2 and t3 are free
                            for r2 in range(1, num_rounds + 1):
                                if r2 == rn:
                                    continue
                                names_in_r2 = {m.team1.name for m in matches if m.round_num == r2} | {m.team2.name for m in matches if m.round_num == r2}
                                if t2.name in names_in_r2 or t3.name in names_in_r2:
                                    continue
                                free_r2 = round_free_courts(r2)
                                if not free_r2:
                                    continue

                                # choose court for moved match (prefer allowed, else any)
                                allowed_free_r2 = [c for c in allowed_courts[lvl] if c in free_r2]
                                move_court = (sorted(allowed_free_r2)[0] if allowed_free_r2 else sorted(free_r2)[0])
                                # perform move
                                bm.round_num = r2
                                bm.start_time = start_time + timedelta(minutes=13*(r2-1))
                                bm.court = move_court
                                # now schedule t1 vs t2 in rn
                                free_now = round_free_courts(rn)
                                if not free_now:
                                    break
                                allowed_free_now = [c for c in allowed_courts[lvl] if c in free_now]
                                court = (sorted(allowed_free_now)[0] if allowed_free_now else sorted(free_now)[0])
                                match = Match(rn, court, t1, t2, start_time + (rn-1) * round_duration)
                                matches.append(match)
                                t1.matches += 1
                                t2.matches += 1
                                t1.opponents.add(t2.name)
                                t2.opponents.add(t1.name)
                                t1.last_round = rn
                                t2.last_round = rn
                                t1.groups_faced.add(t2.group)
                                t2.groups_faced.add(t1.group)
                                swapped = True
                                break
                            if swapped:
                                break
                    if swapped:
                        break
                if swapped:
                    return True
                else:
                    continue
            # best tuple shape depends on mode; extract last two entries as teams
            t1 = best[-2]
            t2 = best[-1]
            # choose court within allowed for level; prefer preferred, then closest to last block
            allowed_free = [c for c in allowed_courts[lvl] if c in free]
            # if no allowed free court, fall back to any free court to preserve total match count priority
            any_free = list(free)
            if not allowed_free and not any_free:
                continue
            pref_free = [c for c in preferred_courts[lvl] if c in allowed_free]
            if pref_free:
                court = sorted(pref_free)[0]
            else:
                if allowed_free:
                    if last_block[lvl]:
                        center = (last_block[lvl][0] + last_block[lvl][-1]) / 2
                        court = min(allowed_free, key=lambda c: abs(c - center))
                    else:
                        court = sorted(allowed_free)[0]
                else:
                    # no allowed court available; pick any free court
                    court = sorted(any_free)[0]
            match = Match(rn, court, t1, t2, start_time + (rn-1) * round_duration)
            matches.append(match)
            t1.matches += 1
            t2.matches += 1
            t1.opponents.add(t2.name)
            t2.opponents.add(t1.name)
            t1.last_round = rn
            t2.last_round = rn
            t1.groups_faced.add(t2.group)
            t2.groups_faced.add(t1.group)
            return True
        return False

    # Multi-pass repair loop to push to full 336 if still short
    for _ in range(3):  # limited passes for performance
        rn = 1
        any_change = False
        while len(matches) < target_total and rn <= num_rounds:
            added_any = False
            while len(matches) < target_total and try_add_one(rn):
                added_any = True
                any_change = True
            if added_any:
                r_matches = [m for m in matches if m.round_num == rn]
                a_ms = [m for m in r_matches if m.team1.level == 'A']
                b_ms = [m for m in r_matches if m.team1.level == 'B']
                c_ms = [m for m in r_matches if m.team1.level == 'C']
                used = set()
                a_targets = [c for c in range(1, courts+1) if c in allowed_courts['A'] and c not in used][:len(a_ms)]
                used.update(a_targets)
                c_targets = []
                for c in range(courts, 0, -1):
                    if c in allowed_courts['C'] and c not in used:
                        c_targets.append(c)
                    if len(c_targets) >= len(c_ms):
                        break
                used.update(c_targets)
                b_targets = [c for c in range(1, courts+1) if c in allowed_courts['B'] and c not in used][:len(b_ms)]
                for m, c in zip(sorted(a_ms, key=lambda x: x.court), sorted(a_targets)):
                    m.court = c
                for m, c in zip(sorted(b_ms, key=lambda x: x.court), sorted(b_targets)):
                    m.court = c
                for m, c in zip(sorted(c_ms, key=lambda x: -x.court), c_targets):
                    m.court = c
            rn += 1
        if len(matches) >= target_total or not any_change:
            break

    # Final aggressive repair to force all teams to TARGET_MATCHES_PER_TEAM matches if capacity allows
    def aggressive_repair():
        # helper lookups refreshed each iteration
        round_team_names = {r: {m.team1.name for m in matches if m.round_num == r} | {m.team2.name for m in matches if m.round_num == r} for r in range(1, num_rounds+1)}
        def free_courts(r):
            used = {m.court for m in matches if m.round_num == r}
            return [c for c in range(1, courts+1) if c not in used]
        changed = False
        underplayed = [t for t in teams if t.matches < TARGET_MATCHES_PER_TEAM]
        # Sort by matches then name for deterministic order
        underplayed.sort(key=lambda x: (x.matches, x.name))
        for t1 in underplayed:
            if t1.matches >= TARGET_MATCHES_PER_TEAM:
                continue
            # attempt partner
            partners = [t2 for t2 in teams if t2.level == t1.level and allow_pair(t1, t2) and t2.matches < TARGET_MATCHES_PER_TEAM]
            partners.sort(key=lambda x: (x.matches, x.name))
            for t2 in partners:
                if t1.matches >= TARGET_MATCHES_PER_TEAM:
                    break
                if t2.matches >= TARGET_MATCHES_PER_TEAM:
                    continue
                # find round with both free and free court
                scheduled = False
                for r in range(1, num_rounds+1):
                    if t1.name in round_team_names[r] or t2.name in round_team_names[r]:
                        continue
                    fc = free_courts(r)
                    if fc:
                        court = fc[0]
                        start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r-1))
                        m = Match(r, court, t1, t2, start_time)
                        matches.append(m)
                        t1.matches += 1
                        t2.matches += 1
                        t1.opponents.add(t2.name)
                        t2.opponents.add(t1.name)
                        t1.groups_faced.add(t2.group)
                        t2.groups_faced.add(t1.group)
                        round_team_names[r].add(t1.name)
                        round_team_names[r].add(t2.name)
                        changed = True
                        scheduled = True
                        break
                if scheduled:
                    continue
                # No direct slot: try displacement
                for r in range(1, num_rounds+1):
                    if t1.name in round_team_names[r] or t2.name in round_team_names[r]:
                        continue
                    # pick a movable saturated match in round r
                    movable = [m for m in matches if m.round_num == r and m.team1.level == t1.level and m.team2.level == t1.level and m.team1.matches >= TARGET_MATCHES_PER_TEAM and m.team2.matches >= TARGET_MATCHES_PER_TEAM]
                    if not movable:
                        continue
                    target_match = movable[0]
                    # find alternative round r2 to relocate target_match
                    for r2 in range(1, num_rounds+1):
                        if r2 == r:
                            continue
                        if target_match.team1.name in round_team_names[r2] or target_match.team2.name in round_team_names[r2]:
                            continue
                        fc2 = free_courts(r2)
                        if not fc2:
                            continue
                        # move
                        new_court = fc2[0]
                        target_match.round_num = r2
                        target_match.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r2-1))
                        target_match.court = new_court
                        round_team_names[r2].add(target_match.team1.name)
                        round_team_names[r2].add(target_match.team2.name)
                        round_team_names[r].remove(target_match.team1.name)
                        round_team_names[r].remove(target_match.team2.name)
                        # now schedule desired t1 vs t2 in freed slot r
                        fc_r = free_courts(r)
                        if not fc_r:
                            # should have at least one freed
                            continue
                        court_new = fc_r[0]
                        start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r-1))
                        m2 = Match(r, court_new, t1, t2, start_time)
                        matches.append(m2)
                        t1.matches += 1
                        t2.matches += 1
                        t1.opponents.add(t2.name)
                        t2.opponents.add(t1.name)
                        t1.groups_faced.add(t2.group)
                        t2.groups_faced.add(t1.group)
                        round_team_names[r].add(t1.name)
                        round_team_names[r].add(t2.name)
                        changed = True
                        scheduled = True
                        break
                    if scheduled:
                        break
            # end partner loop
        return changed

    # run aggressive repair iterations until stable or all done
    iter_guard = 0
    while any(t.matches < TARGET_MATCHES_PER_TEAM for t in teams) and iter_guard < 30:
        if not aggressive_repair():
            break
        iter_guard += 1

    # Final prune: if any accidental overflow (shouldn't happen) cut extras from last rounds deterministically
    for t in teams:
        if t.matches > TARGET_MATCHES_PER_TEAM:
            # remove surplus matches for that team (rare) preferring those against highest matched opponents
            surplus = t.matches - TARGET_MATCHES_PER_TEAM
            related = [m for m in matches if m.team1.name == t.name or m.team2.name == t.name]
            # sort by round descending so we drop later rounds first
            related.sort(key=lambda m: (-m.round_num, m.team1.matches + m.team2.matches))
            for m in related:
                if surplus <= 0:
                    break
                # remove match and adjust opponent stats minimally (not recomputing groups_faced fully for simplicity)
                matches.remove(m)
                other = m.team2 if m.team1.name == t.name else m.team1
                t.matches -= 1
                other.matches -= 1
                # leave opponents sets (slight inconsistency acceptable given rarity)
                surplus -= 1

    # Last-chance fill per level if still below theoretical level totals (e.g., 71 vs 72)
    level_current_totals = {lvl: sum(1 for m in matches if m.team1.level == lvl) for lvl in ['A','B','C']}
    for lvl in ['A','B','C']:
        need = desired_total_matches[lvl] - level_current_totals[lvl]
        if need <= 0:
            continue
        # attempt to add 'need' matches respecting constraints; may temporarily overflow some teams which we'll prune after
        # build team list for level
        lvl_teams = [t for t in teams if t.level == lvl]
        attempts = 0
        while need > 0 and attempts < 200:
            attempts += 1
            # pick an underplayed team first else any team still having available opponent options
            base_pool = [t for t in lvl_teams if t.matches < TARGET_MATCHES_PER_TEAM]
            if not base_pool:
                base_pool = sorted(lvl_teams, key=lambda x: (x.matches, x.name))
            base_pool.sort(key=lambda x: (x.matches, x.name))
            placed = False
            for t1 in base_pool:
                # find candidate opponents (avoid same group and rematch)
                opps = [t2 for t2 in lvl_teams if t2.name != t1.name and allow_pair(t1, t2)]
                # prioritize underplayed opponents
                opps.sort(key=lambda x: (x.matches, x.name))
                for t2 in opps:
                    # find a round with both free + free court
                    for r in range(1, num_rounds+1):
                        names_r = {m.team1.name for m in matches if m.round_num == r} | {m.team2.name for m in matches if m.round_num == r}
                        if t1.name in names_r or t2.name in names_r:
                            continue
                        used_courts_r = {m.court for m in matches if m.round_num == r}
                        free_courts_r = [c for c in range(1, courts+1) if c not in used_courts_r]
                        if not free_courts_r:
                            continue
                        court = free_courts_r[0]
                        start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r-1))
                        m_new = Match(r, court, t1, t2, start_time)
                        matches.append(m_new)
                        t1.matches += 1
                        t2.matches += 1
                        t1.opponents.add(t2.name)
                        t2.opponents.add(t1.name)
                        t1.groups_faced.add(t2.group)
                        t2.groups_faced.add(t1.group)
                        need -= 1
                        placed = True
                        break
                    if placed or need <= 0:
                        break
                if placed or need <= 0:
                    break
            if not placed and need > 0:
                # displacement attempt: free a court in a round r with both teams free but fully occupied
                for t1 in base_pool:
                    if t1.matches >= TARGET_MATCHES_PER_TEAM:
                        continue
                    opps = [t2 for t2 in lvl_teams if t2.name != t1.name and allow_pair(t1, t2)]
                    opps.sort(key=lambda x: (x.matches, x.name))
                    for t2 in opps:
                        if t2.matches > TARGET_MATCHES_PER_TEAM:
                            continue
                        # find round r with both free but no free courts
                        for r in range(1, num_rounds+1):
                            names_r = {m.team1.name for m in matches if m.round_num == r} | {m.team2.name for m in matches if m.round_num == r}
                            if t1.name in names_r or t2.name in names_r:
                                continue
                            used_courts_r = {m.court for m in matches if m.round_num == r}
                            if len(used_courts_r) < courts:
                                continue  # skip; has free court, would have been placed earlier
                            # find movable saturated match of same level
                            # allow displacement of any saturated match (regardless of level) to free a court
                            movable = [m for m in matches if m.round_num == r and m.team1.matches >= TARGET_MATCHES_PER_TEAM and m.team2.matches >= TARGET_MATCHES_PER_TEAM]
                            if not movable:
                                continue
                            mv = movable[0]
                            old_court = mv.court
                            # find alternative round r2 to relocate mv
                            for r2 in range(1, num_rounds+1):
                                if r2 == r:
                                    continue
                                names_r2 = {m.team1.name for m in matches if m.round_num == r2} | {m.team2.name for m in matches if m.round_num == r2}
                                if mv.team1.name in names_r2 or mv.team2.name in names_r2:
                                    continue
                                used_courts_r2 = {m.court for m in matches if m.round_num == r2}
                                free_r2 = [c for c in range(1, courts+1) if c not in used_courts_r2]
                                if not free_r2:
                                    continue
                                mv.round_num = r2
                                mv.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r2-1))
                                mv.court = free_r2[0]
                                # schedule new match in r using old_court
                                court_new = old_court
                                start_time_new = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r-1))
                                m_new = Match(r, court_new, t1, t2, start_time_new)
                                matches.append(m_new)
                                t1.matches += 1
                                t2.matches += 1
                                t1.opponents.add(t2.name)
                                t2.opponents.add(t1.name)
                                t1.groups_faced.add(t2.group)
                                t2.groups_faced.add(t1.group)
                                need -= 1
                                placed = True
                                break
                            if placed or need <= 0:
                                break
                        if placed or need <= 0:
                            break
                    if placed or need <= 0:
                        break
                if not placed:
                    break  # cannot place further; exit loop
        # refresh current totals for potential next level iteration
        level_current_totals[lvl] = sum(1 for m in matches if m.team1.level == lvl)

    # Prune any overflow again after last-chance fill
    for t in teams:
        if t.matches > TARGET_MATCHES_PER_TEAM:
            surplus = t.matches - TARGET_MATCHES_PER_TEAM
            related = [m for m in matches if m.team1.name == t.name or m.team2.name == t.name]
            related.sort(key=lambda m: (-m.round_num, m.team1.matches + m.team2.matches))
            for m in related:
                if surplus <= 0:
                    break
                matches.remove(m)
                other = m.team2 if m.team1.name == t.name else m.team1
                t.matches -= 1
                other.matches -= 1
                surplus -= 1

    # Final force fill: direct injection until global target reached
    if len(matches) < target_total:
        def teams_underplayed():
            return [t for t in teams if t.matches < TARGET_MATCHES_PER_TEAM]
        def round_names(r):
            return {m.team1.name for m in matches if m.round_num == r} | {m.team2.name for m in matches if m.round_num == r}
        def free_courts(r):
            used = {m.court for m in matches if m.round_num == r}
            return [c for c in range(1, courts+1) if c not in used]
        attempts = 0
        while len(matches) < target_total and attempts < 200:
            attempts += 1
            ups = teams_underplayed()
            if len(ups) < 2:
                break
            # group by level for pairing preference
            ups.sort(key=lambda x: (x.level, x.matches, x.name))
            paired = False
            for i in range(len(ups)):
                t1 = ups[i]
                # build partner pool: all same-level teams (including saturated) not same group and not rematch
                partner_pool = [t2 for t2 in teams if t2.level == t1.level and allow_pair(t1, t2)]
                # prioritize underplayed then by matches then name
                partner_pool.sort(key=lambda x: (x.matches >= TARGET_MATCHES_PER_TEAM, x.matches, x.name))
                for t2 in partner_pool:
                    # try direct placement
                    placed = False
                    for r in range(1, num_rounds+1):
                        rnms = round_names(r)
                        if t1.name in rnms or t2.name in rnms:
                            continue
                        fc = free_courts(r)
                        if fc:
                            court = fc[0]
                            start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r-1))
                            m_new = Match(r, court, t1, t2, start_time)
                            matches.append(m_new)
                            t1.matches += 1
                            t2.matches += 1
                            t1.opponents.add(t2.name)
                            t2.opponents.add(t1.name)
                            t1.groups_faced.add(t2.group)
                            t2.groups_faced.add(t1.group)
                            placed = True
                            paired = True
                            break
                    if placed:
                        break
                    # displacement: move any saturated match from candidate round
                    if not placed:
                        for r in range(1, num_rounds+1):
                            rnms = round_names(r)
                            if t1.name in rnms or t2.name in rnms:
                                continue
                            if not free_courts(r):  # if full
                                # find saturated match to move
                                sat = [m for m in matches if m.round_num == r and m.team1.matches >= TARGET_MATCHES_PER_TEAM and m.team2.matches >= TARGET_MATCHES_PER_TEAM]
                                if not sat:
                                    continue
                                mv = sat[0]
                                old_court = mv.court
                                # find target r2
                                for r2 in range(1, num_rounds+1):
                                    if r2 == r:
                                        continue
                                    rnms2 = round_names(r2)
                                    if mv.team1.name in rnms2 or mv.team2.name in rnms2:
                                        continue
                                    fc2 = free_courts(r2)
                                    if not fc2:
                                        continue
                                    mv.round_num = r2
                                    mv.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r2-1))
                                    mv.court = fc2[0]
                                    # now place new
                                    start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r-1))
                                    m_new = Match(r, old_court, t1, t2, start_time)
                                    matches.append(m_new)
                                    t1.matches += 1
                                    t2.matches += 1
                                    t1.opponents.add(t2.name)
                                    t2.opponents.add(t1.name)
                                    t1.groups_faced.add(t2.group)
                                    t2.groups_faced.add(t1.group)
                                    paired = True
                                    placed = True
                                    break
                                if placed:
                                    break
                        if placed:
                            break
                if paired:
                    break
            if not paired:
                break
        # prune overflow again
        for t in teams:
            if t.matches > TARGET_MATCHES_PER_TEAM:
                surplus = t.matches - TARGET_MATCHES_PER_TEAM
                related = [m for m in matches if m.team1.name == t.name or m.team2.name == t.name]
                related.sort(key=lambda m: (-m.round_num, m.team1.matches + m.team2.matches))
                for m in related:
                    if surplus <= 0:
                        break
                    matches.remove(m)
                    other = m.team2 if m.team1.name == t.name else m.team1
                    t.matches -= 1
                    other.matches -= 1
                    surplus -= 1

    return sorted(matches, key=lambda m: (m.round_num, m.court))

def compute_diversity_score(teams: List[Team]) -> int:
    # 分散スコア: 全ペアの対戦グループ種類総和
    return sum(len(t.groups_faced) for t in teams)


def _max_team_streak(matches: List[Match]) -> int:
    rounds_map: Dict[str, List[int]] = defaultdict(list)
    for m in matches:
        rounds_map[m.team1.name].append(int(m.round_num))
        rounds_map[m.team2.name].append(int(m.round_num))
    best = 0
    for rs in rounds_map.values():
        if not rs:
            continue
        sr = sorted(rs)
        cur = 1
        mx = 1
        for i in range(1, len(sr)):
            if sr[i] == sr[i - 1] + 1:
                cur += 1
                mx = max(mx, cur)
            else:
                cur = 1
        best = max(best, mx)
    return best


def _count_consecutive_pairs(matches: List[Match]) -> int:
    """Count adjacent-round occurrences (1 gap) across all teams.

    This is used as a soft objective: fewer back-to-backs is better,
    while still keeping matches-per-team fixed.
    """

    rounds_map: Dict[str, List[int]] = defaultdict(list)
    for m in matches:
        rounds_map[m.team1.name].append(int(m.round_num))
        rounds_map[m.team2.name].append(int(m.round_num))
    total = 0
    for rs in rounds_map.values():
        if not rs:
            continue
        sr = sorted(rs)
        for i in range(1, len(sr)):
            if sr[i] == sr[i - 1] + 1:
                total += 1
    return total


def refresh_team_stats(teams: List[Team], matches: List[Match]) -> None:
    for t in teams:
        t.matches = 0
        t.opponents = set()
        t.groups_faced = set()
        t.last_round = 0
    for match in matches:
        for team, opp in ((match.team1, match.team2), (match.team2, match.team1)):
            team.matches += 1
            team.opponents.add(opp.name)
            team.groups_faced.add(opp.group)
            team.last_round = max(team.last_round, match.round_num)


def boost_group_diversity(matches: List[Match], teams: List[Team], min_unique: int = 4,
                          max_swaps: int = 200) -> List[Match]:
    """Swap opponents within the same level to increase unique opponent groups."""
    refresh_team_stats(teams, matches)
    level_groups: Dict[str, Set[str]] = {
        lvl: {t.group for t in teams if t.level == lvl}
        for lvl in ['A', 'B', 'C']
    }

    def build_team_match_map() -> Dict[str, List[Match]]:
        mapping: Dict[str, List[Match]] = defaultdict(list)
        for match in matches:
            mapping[match.team1.name].append(match)
            mapping[match.team2.name].append(match)
        return mapping

    def missing_groups(team: Team) -> List[str]:
        return [g for g in level_groups[team.level]
                if g != team.group and g not in team.groups_faced]

    def can_pair(a: Team, b: Team) -> bool:
        if a.name == b.name:
            return False
        if a.group == b.group:
            return False
        if b.name in a.opponents:
            return False
        return True

    def assign_pair(match_obj: Match, left: Team, right: Team) -> None:
        match_obj.team1 = left
        match_obj.team2 = right

    def extract_candidate(match_obj: Match, target_group: str) -> Tuple[Optional[Team], Optional[Team]]:
        if match_obj.team1.group == target_group:
            return match_obj.team1, match_obj.team2
        if match_obj.team2.group == target_group:
            return match_obj.team2, match_obj.team1
        return None, None

    def try_simple_swap(team: Team, team_matches_map: Dict[str, List[Match]]) -> bool:
        team_matches = team_matches_map.get(team.name, [])
        if not team_matches:
            return False
        for target_group in missing_groups(team):
            candidate_matches = [m for m in matches if m.team1.level == team.level
                                 and (m.team1.group == target_group or m.team2.group == target_group)]
            if not candidate_matches:
                continue
            random.shuffle(candidate_matches)
            for match_team in team_matches:
                opp1 = match_team.team2 if match_team.team1 == team else match_team.team1
                for match_candidate in candidate_matches:
                    candidate, opp2 = extract_candidate(match_candidate, target_group)
                    if not candidate or not opp2:
                        continue
                    if not can_pair(team, candidate):
                        continue
                    if not can_pair(opp1, opp2):
                        continue
                    assign_pair(match_team, team, candidate)
                    assign_pair(match_candidate, opp1, opp2)
                    return True
        return False

    def try_three_match_swap(team: Team, team_matches_map: Dict[str, List[Match]]) -> bool:
        team_matches = team_matches_map.get(team.name, [])
        if not team_matches:
            return False
        for target_group in missing_groups(team):
            candidate_matches = [m for m in matches if m.team1.level == team.level
                                 and (m.team1.group == target_group or m.team2.group == target_group)]
            if not candidate_matches:
                continue
            random.shuffle(candidate_matches)
            for match_team in team_matches:
                opp1 = match_team.team2 if match_team.team1 == team else match_team.team1
                for match_candidate in candidate_matches:
                    candidate, opp2 = extract_candidate(match_candidate, target_group)
                    if not candidate or not opp2:
                        continue
                    if candidate.name in {team.name, opp1.name}:
                        continue
                    if opp2.name in {team.name, opp1.name}:
                        continue
                    if candidate.name in team.opponents:
                        continue
                    for bridge in matches:
                        if bridge in (match_team, match_candidate):
                            continue
                        if bridge.team1.level != team.level:
                            continue
                        endpoints = [(bridge.team1, bridge.team2), (bridge.team2, bridge.team1)]
                        for alt, opp_alt in endpoints:
                            involved = {team.name, opp1.name, candidate.name, opp2.name, alt.name, opp_alt.name}
                            if len(involved) < 6:
                                continue
                            if not can_pair(opp1, alt):
                                continue
                            if not can_pair(opp2, opp_alt):
                                continue
                            assign_pair(match_team, team, candidate)
                            assign_pair(match_candidate, opp1, alt)
                            assign_pair(bridge, opp2, opp_alt)
                            return True
        return False

    swaps = 0
    while swaps < max_swaps:
        low_diversity = [t for t in teams if len(t.groups_faced) < min_unique]
        if not low_diversity:
            break
        low_diversity.sort(key=lambda t: (len(t.groups_faced), t.level, t.group, t.name))
        team_matches_map = build_team_match_map()
        progressed = False
        for team in low_diversity:
            if try_simple_swap(team, team_matches_map):
                swaps += 1
                refresh_team_stats(teams, matches)
                progressed = True
                break
            if try_three_match_swap(team, team_matches_map):
                swaps += 1
                refresh_team_stats(teams, matches)
                progressed = True
                break
        if not progressed:
            break
    return matches

def rebalance_vertical_distribution(matches: List[Match], teams: List[Team], num_rounds: int, courts: int) -> List[Match]:
    """各レベルの試合数をラウンド方向(1..num_rounds)へ均等割り当てに近づけ、
    その後 A 左 / B 中央 / C 右 のコート帯を強制再配置する。
    既存試合を移動するだけで新規生成はしないので全ペアTARGET_MATCHES_PER_TEAM試合の成立を壊さない。
    アルゴリズム:
      1. 現在の per-round level counts を取得
      2. 各レベル total を公平分割: base = total//num_rounds, remainder 個のラウンドに +1
      3. deficit ラウンドに対し surplus ラウンドから移動可能な試合を探し移動
         - 移動候補は source の同レベル試合で source level count > target(source)
         - 移動先ラウンドで両チームが未出場 & 容量 (< courts)
      4. 全レベル調整後、各ラウンドでコート再割当 (A 左詰 / C 右詰 / B 中央連続ブロック)
    """
    # インデックス構築
    round_matches: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds+1)}
    for m in matches:
        round_matches[m.round_num].append(m)

    def level_counts_per_round(level: str) -> Dict[int, int]:
        return {r: sum(1 for m in round_matches[r] if m.team1.level == level) for r in range(1, num_rounds+1)}

    def participants_in_round(r: int) -> Set[str]:
        return {m.team1.name for m in round_matches[r]} | {m.team2.name for m in round_matches[r]}

    level_priority = {'B': 0, 'C': 1, 'A': 2}

    def ensure_space(target_round: int, avoid_level: str | None, counts_map: Dict[int, int]) -> bool:
        if len(round_matches[target_round]) < courts:
            return True
        movable = round_matches[target_round][:]
        if avoid_level:
            filtered = [m for m in movable if m.team1.level != avoid_level]
            if filtered:
                movable = filtered
        movable.sort(key=lambda m: (level_priority.get(m.team1.level, 3), -m.round_num))
        search = list(range(target_round+1, num_rounds+1)) + list(range(1, target_round))
        for mv in movable:
            for r2 in search:
                if len(round_matches[r2]) >= courts:
                    continue
                names2 = participants_in_round(r2)
                if mv.team1.name in names2 or mv.team2.name in names2:
                    continue
                round_matches[target_round].remove(mv)
                round_matches[r2].append(mv)
                mv.round_num = r2
                mv.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r2-1))
                if avoid_level and mv.team1.level == avoid_level:
                    counts_map[target_round] -= 1
                    counts_map[r2] += 1
                return True
        return False

    totals = {lvl: sum(1 for m in matches if m.team1.level == lvl) for lvl in ['A','B','C']}
    targets: Dict[str, Dict[int,int]] = {lvl: {} for lvl in ['A','B','C']}
    for lvl in ['A','B','C']:
        total = totals[lvl]
        base = total // num_rounds
        rem = total - base * num_rounds
        # 先頭 rem ラウンドに +1 (視覚的に早いラウンドへ分散)
        for r in range(1, num_rounds+1):
            targets[lvl][r] = base + (1 if r <= rem else 0)

    # 移動ヘルパー
    def can_move(match: Match, dest_round: int) -> bool:
        if len(round_matches[dest_round]) >= courts:
            return False
        names_dest = {mm.team1.name for mm in round_matches[dest_round]} | {mm.team2.name for mm in round_matches[dest_round]}
        if match.team1.name in names_dest or match.team2.name in names_dest:
            return False
        return True

    # レベルごとにバランス調整
    for lvl in ['A','B','C']:
        counts = level_counts_per_round(lvl)
        # deficit rounds sorted by descending deficit
        def_rounds = [r for r in range(1, num_rounds+1) if counts[r] < targets[lvl][r]]
        sur_rounds = [r for r in range(1, num_rounds+1) if counts[r] > targets[lvl][r]]
        # なるべく前半ラウンドを優先して埋める
        def_rounds.sort()
        # Surplus: 後半から優先的に抜く (既存偏り除去)
        sur_rounds.sort(reverse=True)
        for r_def in def_rounds:
            need = targets[lvl][r_def] - counts[r_def]
            if need <= 0:
                continue
            # 充足するまで surplus source から移動
            for r_src in list(sur_rounds):
                if need <= 0:
                    break
                if counts[r_src] <= targets[lvl][r_src]:
                    continue
                # 候補試合
                movable = [m for m in round_matches[r_src] if m.team1.level == lvl]
                # 安定順: 遅いラウンド(大きい)の試合を優先移動
                for m in movable:
                    if counts[r_src] <= targets[lvl][r_src] or need <= 0:
                        break
                    if len(round_matches[r_def]) >= courts:
                        if not ensure_space(r_def, lvl, counts):
                            continue
                    if not can_move(m, r_def):
                        continue
                    # 移動実行
                    round_matches[r_src].remove(m)
                    old_round = m.round_num
                    m.round_num = r_def
                    m.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r_def-1))
                    round_matches[r_def].append(m)
                    counts[r_src] -= 1
                    counts[r_def] += 1
                    need -= 1
                # surplus 更新
                if counts[r_src] <= targets[lvl][r_src] and r_src in sur_rounds:
                    sur_rounds.remove(r_src)
            # need 未達でも次へ (これ以上安全に移せない)

    # 最終コート帯再割当
    for r in range(1, num_rounds+1):
        rm = round_matches[r]
        if not rm:
            continue
        a_ms = [m for m in rm if m.team1.level == 'A']
        b_ms = [m for m in rm if m.team1.level == 'B']
        c_ms = [m for m in rm if m.team1.level == 'C']
        nA, nB, nC = len(a_ms), len(b_ms), len(c_ms)
        a_targets = list(range(1, nA+1))
        c_targets = list(range(courts, courts - nC, -1))
        used = set(a_targets) | set(c_targets)
        # B block contiguous central
        b_targets: List[int] = []
        if nB > 0:
            center = (courts + 1)/2
            candidates = []
            free = [c for c in range(1, courts+1) if c not in used]
            # 連続ブロック探索
            free_set = set(free)
            for start in range(1, courts - nB + 2):
                block = list(range(start, start + nB))
                if all(c in free_set for c in block):
                    b_center = (block[0] + block[-1])/2
                    candidates.append((abs(b_center - center), block))
            if candidates:
                candidates.sort(key=lambda x: x[0])
                b_targets = candidates[0][1]
            else:
                b_targets = free[:nB]
        for m,c in zip(a_ms, a_targets):
            m.court = c
        for m,c in zip(b_ms, b_targets):
            m.court = c
        for m,c in zip(c_ms, c_targets):
            m.court = c

    # 返却
    flat = [m for r in range(1, num_rounds+1) for m in round_matches[r]]
    return sorted(flat, key=lambda m: (m.round_num, m.court))

def enforce_segments_and_quotas(matches: List[Match], teams: List[Team], num_rounds: int, courts: int,
                                seg_A: Tuple[int,int]=(1,4), seg_B: Tuple[int,int]=(5,12), seg_C: Tuple[int,int]=(13,15)) -> List[Match]:
    """レベル毎のコートセグメントとラウンド均等クォータを強制再配置。
    デフォルト: A=1-4, B=5-12, C=13-15 (コート15面前提)
    クォータ算出: total_level_matches / num_rounds を床/余り分配。
    既存 matches を移動のみで調整。移動不可なら残差を許容。
    """
    # セグメント正規化
    def norm_seg(seg):
        a,b = seg
        if a>b:
            a,b = b,a
        return (max(1,a), min(courts,b))
    seg_A = norm_seg(seg_A); seg_B = norm_seg(seg_B); seg_C = norm_seg(seg_C)
    seg_map = {'A': seg_A, 'B': seg_B, 'C': seg_C}
    # ラウンド別配列
    round_matches: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds+1)}
    for m in matches:
        round_matches[m.round_num].append(m)
    # 現在のコートは改めて後段で再割当するため一旦 court=0 化
    for m in matches:
        m.court = 0

    # クォータ計算
    totals = {lvl: sum(1 for m in matches if m.team1.level == lvl) for lvl in ['A','B','C']}
    targets: Dict[str, Dict[int,int]] = {lvl: {} for lvl in ['A','B','C']}
    for lvl in ['A','B','C']:
        total = totals[lvl]
        base = total // num_rounds
        rem = total - base * num_rounds
        for r in range(1, num_rounds+1):
            targets[lvl][r] = base + (1 if r <= rem else 0)

    # helper: teams present in round
    def names_in_round(r: int) -> Set[str]:
        return {mm.team1.name for mm in round_matches[r]} | {mm.team2.name for mm in round_matches[r]}

    # level counts per round
    def level_counts(lvl: str) -> Dict[int,int]:
        return {r: sum(1 for m in round_matches[r] if m.team1.level == lvl) for r in range(1,num_rounds+1)}

    # 移動可能性チェック
    def can_place(match: Match, dest_round: int) -> bool:
        ns = names_in_round(dest_round)
        if match.team1.name in ns or match.team2.name in ns:
            return False
        # 容量チェック (総試合数 < courts)
        if len(round_matches[dest_round]) >= courts:
            return False
        return True

    # 調整ループ (各レベル)
    for lvl in ['A','B','C']:
        counts = level_counts(lvl)
        deficit = [r for r in range(1,num_rounds+1) if counts[r] < targets[lvl][r]]
        surplus = [r for r in range(1,num_rounds+1) if counts[r] > targets[lvl][r]]
        deficit.sort()  # 早いラウンド優先で埋める
        surplus.sort(reverse=True)  # 遅いラウンドから抜く
        for r_def in deficit:
            need = targets[lvl][r_def] - counts[r_def]
            if need <= 0:
                continue
            for r_src in list(surplus):
                if need <= 0:
                    break
                if counts[r_src] <= targets[lvl][r_src]:
                    continue
                # 移動候補 (source ラウンドの lvl 試合)
                movable = [m for m in round_matches[r_src] if m.team1.level == lvl]
                # 安定順: 最も遅いラウンドの前半試合を先に移動
                for mv in movable:
                    if counts[r_src] <= targets[lvl][r_src] or need <= 0:
                        break
                    if not can_place(mv, r_def):
                        continue
                    # 移動
                    round_matches[r_src].remove(mv)
                    mv.round_num = r_def
                    mv.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r_def-1))
                    round_matches[r_def].append(mv)
                    counts[r_src] -= 1
                    counts[r_def] += 1
                    need -= 1
                if counts[r_src] <= targets[lvl][r_src]:
                    surplus.remove(r_src)
            # need 未充足は許容

    # コート割当: セグメント内で A 左詰 / B 左詰(セグメント内中央連続は後続改善可) / C 右詰
    for r in range(1,num_rounds+1):
        rm = round_matches[r]
        if not rm:
            continue
        a_ms = [m for m in rm if m.team1.level=='A']
        b_ms = [m for m in rm if m.team1.level=='B']
        c_ms = [m for m in rm if m.team1.level=='C']
        segA_start, segA_end = seg_A
        segB_start, segB_end = seg_B
        segC_start, segC_end = seg_C
        # A 左詰
        a_targets = (list(range(segA_start, segA_end + 1)) * ((len(a_ms) + (segA_end - segA_start)) // (segA_end - segA_start + 1)))[:len(a_ms)]
        # C 右詰 (逆順)
        c_targets_full = (list(range(segC_start, segC_end + 1)) * ((len(c_ms) + (segC_end - segC_start)) // (segC_end - segC_start + 1)))[:len(c_ms)]
        c_targets_full.sort(reverse=True)
        c_targets = c_targets_full
        # B セグメント内連続ブロック (まず単純左詰)
        b_targets = (list(range(segB_start, segB_end + 1)) * ((len(b_ms) + (segB_end - segB_start)) // (segB_end - segB_start + 1)))[:len(b_ms)]
        # 衝突除去 (他レベル割当外し) — セグメントが明確に分離されている前提なので不要だが保険
        used = set()
        for m,c in zip(a_ms,a_targets):
            m.court = c; used.add(c)
        for m,c in zip(c_ms,c_targets):
            m.court = c; used.add(c)
        # B 割当フィルタ (未使用のみ)
        b_targets = [c for c in b_targets if c not in used][:len(b_ms)]
        for m,c in zip(b_ms,b_targets):
            m.court = c; used.add(c)
        # 余った試合 (理論上なし) はセグメント内空きコート再検索
        leftover = [m for m in rm if m.court == 0]
        if leftover:
            # fallback: 全コート探索し未使用を割当
            free_all = [c for c in range(1,courts+1) if c not in used]
            for m,c in zip(leftover, free_all):
                m.court = c

    # 平坦化返却
    flat = [m for r in range(1,num_rounds+1) for m in round_matches[r]]
    return sorted(flat, key=lambda m: (m.round_num, m.court))

def balanced_round_reassignment(matches: List[Match], num_rounds: int, courts: int) -> List[Match]:
    """各ラウンドに A/B/C を必ず含め、ターゲット分布 (A:3or4, B:8or9, C:3or4) を満たすように
    既存同レベル試合集合を再割当する。 チーム重複はラウンド内禁止を維持。 失敗時は元の matches 返す。
    前提: matches は同レベル対戦のみで総数 336 (=75A+189B+72C)。
    分配ロジック: baseline A=3 B=8 C=3 を全23ラウンドへ、余剰 A=6, B=5, C=3 を異なるラウンドへ均等散布。
    Aレベルを前半ラウンドに優先配置。
    """
    # レベル別試合抽出
    level_matches = {
        'A': [m for m in matches if m.team1.level == 'A'],
        'B': [m for m in matches if m.team1.level == 'B'],
        'C': [m for m in matches if m.team1.level == 'C'],
    }
    # ターゲット計算
    base = {'A':3, 'B':8, 'C':3}
    extras = {'A':6, 'B':5, 'C':3}
    # ラウンド配列
    targets = {r:{'A':base['A'], 'B':base['B'], 'C':base['C']} for r in range(1, num_rounds+1)}
    def spaced_indices(k: int, offset: float = 0.0) -> List[int]:
        if k <= 0:
            return []
        step = num_rounds / k
        return [int((offset + i * step) % num_rounds) for i in range(k)]

    capacity = []
    for r in range(1, num_rounds+1):
        used = sum(targets[r][lvl] for lvl in ['A','B','C'])
        capacity.append(courts - used)
    # distribute extras while respecting per-round capacity (>=0)
    priority = (('A', 0.0), ('C', num_rounds * 0.35), ('B', num_rounds * 0.65))
    for lvl, offset in priority:
        rem = extras[lvl]
        if rem <= 0:
            continue
        for idx in spaced_indices(rem, offset):
            if rem <= 0:
                break
            picked = None
            for shift in range(num_rounds):
                ridx = (idx + shift) % num_rounds
                if capacity[ridx] > 0:
                    picked = ridx
                    break
            if picked is None:
                continue
            targets[picked+1][lvl] += 1
            capacity[picked] -= 1
            rem -= 1
    if any(cap < 0 for cap in capacity):
        return matches
    # 割当構造
    new_rounds = {r: [] for r in range(1, num_rounds+1)}
    used_names_per_round = {r: set() for r in range(1, num_rounds+1)}
    # レベル毎に決定的並び (チーム名 + 元ラウンドで安定)
    for lvl in ['A','B','C']:
        level_matches[lvl].sort(key=lambda m: (m.team1.name, m.team2.name, m.round_num))
    # 割当アルゴリズム: 各レベル試合リストを走査し最初に収容可能ラウンドへ配置
    for lvl in ['A','B','C']:
        remaining = level_matches[lvl][:]
        # ラウンド順序: Aは前半優先, 他のレベルは余剰のあるラウンドを先に
        if lvl == 'A':
            round_order = list(range(1, num_rounds+1))  # 前から順
        else:
            round_order = sorted(range(1,num_rounds+1), key=lambda r: (-targets[r][lvl], r))
        # ラウンド毎の残枠カウンタ
        slots = {r: targets[r][lvl] for r in range(1,num_rounds+1)}
        for match in remaining:
            placed = False
            # 可能なラウンドを優先順序で検索
            for r in round_order:
                if slots[r] <= 0:
                    continue
                names = used_names_per_round[r]
                if match.team1.name in names or match.team2.name in names:
                    continue
                # 配置
                new_rounds[r].append(match)
                names.add(match.team1.name); names.add(match.team2.name)
                slots[r] -= 1
                placed = True
                break
            if not placed:
                # バックオフ: 使用可能ラウンドの中でどちらかが重複するものを避けつつ再探索 (緩和で重複許容はしない)
                # シンプルフォールバック: 元ラウンドに残す
                orig_r = match.round_num
                # 元ラウンドが満杯または重複なら次の空きラウンドへ (緩和)
                for r in range(1,num_rounds+1):
                    if slots[r] <= 0:
                        continue
                    names = used_names_per_round[r]
                    if match.team1.name in names or match.team2.name in names:
                        continue
                    new_rounds[r].append(match)
                    names.add(match.team1.name); names.add(match.team2.name)
                    slots[r] -= 1
                    placed = True
                    break
                if not placed:
                    # どうしても入らない → 元スケジュール維持 (失敗フラグ)
                    return matches
    # コート再配置: A 左 / B 中央 / C 右 （セグメント利用）
    for r in range(1, num_rounds+1):
        rm = new_rounds[r]
        a_ms = [m for m in rm if m.team1.level == 'A']
        b_ms = [m for m in rm if m.team1.level == 'B']
        c_ms = [m for m in rm if m.team1.level == 'C']
        courts_left = list(range(1, courts+1))
        # A 左: 1..4 から優先
        a_band = [c for c in range(1,5)]
        for m,c in zip(a_ms, a_band[:len(a_ms)]):
            m.round_num = r
            m.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r-1))
            m.court = c; courts_left.remove(c)
        # C 右: 13..15 優先
        c_band = [c for c in range(courts, courts-3, -1)]  # 15,14,13
        for m,c in zip(c_ms, c_band[:len(c_ms)]):
            m.round_num = r
            m.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r-1))
            if c in courts_left:
                m.court = c; courts_left.remove(c)
            else:
                m.court = courts_left.pop(0)
        # B 中央: 残りから中央に近い順
        b_ms_sorted = b_ms
        center = (courts+1)/2
        courts_left.sort(key=lambda c: abs(c - center))
        for m,c in zip(b_ms_sorted, courts_left[:len(b_ms_sorted)]):
            m.round_num = r
            m.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r-1))
            m.court = c
    # 平坦化
    reassigned = [m for r in range(1, num_rounds+1) for m in new_rounds[r]]
    return sorted(reassigned, key=lambda m: (m.round_num, m.court))


def tighten_level_bands(matches: List[Match], num_rounds: int, courts: int,
                        segments: Dict[str, Tuple[int, int]] | None = None) -> List[Match]:
    """Final per-round swap pass to keep A/B/C matches inside their preferred court bands.
    If a band's courts are fully occupied by the same level, no action is taken."""
    segs = segments or LEVEL_SEGMENTS
    round_lookup: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds+1)}
    for m in matches:
        round_lookup[m.round_num].append(m)

    def band_distance(level: str, court: int) -> int:
        start, end = segs[level]
        if court < start:
            return start - court
        if court > end:
            return court - end
        return 0

    def swap_into_band(level: str, rm: List[Match]):
        offenders = [m for m in rm if m.team1.level == level and band_distance(level, m.court) > 0]
        for offender in offenders:
            best: Tuple[int, Match] | None = None
            for candidate in rm:
                if candidate is offender:
                    continue
                before = band_distance(level, offender.court) + band_distance(candidate.team1.level, candidate.court)
                after = band_distance(level, candidate.court) + band_distance(candidate.team1.level, offender.court)
                gain = before - after
                if gain <= 0:
                    continue
                if best is None or gain > best[0]:
                    best = (gain, candidate)
            if best:
                donor = best[1]
                offender.court, donor.court = donor.court, offender.court

    for r in range(1, num_rounds+1):
        rm = round_lookup[r]
        if not rm:
            continue
        for level in ['A', 'C', 'B']:
            swap_into_band(level, rm)

    return sorted(matches, key=lambda m: (m.round_num, m.court))


def reduce_back_to_back(matches: List[Match], num_rounds: int, courts: int,
                        max_iterations: int = 400) -> List[Match]:
    """Relocate matches away from consecutive rounds when spare slots exist."""
    if not matches:
        return matches

    round_matches: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds + 1)}
    round_courts: Dict[int, Set[int]] = {r: set() for r in range(1, num_rounds + 1)}
    round_names: Dict[int, Set[str]] = {r: set() for r in range(1, num_rounds + 1)}
    for match in matches:
        round_matches.setdefault(match.round_num, []).append(match)
        round_courts.setdefault(match.round_num, set()).add(match.court)
        round_names.setdefault(match.round_num, set()).add(match.team1.name)
        round_names[match.round_num].add(match.team2.name)

    def build_team_rounds() -> Dict[str, List[int]]:
        mapping: Dict[str, List[int]] = defaultdict(list)
        for match in matches:
            mapping[match.team1.name].append(match.round_num)
            mapping[match.team2.name].append(match.round_num)
        for rounds in mapping.values():
            rounds.sort()
        return mapping

    def neighbor_penalty(rounds: List[int], center: int) -> int:
        prev_round = max((r for r in rounds if r < center), default=None)
        next_round = min((r for r in rounds if r > center), default=None)
        penalty = 0
        if prev_round is not None and center - prev_round == 1:
            penalty += 1
        if next_round is not None and next_round - center == 1:
            penalty += 1
        return penalty

    def match_penalty(match: Match, team_rounds_map: Dict[str, List[int]]) -> int:
        return neighbor_penalty(team_rounds_map[match.team1.name], match.round_num) + \
            neighbor_penalty(team_rounds_map[match.team2.name], match.round_num)

    def penalty_for_round(team_name: str, current_round: int, new_round: int,
                          team_rounds_map: Dict[str, List[int]]) -> int:
        rounds = [r for r in team_rounds_map[team_name] if r != current_round]
        rounds.append(new_round)
        rounds.sort()
        idx = rounds.index(new_round)
        penalty = 0
        if idx > 0 and new_round - rounds[idx - 1] == 1:
            penalty += 1
        if idx + 1 < len(rounds) and rounds[idx + 1] - new_round == 1:
            penalty += 1
        return penalty

    def penalty_if_moved(match: Match, new_round: int, team_rounds_map: Dict[str, List[int]]) -> int:
        return (penalty_for_round(match.team1.name, match.round_num, new_round, team_rounds_map) +
                penalty_for_round(match.team2.name, match.round_num, new_round, team_rounds_map))

    def min_gap(team_name: str, round_num: int, current_round: Optional[int],
                team_rounds_map: Dict[str, List[int]]) -> float:
        rounds = list(team_rounds_map[team_name])
        if current_round is not None:
            rounds = [r for r in rounds if r != current_round]
            rounds.append(round_num)
        rounds.sort()
        idx = rounds.index(round_num)
        left_gap = round_num - rounds[idx - 1] if idx > 0 else float('inf')
        right_gap = rounds[idx + 1] - round_num if idx + 1 < len(rounds) else float('inf')
        return min(left_gap, right_gap)

    def collect_offenders(team_rounds_map: Dict[str, List[int]]) -> List[Tuple[int, float, Match]]:
        offenders: List[Tuple[int, float, Match]] = []
        for match in matches:
            penalty = match_penalty(match, team_rounds_map)
            if penalty == 0:
                continue
            gap = min(
                min_gap(match.team1.name, match.round_num, None, team_rounds_map),
                min_gap(match.team2.name, match.round_num, None, team_rounds_map)
            )
            offenders.append((penalty, gap, match))
        offenders.sort(key=lambda x: (-x[0], x[1]))
        return offenders

    def pick_court(round_num: int) -> Optional[int]:
        for court in range(1, courts + 1):
            if court not in round_courts.setdefault(round_num, set()):
                return court
        return None

    def try_move_match(match: Match, team_rounds_map: Dict[str, List[int]]) -> bool:
        current_penalty = match_penalty(match, team_rounds_map)
        if current_penalty == 0:
            return False
        candidates: List[Tuple[int, float, int, int]] = []
        for round_num in range(1, num_rounds + 1):
            if round_num == match.round_num:
                continue
            if len(round_matches.setdefault(round_num, [])) >= courts:
                continue
            names = round_names.setdefault(round_num, set())
            if match.team1.name in names or match.team2.name in names:
                continue
            new_penalty = penalty_if_moved(match, round_num, team_rounds_map)
            if new_penalty >= current_penalty:
                continue
            gap_score = min(
                min_gap(match.team1.name, round_num, match.round_num, team_rounds_map),
                min_gap(match.team2.name, round_num, match.round_num, team_rounds_map)
            )
            load = len(round_matches[round_num])
            candidates.append((new_penalty, -gap_score, load, round_num))
        if not candidates:
            return False
        candidates.sort(key=lambda x: (x[0], x[1], x[2], x[3]))
        target_round = candidates[0][3]
        new_court = pick_court(target_round)
        if new_court is None:
            return False

        origin_round = match.round_num
        if match in round_matches.setdefault(origin_round, []):
            round_matches[origin_round].remove(match)
        round_courts.setdefault(origin_round, set()).discard(match.court)
        round_names.setdefault(origin_round, set()).discard(match.team1.name)
        round_names[origin_round].discard(match.team2.name)

        match.round_num = target_round
        match.court = new_court
        # Time is applied later via apply_round_times().
        round_matches[target_round].append(match)
        round_courts[target_round].add(new_court)
        round_names[target_round].add(match.team1.name)
        round_names[target_round].add(match.team2.name)
        return True

    attempts = 0
    stagnation: Set[int] = set()
    while attempts < max_iterations:
        team_rounds_map = build_team_rounds()
        offenders = collect_offenders(team_rounds_map)
        if not offenders:
            break
        progressed = False
        for _, _, match in offenders:
            match_id = id(match)
            if match_id in stagnation:
                continue
            if try_move_match(match, team_rounds_map):
                attempts += 1
                progressed = True
                stagnation.clear()
                break
            stagnation.add(match_id)
        if not progressed:
            break

    return sorted(matches, key=lambda m: (m.round_num, m.court))


def reduce_max_consecutive_streak(matches: List[Match], num_rounds: int, courts: int,
                                 max_consecutive: int,
                                 max_iterations: int = 800) -> List[Match]:
    """Try to eliminate streaks longer than max_consecutive by moving matches into spare slots.

    This is a best-effort heuristic. If the schedule is tight, some teams may still end up
    exceeding the requested max_consecutive.
    """
    if not matches or max_consecutive <= 0:
        return matches
    if max_consecutive == 1:
        # Very restrictive; still attempt but often infeasible.
        pass

    round_matches: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds + 1)}
    round_courts: Dict[int, Set[int]] = {r: set() for r in range(1, num_rounds + 1)}
    round_names: Dict[int, Set[str]] = {r: set() for r in range(1, num_rounds + 1)}
    for match in matches:
        round_matches.setdefault(match.round_num, []).append(match)
        round_courts.setdefault(match.round_num, set()).add(match.court)
        round_names.setdefault(match.round_num, set()).add(match.team1.name)
        round_names[match.round_num].add(match.team2.name)

    def build_team_rounds() -> Dict[str, List[int]]:
        mapping: Dict[str, List[int]] = defaultdict(list)
        for m in matches:
            mapping[m.team1.name].append(m.round_num)
            mapping[m.team2.name].append(m.round_num)
        for rs in mapping.values():
            rs.sort()
        return mapping

    def max_streak(rounds: List[int]) -> int:
        if not rounds:
            return 0
        sr = sorted(rounds)
        best = 1
        cur = 1
        for i in range(1, len(sr)):
            if sr[i] == sr[i - 1] + 1:
                cur += 1
                if cur > best:
                    best = cur
            else:
                cur = 1
        return best

    def violation(rounds: List[int]) -> int:
        return max(0, max_streak(rounds) - max_consecutive)

    def find_longest_streak_segment(rounds: List[int]) -> Tuple[int, int]:
        """Return (start_round, end_round) inclusive for the longest consecutive segment."""
        if not rounds:
            return (0, 0)
        sr = sorted(rounds)
        best_s = sr[0]
        best_e = sr[0]
        cur_s = sr[0]
        cur_e = sr[0]
        for i in range(1, len(sr)):
            if sr[i] == sr[i - 1] + 1:
                cur_e = sr[i]
            else:
                if (cur_e - cur_s) > (best_e - best_s):
                    best_s, best_e = cur_s, cur_e
                cur_s = sr[i]
                cur_e = sr[i]
        if (cur_e - cur_s) > (best_e - best_s):
            best_s, best_e = cur_s, cur_e
        return best_s, best_e

    def pick_court(r: int) -> Optional[int]:
        for c in range(1, courts + 1):
            if c not in round_courts.setdefault(r, set()):
                return c
        return None

    def team_rounds_if_moved(team_rounds_map: Dict[str, List[int]], team_name: str,
                             current_round: int, new_round: int) -> List[int]:
        rs = [r for r in team_rounds_map.get(team_name, []) if r != current_round]
        rs.append(new_round)
        rs.sort()
        return rs

    def total_violation_for_teams(team_rounds_map: Dict[str, List[int]], team_names: Tuple[str, str]) -> int:
        return violation(team_rounds_map.get(team_names[0], [])) + violation(team_rounds_map.get(team_names[1], []))

    def choose_problem_match(team_name: str, team_rounds_map: Dict[str, List[int]]) -> Optional[Match]:
        rounds = team_rounds_map.get(team_name, [])
        if violation(rounds) <= 0:
            return None
        seg_s, seg_e = find_longest_streak_segment(rounds)
        if seg_s == 0:
            return None
        # Move the (max_consecutive+1)-th match inside the segment to break the streak.
        target_round = seg_s + max_consecutive
        if target_round > seg_e:
            target_round = seg_e
        candidates = [m for m in round_matches.get(target_round, []) if (m.team1.name == team_name or m.team2.name == team_name)]
        if candidates:
            # Prefer moving the match where the other team also violates.
            candidates.sort(
                key=lambda m: (
                    -total_violation_for_teams(team_rounds_map, (m.team1.name, m.team2.name)),
                    m.court,
                )
            )
            return candidates[0]
        # Fallback: any match in the segment.
        for r in range(seg_s, seg_e + 1):
            for m in round_matches.get(r, []):
                if m.team1.name == team_name or m.team2.name == team_name:
                    return m
        return None

    def try_move(match: Match, team_rounds_map: Dict[str, List[int]]) -> bool:
        t1 = match.team1.name
        t2 = match.team2.name
        before = violation(team_rounds_map.get(t1, [])) + violation(team_rounds_map.get(t2, []))
        if before <= 0:
            return False

        origin_round = match.round_num
        candidates: List[Tuple[int, int, int, int]] = []
        for r in range(1, num_rounds + 1):
            if r == origin_round:
                continue
            if len(round_matches.get(r, [])) >= courts:
                continue
            names = round_names.setdefault(r, set())
            if t1 in names or t2 in names:
                continue

            rs1 = team_rounds_if_moved(team_rounds_map, t1, origin_round, r)
            rs2 = team_rounds_if_moved(team_rounds_map, t2, origin_round, r)
            after = violation(rs1) + violation(rs2)
            # Prefer strict improvement; allow equal only if it increases spacing by spreading load.
            if after > before:
                continue
            load = len(round_matches.get(r, []))
            dist = abs(r - origin_round)
            candidates.append((after, load, dist, r))

        if not candidates:
            # If there is no spare slot that works, try a round-swap.

            def rounds_if_swapped(team_name: str, from_round: int, to_round: int) -> List[int]:
                rs = list(team_rounds_map.get(team_name, []))
                rs = [r for r in rs if r != from_round]
                rs.append(to_round)
                rs.sort()
                return rs

            def total_violation_for(team_names: tuple[str, ...], swaps: dict[str, tuple[int, int]]) -> int:
                total = 0
                for name in team_names:
                    if name in swaps:
                        from_r, to_r = swaps[name]
                        total += violation(rounds_if_swapped(name, from_r, to_r))
                    else:
                        total += violation(team_rounds_map.get(name, []))
                return total

            origin_names_wo = set(round_names.get(origin_round, set()))
            origin_names_wo.discard(t1)
            origin_names_wo.discard(t2)

            best_swap: tuple[int, int, int, Match] | None = None
            for r in range(1, num_rounds + 1):
                if r == origin_round:
                    continue
                names_r = round_names.setdefault(r, set())
                if t1 in names_r or t2 in names_r:
                    continue

                for other in list(round_matches.get(r, [])):
                    u1 = other.team1.name
                    u2 = other.team2.name
                    # After removing `match` from origin_round, `u1/u2` must be free there.
                    if u1 in origin_names_wo or u2 in origin_names_wo:
                        continue
                    # `other` cannot include t1/t2 due to the round-name check above.

                    before_all = (
                        violation(team_rounds_map.get(t1, []))
                        + violation(team_rounds_map.get(t2, []))
                        + violation(team_rounds_map.get(u1, []))
                        + violation(team_rounds_map.get(u2, []))
                    )

                    swaps = {
                        t1: (origin_round, r),
                        t2: (origin_round, r),
                        u1: (r, origin_round),
                        u2: (r, origin_round),
                    }
                    after_all = total_violation_for((t1, t2, u1, u2), swaps)
                    if after_all > before_all:
                        continue

                    dist = abs(r - origin_round)
                    # Prefer lower violation, closer swap, then stable ordering.
                    cand = (after_all, dist, other.court, other)
                    if best_swap is None or cand < best_swap:
                        best_swap = cand

            if best_swap is None:
                return False

            other = best_swap[3]
            target_round = other.round_num
            origin_court = match.court
            target_court = other.court

            # Update round bookkeeping: remove from old rounds.
            if match in round_matches.get(origin_round, []):
                round_matches[origin_round].remove(match)
            if other in round_matches.get(target_round, []):
                round_matches[target_round].remove(other)

            round_names.setdefault(origin_round, set()).discard(t1)
            round_names[origin_round].discard(t2)
            round_names.setdefault(target_round, set()).discard(other.team1.name)
            round_names[target_round].discard(other.team2.name)

            round_courts.setdefault(origin_round, set()).discard(origin_court)
            round_courts.setdefault(target_round, set()).discard(target_court)

            # Apply swap (reuse the freed courts so court-uniqueness holds).
            match.round_num = target_round
            match.court = target_court
            other.round_num = origin_round
            other.court = origin_court

            # Add back to new rounds.
            round_matches[target_round].append(match)
            round_courts[target_round].add(match.court)
            round_names[target_round].add(t1)
            round_names[target_round].add(t2)

            round_matches[origin_round].append(other)
            round_courts[origin_round].add(other.court)
            round_names[origin_round].add(other.team1.name)
            round_names[origin_round].add(other.team2.name)
            return True
        candidates.sort(key=lambda x: (x[0], x[1], x[2], x[3]))
        target_round = candidates[0][3]
        new_court = pick_court(target_round)
        if new_court is None:
            return False

        # Apply move
        if match in round_matches.get(origin_round, []):
            round_matches[origin_round].remove(match)
        round_courts.setdefault(origin_round, set()).discard(match.court)
        round_names.setdefault(origin_round, set()).discard(t1)
        round_names[origin_round].discard(t2)

        match.round_num = target_round
        match.court = new_court
        round_matches[target_round].append(match)
        round_courts[target_round].add(new_court)
        round_names[target_round].add(t1)
        round_names[target_round].add(t2)
        return True

    attempts = 0
    while attempts < max_iterations:
        team_rounds_map = build_team_rounds()
        offenders = [
            (violation(rs), name)
            for name, rs in team_rounds_map.items()
            if violation(rs) > 0
        ]
        if not offenders:
            break
        offenders.sort(key=lambda x: (-x[0], x[1]))
        progressed = False
        for _, team_name in offenders[:20]:
            m = choose_problem_match(team_name, team_rounds_map)
            if m is None:
                continue
            if try_move(m, team_rounds_map):
                attempts += 1
                progressed = True
                break
        if not progressed:
            break

    return sorted(matches, key=lambda m: (m.round_num, m.court))


def eliminate_mid_session_court_gaps(matches: List[Match], num_rounds: int, courts: int) -> List[Match]:
    """Shift matches forward so early rounds remain fully utilized when possible."""
    if not matches:
        return matches

    round_matches: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds + 1)}
    round_names: Dict[int, Set[str]] = {r: set() for r in range(1, num_rounds + 1)}
    round_courts: Dict[int, Set[int]] = {r: set() for r in range(1, num_rounds + 1)}
    for match in matches:
        round_matches[match.round_num].append(match)
        round_names[match.round_num].add(match.team1.name)
        round_names[match.round_num].add(match.team2.name)
        round_courts[match.round_num].add(match.court)

    for r in range(1, num_rounds):
        while len(round_matches[r]) < courts:
            available_courts = [c for c in range(1, courts + 1) if c not in round_courts[r]]
            if not available_courts:
                break
            moved = False
            for donor_round in range(num_rounds, r, -1):
                if donor_round == r or not round_matches[donor_round]:
                    continue
                for match in list(round_matches[donor_round]):
                    if match.team1.name in round_names[r] or match.team2.name in round_names[r]:
                        continue
                    round_matches[donor_round].remove(match)
                    round_names[donor_round].discard(match.team1.name)
                    round_names[donor_round].discard(match.team2.name)
                    round_courts[donor_round].discard(match.court)

                    new_court = available_courts.pop(0)
                    match.round_num = r
                    match.court = new_court
                    # Time is applied later via apply_round_times().
                    round_matches[r].append(match)
                    round_names[r].add(match.team1.name)
                    round_names[r].add(match.team2.name)
                    round_courts[r].add(new_court)
                    moved = True
                    break
                if moved:
                    break
            if not moved:
                break

    flattened: List[Match] = []
    for r in range(1, num_rounds + 1):
        flattened.extend(round_matches[r])
    return sorted(flattened, key=lambda m: (m.round_num, m.court))


def ensure_round_one_full(matches: List[Match], num_rounds: int, courts: int) -> List[Match]:
    """Move matches from later rounds so that round 1 uses all courts when possible."""
    if not matches:
        return matches

    round_lookup: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds + 1)}
    for match in matches:
        round_lookup.setdefault(match.round_num, []).append(match)

    round_one = round_lookup.get(1, [])
    if len(round_one) >= courts:
        return matches

    used_names: Set[str] = set()
    used_courts: Set[int] = set()
    for match in round_one:
        used_names.add(match.team1.name)
        used_names.add(match.team2.name)
        used_courts.add(match.court)
    available_courts = [c for c in range(1, courts + 1) if c not in used_courts]

    for donor_round in range(2, num_rounds + 1):
        if len(round_lookup[1]) >= courts or not available_courts:
            break
        donor_matches = list(round_lookup.get(donor_round, []))
        for match in donor_matches:
            if len(round_lookup[1]) >= courts or not available_courts:
                break
            if match.team1.name in used_names or match.team2.name in used_names:
                continue
            round_lookup[donor_round].remove(match)
            target_court = available_courts.pop(0)
            match.round_num = 1
            match.court = target_court
            # Time is applied later via apply_round_times().
            round_lookup[1].append(match)
            used_names.add(match.team1.name)
            used_names.add(match.team2.name)

    flattened: List[Match] = []
    for r in range(1, num_rounds + 1):
        flattened.extend(round_lookup.get(r, []))
    return sorted(flattened, key=lambda m: (m.round_num, m.court))


def compact_court_usage(matches: List[Match], num_rounds: int, courts: int,
                        band_overrides: Dict[str, Tuple[int, int]] | None = None) -> List[Match]:
    """Reassign courts per round so court numbers are packed (no holes).

    This is important when referee assignment depends on the *previous* match on the
    same court: internal blanks make that ambiguous. The output uses courts 1..K for
    each round (K = number of matches that round), leaving unused courts only at the end.
    """

    round_lookup: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds + 1)}
    for match in matches:
        round_lookup.setdefault(match.round_num, []).append(match)

    level_order = {'A': 0, 'B': 1, 'C': 2}
    last_active_round = max((m.round_num for m in matches), default=0)

    def order_within_level(r_matches: List[Match]) -> List[Match]:
        return sorted(r_matches, key=lambda m: (m.court, m.team1.name, m.team2.name))

    def assign_packed_prefix(rm: List[Match]) -> None:
        ordered = sorted(
            rm,
            key=lambda m: (
                level_order.get(m.team1.level, 9),
                m.court,
                m.team1.name,
                m.team2.name,
            ),
        )
        for idx, match in enumerate(ordered, start=1):
            match.court = idx

    def assign_final_round_prefer_bands(rm: List[Match]) -> None:
        a_matches = order_within_level([m for m in rm if m.team1.level == 'A'])
        b_matches = order_within_level([m for m in rm if m.team1.level == 'B'])
        c_matches = order_within_level([m for m in rm if m.team1.level == 'C'])
        other_matches = order_within_level([m for m in rm if m.team1.level not in {'A', 'B', 'C'}])

        # Final round has no “next round”, so we can prioritize readability:
        # A → left courts, C → right courts, B → near the center.
        available = list(range(1, courts + 1))
        assigned: Dict[int, Match] = {}

        def take_left(n: int) -> List[int]:
            nonlocal available
            picked = available[:n]
            available = available[n:]
            return picked

        def take_right(n: int) -> List[int]:
            nonlocal available
            picked = available[-n:] if n > 0 else []
            available = available[:-n] if n > 0 else available
            return picked

        def take_middle(n: int) -> List[int]:
            nonlocal available
            if n <= 0:
                return []
            mid = (courts + 1) / 2.0
            available_sorted = sorted(available, key=lambda c: (abs(c - mid), c))
            picked = available_sorted[:n]
            picked_set = set(picked)
            available = [c for c in available if c not in picked_set]
            return sorted(picked)

        for match, court in zip(a_matches, take_left(len(a_matches))):
            assigned[court] = match
        for match, court in zip(c_matches, take_right(len(c_matches))):
            assigned[court] = match

        b_courts = take_middle(len(b_matches))
        for match, court in zip(b_matches, b_courts):
            assigned[court] = match

        # Any remaining (unexpected) levels: fill remaining courts from left.
        for match, court in zip(other_matches, available):
            assigned[court] = match

        for court, match in assigned.items():
            match.court = court

    for round_num in range(1, num_rounds + 1):
        rm = round_lookup.get(round_num, [])
        if not rm:
            continue
        if round_num == last_active_round and len(rm) < courts:
            assign_final_round_prefer_bands(rm)
        else:
            assign_packed_prefix(rm)

    return sorted(matches, key=lambda m: (m.round_num, m.court))


def detect_collisions(matches: List[Match]) -> Dict[Tuple[int,int], List[Match]]:
    slot_map: Dict[Tuple[int,int], List[Match]] = {}
    for m in matches:
        slot_map.setdefault((m.round_num, m.court), []).append(m)
    collisions = {k:v for k,v in slot_map.items() if len(v) > 1}
    return collisions

def repair_collisions(matches: List[Match], num_rounds: int, courts: int) -> List[Match]:
    # 重複 (round,court) の試合を同ラウンド空きコートへ、なければ近い後続ラウンドへ移動
    collisions = detect_collisions(matches)
    if not collisions:
        return matches
    # ラウンド毎使用コートと出場チーム収集
    round_used_courts: Dict[int, Set[int]] = {r:set() for r in range(1,num_rounds+1)}
    round_team_names: Dict[int, Set[str]] = {r:set() for r in range(1,num_rounds+1)}
    for m in matches:
        round_used_courts[m.round_num].add(m.court)
        round_team_names[m.round_num].add(m.team1.name)
        round_team_names[m.round_num].add(m.team2.name)
    # 衝突分を処理
    moved: List[Match] = []
    for (r,c), dup_list in collisions.items():
        # 先頭はそのまま、2件目以降を動かす
        for m in dup_list[1:]:
            # 同ラウンド空きコート試行
            placed = False
            for alt_c in range(1, courts+1):
                if alt_c in round_used_courts[r]:
                    continue
                # チーム重複防止
                names = round_team_names[r]
                if m.team1.name in names or m.team2.name in names:
                    continue
                m.court = alt_c
                round_used_courts[r].add(alt_c)
                placed = True
                break
            if placed:
                continue
            # 後続ラウンド探索
            for r2 in range(r+1, num_rounds+1):
                names2 = round_team_names[r2]
                # チームすでにそのラウンド出場しているなら不可
                if m.team1.name in names2 or m.team2.name in names2:
                    continue
                # 空きコート検索
                for alt_c in range(1, courts+1):
                    if alt_c in round_used_courts[r2]:
                        continue
                    # 移動
                    m.round_num = r2
                    m.court = alt_c
                    # 時刻再計算
                    m.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r2-1))
                    round_used_courts[r2].add(alt_c)
                    round_team_names[r2].add(m.team1.name)
                    round_team_names[r2].add(m.team2.name)
                    moved.append(m)
                    placed = True
                    break
                if placed:
                    break
            if not placed:
                # 前方探索 (例外的) - なるべく近い前ラウンドへ
                for r2 in range(r-1, 0, -1):
                    names2 = round_team_names[r2]
                    if m.team1.name in names2 or m.team2.name in names2:
                        continue
                    for alt_c in range(1, courts+1):
                        if alt_c in round_used_courts[r2]:
                            continue
                        m.round_num = r2
                        m.court = alt_c
                        m.start_time = datetime(2025, 11, 26, 12, 50) + timedelta(minutes=13*(r2-1))
                        round_used_courts[r2].add(alt_c)
                        round_team_names[r2].add(m.team1.name); round_team_names[r2].add(m.team2.name)
                        moved.append(m)
                        placed = True
                        break
                    if placed:
                        break
            # ここで placed=False なら修復失敗 → そのまま保持
    # 再衝突確認 (無限ループ回避で 1 回のみ)
    return sorted(matches, key=lambda m: (m.round_num, m.court))


def normalize_round_capacity(matches: List[Match], num_rounds: int, courts: int) -> List[Match]:
    """Ensure each round has at most `courts` matches and court numbers stay within 1..courts.

    When collision repair fails to fully resolve overlaps, a round can temporarily exceed
    court capacity. Downstream court compaction must never "create" court 16 etc; instead,
    we move overflow matches into rounds that still have capacity.
    """

    if not matches:
        return matches

    base_time = datetime(2025, 11, 26, 12, 50)

    round_matches: Dict[int, List[Match]] = {r: [] for r in range(1, num_rounds + 1)}
    round_names: Dict[int, Set[str]] = {r: set() for r in range(1, num_rounds + 1)}

    for m in matches:
        r = min(max(1, m.round_num), num_rounds)
        m.round_num = r
        round_matches[r].append(m)
        round_names[r].add(m.team1.name)
        round_names[r].add(m.team2.name)

    def place_into_round(match: Match, target_round: int) -> bool:
        if len(round_matches[target_round]) >= courts:
            return False
        names = round_names[target_round]
        if match.team1.name in names or match.team2.name in names:
            return False
        match.round_num = target_round
        match.start_time = base_time + timedelta(minutes=13 * (target_round - 1))
        round_matches[target_round].append(match)
        names.add(match.team1.name)
        names.add(match.team2.name)
        return True

    # 1) Move overflow matches out of overfull rounds.
    for r in range(1, num_rounds + 1):
        while len(round_matches[r]) > courts:
            overflow = sorted(round_matches[r], key=lambda m: (m.court, m.team1.name, m.team2.name))[-1]
            round_matches[r].remove(overflow)
            round_names[r].discard(overflow.team1.name)
            round_names[r].discard(overflow.team2.name)

            moved = False
            for r2 in range(r + 1, num_rounds + 1):
                if place_into_round(overflow, r2):
                    moved = True
                    break
            if not moved:
                for r2 in range(r - 1, 0, -1):
                    if place_into_round(overflow, r2):
                        moved = True
                        break
            if not moved:
                # Put it back if we truly cannot place it anywhere.
                round_matches[r].append(overflow)
                round_names[r].add(overflow.team1.name)
                round_names[r].add(overflow.team2.name)
                break

    # 2) Clamp any out-of-range court numbers and remove duplicates within each round.
    flattened: List[Match] = []
    for r in range(1, num_rounds + 1):
        rm = round_matches[r]
        if not rm:
            continue
        # Reassign courts to a unique packed set within 1..courts.
        ordered = sorted(rm, key=lambda m: (m.court, m.team1.name, m.team2.name))
        for idx, m in enumerate(ordered, start=1):
            if idx > courts:
                # Should not happen after overflow move, but guard anyway.
                break
            m.court = idx
        flattened.extend(rm)

    return sorted(flattened, key=lambda m: (m.round_num, m.court))

def write_to_excel(matches: List[Match], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "対戦表"
    
    ws.append(["ラウンド", "コート", "チーム1", "チーム2", "開始時間"])
    
    for match in matches:
        ws.append([
            match.round_num,
            match.court,
            match.team1.name,
            match.team2.name,
            match.start_time.strftime("%H:%M")
        ])
    
    wb.save(output_path)

def write_to_excel_like_summary(
    matches: List[Match],
    teams: List[Team],
    output_path: str,
    allow_court_gaps: bool,
    num_rounds: int,
    courts: int,
    start_time_hhmm: str = DEFAULT_START_TIME_HHMM,
    round_minutes: int = DEFAULT_ROUND_MINUTES,
    excel_include_members: bool = False,
    excel_members_below: bool = False,
    excel_members_vlookup: bool = False,
    normalize_round_times: bool = True,
):
    wb = openpyxl.Workbook()
    # Ensure formulas (e.g., VLOOKUP for members) are recalculated when the file is opened in Excel.
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
    # Level-based fills
    level_fill = {
        'A': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # light yellow
        'B': PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),  # light green
        'C': PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid"),  # light blue
    }
    
    max_court = courts
    base_time = _base_datetime_from_hhmm(start_time_hhmm)
    round_duration = timedelta(minutes=int(round_minutes))

    # Normalize timestamps for the standard generator outputs.
    # When importing from an external "short list" sheet that already contains exact times,
    # we must preserve the existing Match.start_time.
    if normalize_round_times:
        apply_round_times(matches, start_time_hhmm=start_time_hhmm, round_minutes=int(round_minutes))

    # Sheet 1: 集計表
    ws1 = wb.active
    ws1.title = "集計表"
    ws1.append(["チーム1", "チーム2", "コート", "試合", "開始", "終了"])
    for match in matches:
        end_time = (match.start_time + round_duration).strftime("%H:%M")
        ws1.append([
            match.team1.name,
            match.team2.name,
            match.court,
            match.round_num,
            match.start_time.strftime("%H:%M"),
            end_time,
        ])
    
    # Sheet 2: 対戦表（氏名入り）
    # 2-row layout per round: row 1 = pair names, row 2 = member names (VLOOKUP).
    # This is the only editable match-table sheet.
    ws2 = wb.create_sheet("対戦表（氏名入り）")
    # Columns: 試合, 時間, then for each court two columns: コートN-チーム1, コートN-チーム2
    header = ["試合", "開始", "終了"]
    for c in range(1, max_court + 1):
        header += [f"コート{c}-チーム1", f"コート{c}-チーム2"]
    ws2.append(header)

    def render_excel_team_cell(team: Team | None) -> str:
        if not team:
            return ""
        if (not excel_members_below) and excel_include_members and team.members:
            return f"{team.name}\n{team.members}"
        return team.name

    def round_name_row_index(round_num: int) -> int:
        # Header is row 1.
        stride = 2 if excel_members_below else 1
        return 2 + (round_num - 1) * stride

    def members_vlookup_formula(*, key_cell_ref: str) -> str:
        # Use VLOOKUP for broad Excel compatibility.
        # Requires the workbook to contain the "ペア一覧" sheet with columns:
        # A=ペア名, B=選手名.
        return (
            f"=IF({key_cell_ref}=\"\",\"\",IFERROR(VLOOKUP({key_cell_ref},'ペア一覧'!$A:$B,2,FALSE),\"\"))"
        )

    def apply_level_fill_and_alignment(*, row_idx: int, court: int, fill: PatternFill | None) -> None:
        col_team1 = 3 + (court - 1) * 2 + 1
        col_team2 = col_team1 + 1
        if fill:
            ws2.cell(row=row_idx, column=col_team1).fill = fill
            ws2.cell(row=row_idx, column=col_team2).fill = fill
        ws2.cell(row=row_idx, column=col_team1).alignment = Alignment(wrap_text=True, vertical="top")
        ws2.cell(row=row_idx, column=col_team2).alignment = Alignment(wrap_text=True, vertical="top")

    for round_num in range(1, num_rounds + 1):
        # derive time from any match in round or compute from start
        round_start = None
        any_match = next((m for m in matches if m.round_num == round_num), None)
        if any_match:
            round_start = any_match.start_time.strftime("%H:%M")
        else:
            # fallback: compute from configured start time
            round_start = (base_time + timedelta(minutes=int(round_minutes) * (round_num - 1))).strftime("%H:%M")
        # compute end
        start_dt = datetime.strptime(round_start, "%H:%M")
        round_end = (start_dt + round_duration).strftime("%H:%M")
        if not excel_members_below:
            row = [round_num, round_start, round_end]
            for court in range(1, max_court + 1):
                match = next((m for m in matches if m.round_num == round_num and m.court == court), None)
                if match:
                    row += [render_excel_team_cell(match.team1), render_excel_team_cell(match.team2)]
                else:
                    row += ["", ""]
            ws2.append(row)
            row_idx_name = ws2.max_row

            # Apply fills per court cells based on level (use team1 level)
            for court in range(1, max_court + 1):
                match = next((m for m in matches if m.round_num == round_num and m.court == court), None)
                if not match:
                    continue
                fill = level_fill.get(match.team1.level)
                apply_level_fill_and_alignment(row_idx=row_idx_name, court=court, fill=fill)
        else:
            # Two-row layout: row 1 = pair names, row 2 = member names.
            row_names = [round_num, round_start, round_end]
            for court in range(1, max_court + 1):
                match = next((m for m in matches if m.round_num == round_num and m.court == court), None)
                if match:
                    row_names += [match.team1.name, match.team2.name]
                else:
                    row_names += ["", ""]
            ws2.append(row_names)
            row_idx_name = ws2.max_row

            row_members: list[Any] = ["", "", ""]
            for court in range(1, max_court + 1):
                match = next((m for m in matches if m.round_num == round_num and m.court == court), None)
                if not match:
                    row_members += ["", ""]
                    continue

                if excel_include_members:
                    row_members += [match.team1.members or "", match.team2.members or ""]
                elif excel_members_vlookup:
                    col_team1 = 3 + (court - 1) * 2 + 1
                    col_team2 = col_team1 + 1
                    key1 = f"{get_column_letter(col_team1)}{row_idx_name}"
                    key2 = f"{get_column_letter(col_team2)}{row_idx_name}"
                    row_members += [members_vlookup_formula(key_cell_ref=key1), members_vlookup_formula(key_cell_ref=key2)]
                else:
                    row_members += ["", ""]
            ws2.append(row_members)
            row_idx_members = ws2.max_row

            # Apply fills/alignment on both rows.
            for court in range(1, max_court + 1):
                match = next((m for m in matches if m.round_num == round_num and m.court == court), None)
                if not match:
                    continue
                fill = level_fill.get(match.team1.level)
                apply_level_fill_and_alignment(row_idx=row_idx_name, court=court, fill=fill)
                apply_level_fill_and_alignment(row_idx=row_idx_members, court=court, fill=fill)

    # Sheet 2b: 対戦表（ペア名のみ）
    # - Links from "対戦表（氏名入り）" (pair-name row) so edits propagate.
    # - Intended for copy/paste into the macro summary sheet.
    ws2_pairs = wb.create_sheet("対戦表（ペア名のみ）")
    ws2_pairs.append(header)

    def ref_or_blank(*, ref: str) -> str:
        # Excel shows 0 when referencing an empty cell (e.g., =A1 where A1 is blank).
        # For copy/paste workflows, blank is much easier than 0.
        return f"=IF({ref}=0,\"\",{ref})"

    for round_num in range(1, num_rounds + 1):
        # ws2: header row 1, then each round takes 2 rows (names, members) when excel_members_below=True
        src_row_names = round_name_row_index(round_num)
        dst_row = ws2_pairs.max_row + 1

        ws2_pairs.cell(row=dst_row, column=1).value = ref_or_blank(ref=f"'対戦表（氏名入り）'!A{src_row_names}")
        ws2_pairs.cell(row=dst_row, column=2).value = ref_or_blank(ref=f"'対戦表（氏名入り）'!B{src_row_names}")
        ws2_pairs.cell(row=dst_row, column=3).value = ref_or_blank(ref=f"'対戦表（氏名入り）'!C{src_row_names}")

        for court in range(1, max_court + 1):
            col_team1 = 3 + (court - 1) * 2 + 1
            col_team2 = col_team1 + 1
            ref1 = f"'対戦表（氏名入り）'!{get_column_letter(col_team1)}{src_row_names}"
            ref2 = f"'対戦表（氏名入り）'!{get_column_letter(col_team2)}{src_row_names}"
            ws2_pairs.cell(row=dst_row, column=col_team1).value = ref_or_blank(ref=ref1)
            ws2_pairs.cell(row=dst_row, column=col_team2).value = ref_or_blank(ref=ref2)

            match = next((m for m in matches if m.round_num == round_num and m.court == court), None)
            if match:
                fill = level_fill.get(match.team1.level)
                # Reuse the same look as the main sheet.
                ws2_pairs.cell(row=dst_row, column=col_team1).alignment = Alignment(wrap_text=True, vertical="top")
                ws2_pairs.cell(row=dst_row, column=col_team2).alignment = Alignment(wrap_text=True, vertical="top")
                if fill:
                    ws2_pairs.cell(row=dst_row, column=col_team1).fill = fill
                    ws2_pairs.cell(row=dst_row, column=col_team2).fill = fill

    # Sheet 3: ペア一覧（試合数）
    ws3 = wb.create_sheet("ペア一覧")
    ws3.append(["ペア名", "選手名", "レベル", "グループ", "試合数"])
    # Ensure stable order by level then group then name
    for t in sorted(teams, key=lambda x: (x.level, x.group, x.name)):
        ws3.append([t.name, t.members, t.level, t.group, t.matches])

    # Sheet 4: 全対戦リスト（両視点）
    ws4 = wb.create_sheet("全対戦リスト")
    ws4.append(["試合", "コート", "開始", "終了", "ペア名", "選手名", "相手ペア名", "相手選手名"])
    short_entries: list[dict[str, Any]] = []
    for match in matches:
        time_str = match.start_time.strftime("%H:%M")
        end_str = (match.start_time + round_duration).strftime("%H:%M")
        rows = [
            (match.team1, match.team2),
            (match.team2, match.team1),
        ]
        for cur, opp in rows:
            ws4.append([
                match.round_num,
                match.court,
                time_str,
                end_str,
                cur.name,
                cur.members,
                opp.name,
                opp.members,
            ])
            short_entries.append({
                "round": match.round_num,
                "court": match.court,
                "time": time_str,
                "pair": cur.name,
                "members": cur.members,
                "opp": opp.name,
                "opp_members": opp.members,
            })

    short_headers = ["試合", "コート", "時間", "ペア名", "選手名", "相手ペア名", "相手選手名"]
    ws4_team = wb.create_sheet("対戦一覧短縮（チーム順）")
    ws4_round = wb.create_sheet("対戦一覧短縮（試合順）")
    ws4_team.append(short_headers)
    ws4_round.append(short_headers)

    def normalize_name(value: str) -> str:
        return value.casefold() if isinstance(value, str) else ""

    for entry in sorted(short_entries, key=lambda e: (normalize_name(e["pair"]), e["round"], e["court"])):
        ws4_team.append([
            entry["round"],
            entry["court"],
            entry["time"],
            entry["pair"],
            entry["members"],
            entry["opp"],
            entry["opp_members"],
        ])

    for entry in sorted(short_entries, key=lambda e: (e["round"], e["court"], normalize_name(e["pair"]))):
        ws4_round.append([
            entry["round"],
            entry["court"],
            entry["time"],
            entry["pair"],
            entry["members"],
            entry["opp"],
            entry["opp_members"],
        ])
    
    # Styling helpers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # blue
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_sheet(ws):
        # headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # borders for all data cells
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in r:
                c.border = border

    def format_personal_schedule_sheet(ws):
        ws.freeze_panes = "A2"
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.alignment = header_alignment

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 2:
            return

        zebra_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        zebra_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        for row_idx in range(2, max_row + 1):
            row_fill = zebra_gray if row_idx % 2 == 0 else zebra_white
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.alignment = wrap_alignment

        if max_col >= 3:
            start_letter = get_column_letter(3)
            end_letter = get_column_letter(max_col)
            orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            formula = f'LEN(TRIM({start_letter}2))>0'
            ws.conditional_formatting.add(
                f"{start_letter}2:{end_letter}{max_row}",
                FormulaRule(formula=[formula], fill=orange_fill),
            )

        base_height = 17
        chars_per_line = 18
        for row_idx in range(2, max_row + 1):
            max_lines = 1
            for col_idx in range(1, max_col + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if not val:
                    continue
                text = str(val)
                explicit = text.count("\n") + 1
                approx = max(1, math.ceil(len(text) / chars_per_line))
                max_lines = max(max_lines, explicit, approx)
            ws.row_dimensions[row_idx].height = base_height * max_lines

    for ws in [ws1, ws2, ws2_pairs, ws3, ws4, ws4_team, ws4_round]:
        style_sheet(ws)

    if excel_members_below:
        thick = Side(style="thick", color="000000")
        dotted = Side(style="dotted", color="000000")

        # Keep header visible + lock the left time columns.
        # Freeze top row and columns A-C.
        ws2.freeze_panes = "D2"
        for cell in ws2[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Give header a bit more height so "コートXX-チームY" is readable.
        ws2.row_dimensions[1].height = 36

        def apply_team_block_border(*, name_row: int, members_row: int, col: int, opponent_side: str) -> None:
            """Border a (pair+members) block.

            - Outer edges are thick.
            - The opponent-facing edge is dotted (so 'vs' boundary is visually obvious).
            """
            if opponent_side not in ("left", "right"):
                raise ValueError("opponent_side must be 'left' or 'right'")

            top_cell = ws2.cell(row=name_row, column=col)
            bot_cell = ws2.cell(row=members_row, column=col)

            left_side = dotted if opponent_side == "left" else thick
            right_side = dotted if opponent_side == "right" else thick

            # Keep the inner border between name/members thin.
            top_cell.border = Border(
                left=left_side,
                right=right_side,
                top=thick,
                bottom=top_cell.border.bottom,
            )
            bot_cell.border = Border(
                left=left_side,
                right=right_side,
                top=bot_cell.border.top,
                bottom=thick,
            )

        for round_num in range(1, num_rounds + 1):
            name_row = round_name_row_index(round_num)
            members_row = name_row + 1
            for court in range(1, max_court + 1):
                col_team1 = 3 + (court - 1) * 2 + 1
                col_team2 = col_team1 + 1
                # Dotted line between opponents (team1 right edge / team2 left edge)
                apply_team_block_border(name_row=name_row, members_row=members_row, col=col_team1, opponent_side="right")
                apply_team_block_border(name_row=name_row, members_row=members_row, col=col_team2, opponent_side="left")

        # Also draw thick/dotted borders for the 1-row "対戦表（ペア名のみ）" sheet.
        ws2_pairs.freeze_panes = "D2"
        for cell in ws2_pairs[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws2_pairs.row_dimensions[1].height = 36

        def apply_vs_border_single_row(*, row_idx: int, col: int, opponent_side: str) -> None:
            if opponent_side not in ("left", "right"):
                raise ValueError("opponent_side must be 'left' or 'right'")
            cell = ws2_pairs.cell(row=row_idx, column=col)
            left_side = dotted if opponent_side == "left" else thick
            right_side = dotted if opponent_side == "right" else thick
            cell.border = Border(left=left_side, right=right_side, top=thick, bottom=thick)

        for round_num in range(1, num_rounds + 1):
            row_idx = 1 + round_num
            for court in range(1, max_court + 1):
                col_team1 = 3 + (court - 1) * 2 + 1
                col_team2 = col_team1 + 1
                apply_vs_border_single_row(row_idx=row_idx, col=col_team1, opponent_side="right")
                apply_vs_border_single_row(row_idx=row_idx, col=col_team2, opponent_side="left")

    # 表示試合数 vs 実試合数サマリ (対戦表シートに追加)
    def is_round_row(row_idx: int) -> bool:
        v = ws2.cell(row=row_idx, column=1).value
        if v is None:
            return False
        if isinstance(v, int):
            return True
        if isinstance(v, float) and v.is_integer():
            return True
        if isinstance(v, str):
            s = v.strip()
            return s.isdigit()
        return False

    displayed = 0
    for r in range(2, ws2.max_row + 1):
        if not is_round_row(r):
            continue
        # courts start column 4, step 2 (team1, team2)
        for col in range(4, 4 + max_court * 2, 2):
            t1 = ws2.cell(row=r, column=col).value
            t2 = ws2.cell(row=r, column=col + 1).value if col + 1 <= ws2.max_column else None
            if t1 and t2:
                displayed += 1
    expected = len(matches)
    collision_count = len(detect_collisions(matches)) if 'detect_collisions' in globals() else 0
    ws2.append(["表示試合数", displayed, "実試合数", expected, "不足", expected-displayed, "衝突スロット", collision_count])
    # style last row header cells
    for c in range(1,8):
        cell = ws2.cell(row=ws2.max_row, column=c)
        cell.border = border
        if c in (1,3,5,7):
            cell.font = Font(bold=True)

    # After saving base sheets, append analysis sheets if necessary (but openpyxl allows continue on same wb)
    # Reopen workbook to append more sheets if necessary (but openpyxl allows continue on same wb)
    # Build indices for analysis
    team_by_name = {t.name: t for t in teams}
    rounds_by_team: dict[str, List[int]] = {t.name: [] for t in teams}
    opponents_by_team: dict[str, Set[str]] = {t.name: set() for t in teams}
    groups_by_team: dict[str, Set[str]] = {t.name: set() for t in teams}
    same_group_counts: dict[str, int] = {t.name: 0 for t in teams}
    for m in matches:
        rounds_by_team[m.team1.name].append(m.round_num)
        rounds_by_team[m.team2.name].append(m.round_num)
        opponents_by_team[m.team1.name].add(m.team2.name)
        opponents_by_team[m.team2.name].add(m.team1.name)
        groups_by_team[m.team1.name].add(m.team2.group)
        groups_by_team[m.team2.name].add(m.team1.group)
        if m.team1.group == m.team2.group:
            same_group_counts[m.team1.name] += 1
            same_group_counts[m.team2.name] += 1

    def consec_stats(rounds: List[int]) -> tuple[int, int, float]:
        if not rounds:
            return 0, 0, 0.0
        sr = sorted(rounds)
        max_streak = 1
        streak = 1
        consecutive_pairs = 0
        gaps = []
        for i in range(1, len(sr)):
            gap = sr[i] - sr[i-1]
            gaps.append(gap)
            if gap == 1:
                consecutive_pairs += 1
                streak += 1
                max_streak = max(max_streak, streak)
            else:
                streak = 1
        avg_gap = (sum(gaps) / len(gaps)) if gaps else 0.0
        return max_streak, consecutive_pairs, avg_gap

    # Sheet 5: 対戦分散チェック
    ws5 = wb.create_sheet("対戦分散チェック")
    ws5.append(["ペア名", "選手名", "レベル", "グループ", "試合数", "ユニーク相手ペア数", "対戦グループ数", "同グループ回数", "理論最大グループ数", "対戦グループ一覧", "最大連戦数", "連戦回数", "平均間隔"])
    # Precompute level -> available opponent groups count (excluding own group if present)
    level_groups: dict[str, Set[str]] = {}
    for lvl in ['A', 'B', 'C']:
        level_groups[lvl] = set(t.group for t in teams if t.level == lvl)
    for t in sorted(teams, key=lambda x: (x.level, x.group, x.name)):
        mx, consec_cnt, avg_gap = consec_stats(rounds_by_team[t.name])
        total_groups_in_level = len(level_groups[t.level])
        theoretical_max_groups = max(0, total_groups_in_level - 1)  # cannot face own group
        ws5.append([
            t.name,
            t.members,
            t.level,
            t.group,
            len(rounds_by_team[t.name]),
            len(opponents_by_team[t.name]),
            len(groups_by_team[t.name]),
            same_group_counts.get(t.name, 0),
            theoretical_max_groups,
            ", ".join(sorted(groups_by_team[t.name])) ,
            mx,
            consec_cnt,
            round(avg_gap, 2),
        ])

    # Sheet 6: 個人スケジュール表（マトリクス）
    ws6 = wb.create_sheet("個人スケジュール表")
    header = ["ペア名", "選手名"] + [f"R{r}" for r in range(1, num_rounds + 1)]
    ws6.append(header)
    # Build quick lookup: (round, team_name) -> court/time
    rt_lookup: dict[tuple[int, str], tuple[int, str]] = {}
    for m in matches:
        t = m.start_time.strftime("%H:%M")
        rt_lookup[(m.round_num, m.team1.name)] = (m.court, t)
        rt_lookup[(m.round_num, m.team2.name)] = (m.court, t)
    # Also include opponent name per round
    # Build lookup for opponent per (round, team)
    opp_lookup: dict[tuple[int, str], str] = {}
    opp_members_lookup: dict[tuple[int, str], str] = {}
    for m in matches:
        opp_lookup[(m.round_num, m.team1.name)] = m.team2.name
        opp_lookup[(m.round_num, m.team2.name)] = m.team1.name
        opp_members_lookup[(m.round_num, m.team1.name)] = m.team2.members
        opp_members_lookup[(m.round_num, m.team2.name)] = m.team1.members
    def alphabetical_team_key(team: Team) -> str:
        return team.name.casefold() if isinstance(team.name, str) else ""

    for t in sorted(teams, key=alphabetical_team_key):
        row = [t.name, t.members]
        for r in range(1, num_rounds + 1):
            ct = rt_lookup.get((r, t.name))
            opp = opp_lookup.get((r, t.name))
            oppm = opp_members_lookup.get((r, t.name))
            if ct:
                if opp and oppm:
                    row.append(f"C{ct[0]}  {opp} | {oppm}")
                elif opp:
                    row.append(f"C{ct[0]}  {opp}")
                else:
                    row.append(f"C{ct[0]}")
            else:
                row.append("")
        ws6.append(row)

    # Precompute per-round counts for checklist metrics
    round_counts = Counter(m.round_num for m in matches)
    last_active_round = max(round_counts.keys(), default=0)
    front_gap_rounds = [r for r in range(1, last_active_round) if round_counts.get(r, 0) < courts]

    # Sheet 7: チェックリスト（絶対条件 + 分散の参考値）
    ws7 = wb.create_sheet("チェックリスト")
    ws7.append(["項目", "判定", "詳細"])

    # Quick visibility summary: max streak and whether any 3-in-a-row exists.
    all_team_streaks: dict[str, int] = {}
    for t in teams:
        mx, _consec_cnt, _avg_gap = consec_stats(rounds_by_team[t.name])
        all_team_streaks[t.name] = int(mx)
    global_max_streak = max(all_team_streaks.values(), default=0)
    triple_teams = [name for name, mx in all_team_streaks.items() if mx >= 3]
    triple_ok = len(triple_teams) == 0
    triple_detail = (
        f"3連戦なし / 最大連戦={global_max_streak}"
        if triple_ok
        else f"最大連戦={global_max_streak} / 該当 {len(triple_teams)} ペア: "
             + ", ".join(sorted(triple_teams)[:10])
             + (" ..." if len(triple_teams) > 10 else "")
    )
    ws7.append(["3連戦なし", "yes" if triple_ok else "no", triple_detail])
    row_triple = ws7.max_row

    ws7.append(["最大連戦数（全体）", str(global_max_streak), "目標=2（連戦2まで）"])
    row_max_streak = ws7.max_row

    # Make the 3連戦 summary row stand out.
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
    ng_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # light red
    summary_fill = ok_fill if triple_ok else ng_fill
    for col in range(1, 4):
        cell = ws7.cell(row=row_triple, column=col)
        cell.fill = summary_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    for col in range(1, 4):
        cell = ws7.cell(row=row_max_streak, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    # 1) 全ペアTARGET_MATCHES_PER_TEAM試合
    violators = [t.name for t in teams if len(rounds_by_team[t.name]) != TARGET_MATCHES_PER_TEAM]
    all_target_ok = len(violators) == 0
    detail1 = f"全員{TARGET_MATCHES_PER_TEAM}試合" if all_target_ok else f"未達 {len(violators)} ペア: " + ", ".join(violators[:10]) + (" ..." if len(violators) > 10 else "")
    ws7.append([f"全ペア{TARGET_MATCHES_PER_TEAM}試合", "yes" if all_target_ok else "no", detail1])
    # 2) 同じチームで対戦はないか（同グループ対戦の有無）
    same_group_match_count = sum(1 for m in matches if m.team1.group == m.team2.group)
    same_group_ok = same_group_match_count == 0
    detail2 = "同グループ対戦なし" if same_group_ok else f"同グループ対戦 {same_group_match_count} 試合"
    ws7.append(["同じチームで対戦はないか", "yes" if same_group_ok else "no", detail2])
    # 3) 同じクラス（A,B,C）で対戦されているか
    cross_level_violations = sum(1 for m in matches if m.team1.level != m.team2.level)
    level_ok = cross_level_violations == 0
    detail3 = "全試合同レベル内" if level_ok else f"異レベル対戦 {cross_level_violations} 試合"
    ws7.append(["同じクラス（A,B,C）で対戦されているか", "yes" if level_ok else "no", detail3])
    # 4) 空きコート許容設定（最終試合以降を除く）
    gap_detail = "前半空きなし" if not front_gap_rounds else \
        f"R{front_gap_rounds[0]} など {len(front_gap_rounds)} ラウンドで空き"
    gap_detail += " / 設定=" + ("許容" if allow_court_gaps else "禁止")
    ws7.append(["コート空き許容（最終試合以降除く）", "yes" if allow_court_gaps else "no", gap_detail])
    # 4) 分散（参考）: 平均・中央値の理論最大比
    ratios = []
    for t in teams:
        total_groups_in_level = len(set(tt.group for tt in teams if tt.level == t.level))
        theoretical = max(0, total_groups_in_level - 1)
        faced = len(groups_by_team[t.name])
        if theoretical > 0:
            ratios.append(faced / theoretical)
    if ratios:
        ratios_sorted = sorted(ratios)
        n = len(ratios_sorted)
        avg_ratio = sum(ratios_sorted) / n
        med_ratio = (ratios_sorted[n//2] if n % 2 == 1 else (ratios_sorted[n//2 - 1] + ratios_sorted[n//2]) / 2)
        detail4 = f"平均 {avg_ratio*100:.0f}% / 中央 {med_ratio*100:.0f}%（理論最大比）"
    else:
        detail4 = "-"
    ws7.append(["なるべく分散できているか", "なるべくyes", detail4])

    # Relaxable levers summary (impact small -> large) with attainment metrics
    ws7.append(["", "", ""])  # spacer
    ws7.append(["緩和可能項目（影響小→大）", "意味/設定", "指標/達成率"])
    # Preferred court band adherence
    preferred_courts_report = {
        'A': list(range(1, 4)),
        'B': list(range(5, 14)),
        'C': list(range(13, 16)),
    }
    total_ms = len(matches) if matches else 1
    pref_hits = sum(1 for m in matches if m.court in preferred_courts_report[m.team1.level])
    pref_rate = pref_hits / total_ms
    ws7.append(["コート帯固定", "現在=なるべく（全コート可、紙面は帯寄せ）", f"好み帯占有率 約{pref_rate*100:.0f}%（A:1-3, B:5-13, C:13-15）"])
    # Consecutive avoidance rate
    total_possible_adj = 0
    total_consec = 0
    for t in teams:
        mx, consec_cnt, avg_gap = consec_stats(rounds_by_team[t.name])
        k = len(rounds_by_team[t.name])
        total_possible_adj += max(0, k - 1)
        total_consec += consec_cnt
    avoid_rate = (1 - (total_consec / total_possible_adj)) if total_possible_adj > 0 else 1.0
    ws7.append(["連戦回避", "現在=参考（許容）", f"連戦回避率 約{avoid_rate*100:.0f}%（連戦総数/可能箇所）"])
    # Diversity attainment (reuse ratios from earlier calc)
    ratios = []
    for t in teams:
        total_groups_in_level = len(set(tt.group for tt in teams if tt.level == t.level))
        theoretical = max(0, total_groups_in_level - 1)
        faced = len(groups_by_team[t.name])
        if theoretical > 0:
            ratios.append(faced / theoretical)
    if ratios:
        ratios_sorted = sorted(ratios)
        n = len(ratios_sorted)
        avg_ratio2 = sum(ratios_sorted) / n
        med_ratio2 = (ratios_sorted[n//2] if n % 2 == 1 else (ratios_sorted[n//2 - 1] + ratios_sorted[n//2]) / 2)
        ws7.append(["対戦相手分散", "現在=参考（優先度低）", f"理論最大比 平均{avg_ratio2*100:.0f}% / 中央{med_ratio2*100:.0f}%"])
    else:
        ws7.append(["対戦相手分散", "現在=参考（優先度低）", "-"])

    ws7.append(["", "", ""])
    ws7.append(["対戦グループ数分布", "ペア数", "構成比"])
    group_distribution = Counter(len(groups_by_team[t.name]) for t in teams)
    total_pairs = len(teams) or 1
    for bucket in sorted(group_distribution):
        count = group_distribution[bucket]
        ws7.append([f"{bucket} グループ", count, f"{(count/total_pairs)*100:.0f}%"])
    counts_sorted = sorted((len(groups_by_team[t.name]) for t in teams))
    if counts_sorted:
        avg_groups = sum(counts_sorted) / total_pairs
        if total_pairs % 2 == 1:
            median_groups = counts_sorted[total_pairs // 2]
        else:
            median_groups = (counts_sorted[total_pairs // 2 - 1] + counts_sorted[total_pairs // 2]) / 2
        ws7.append(["平均/中央値", f"{avg_groups:.2f}", f"中央値 {median_groups:.2f}"])

    for ws in [ws5, ws6, ws7]:
        style_sheet(ws)
    format_personal_schedule_sheet(ws6)

    # Sheet 8: ラウンド統計 (利用率とレベル内訳)
    ws8 = wb.create_sheet("ラウンド統計")
    ws8.append(["ラウンド", "総試合", "A試合", "B試合", "C試合", "空きコート", "利用率%", "A帯コート", "B帯コート", "C帯コート"])
    # 帯定義 (報告用: 現在固定値)
    band_A = set(range(1,5))
    band_B = set(range(5,13))
    band_C = set(range(13,16))
    for r in range(1, num_rounds + 1):
        rm = [m for m in matches if m.round_num == r]
        total = len(rm)
        a_cnt = sum(1 for m in rm if m.team1.level == 'A')
        b_cnt = sum(1 for m in rm if m.team1.level == 'B')
        c_cnt = sum(1 for m in rm if m.team1.level == 'C')
        empty = courts - total
        usage = (total / courts * 100) if courts > 0 else 0.0
        a_band_used = sum(1 for m in rm if m.court in band_A and m.team1.level=='A')
        b_band_used = sum(1 for m in rm if m.court in band_B and m.team1.level=='B')
        c_band_used = sum(1 for m in rm if m.court in band_C and m.team1.level=='C')
        ws8.append([r, total, a_cnt, b_cnt, c_cnt, empty, round(usage,1), a_band_used, b_band_used, c_band_used])
    style_sheet(ws8)

    # 利用率サマリ行
    ws8.append([])
    total_matches = len(matches)
    ws8.append(["総試合数", total_matches, "期待", expected_total_matches(len(teams), TARGET_MATCHES_PER_TEAM), "空き合計", num_rounds*courts - total_matches])

    # 最終衝突チェック: もし不足があれば修復して再書き込み
    if expected - displayed > 0:
        print(f"対戦表表示不足 {expected - displayed} 試合検出 → 修復試行")
        matches = repair_collisions(matches, num_rounds, courts)
        print(f"After auto-repair, max round: {max(m.round_num for m in matches)}, min round: {min(m.round_num for m in matches)}")
        print(f"After auto-repair, max court: {max(m.court for m in matches)}, min court: {min(m.court for m in matches)}")
        # 対戦表シート再書き込み
        # まず既存の対戦表データをクリア (ヘッダー以外)
        for row in range(2, ws2.max_row+1):
            for col in range(1, ws2.max_column+1):
                ws2.cell(row=row, column=col).value = None
                # fillはデフォルトに戻す
                ws2.cell(row=row, column=col).fill = PatternFill()
        # 再書き込み
        for round_num in range(1, num_rounds + 1):
            # derive time from any match in round or compute from start
            round_start = None
            any_match = next((m for m in matches if m.round_num == round_num), None)
            if any_match:
                round_start = any_match.start_time.strftime("%H:%M")
            else:
                # fallback: compute from configured start time
                round_start = (base_time + timedelta(minutes=int(round_minutes) * (round_num - 1))).strftime("%H:%M")
            # compute end
            start_dt = datetime.strptime(round_start, "%H:%M")
            round_end = (start_dt + round_duration).strftime("%H:%M")
            row_name_idx = round_name_row_index(round_num)
            row_members_idx = row_name_idx + 1

            # Name row
            ws2.cell(row=row_name_idx, column=1).value = round_num
            ws2.cell(row=row_name_idx, column=2).value = round_start
            ws2.cell(row=row_name_idx, column=3).value = round_end

            # Members row (only for two-row layout)
            if excel_members_below:
                ws2.cell(row=row_members_idx, column=1).value = ""
                ws2.cell(row=row_members_idx, column=2).value = ""
                ws2.cell(row=row_members_idx, column=3).value = ""

            for court in range(1, max_court + 1):
                col_team1 = 3 + (court - 1) * 2 + 1
                col_team2 = col_team1 + 1
                match = next((m for m in matches if m.round_num == round_num and m.court == court), None)
                if match:
                    ws2.cell(row=row_name_idx, column=col_team1).value = match.team1.name
                    ws2.cell(row=row_name_idx, column=col_team2).value = match.team2.name
                else:
                    ws2.cell(row=row_name_idx, column=col_team1).value = ""
                    ws2.cell(row=row_name_idx, column=col_team2).value = ""

                if excel_members_below:
                    if not match:
                        ws2.cell(row=row_members_idx, column=col_team1).value = ""
                        ws2.cell(row=row_members_idx, column=col_team2).value = ""
                    else:
                        if excel_include_members:
                            ws2.cell(row=row_members_idx, column=col_team1).value = match.team1.members or ""
                            ws2.cell(row=row_members_idx, column=col_team2).value = match.team2.members or ""
                        elif excel_members_vlookup:
                            key1 = f"{get_column_letter(col_team1)}{row_name_idx}"
                            key2 = f"{get_column_letter(col_team2)}{row_name_idx}"
                            ws2.cell(row=row_members_idx, column=col_team1).value = members_vlookup_formula(key_cell_ref=key1)
                            ws2.cell(row=row_members_idx, column=col_team2).value = members_vlookup_formula(key_cell_ref=key2)
                        else:
                            ws2.cell(row=row_members_idx, column=col_team1).value = ""
                            ws2.cell(row=row_members_idx, column=col_team2).value = ""

                # 塗り直し
                if match:
                    fill = level_fill.get(match.team1.level)
                    apply_level_fill_and_alignment(row_idx=row_name_idx, court=court, fill=fill)
                    if excel_members_below:
                        apply_level_fill_and_alignment(row_idx=row_members_idx, court=court, fill=fill)
        # サマリ再計算
        displayed = 0
        for r in range(2, ws2.max_row + 1):
            if not is_round_row(r):
                continue
            for col in range(4, 4 + max_court * 2, 2):
                t1 = ws2.cell(row=r, column=col).value
                t2 = ws2.cell(row=r, column=col + 1).value if col + 1 <= ws2.max_column else None
                if t1 and t2:
                    displayed += 1
        collision_count = len(detect_collisions(matches))
        ws2.append(["表示試合数", displayed, "実試合数", expected, "不足", expected-displayed, "衝突スロット", collision_count])

    wb.save(output_path)


def write_personal_schedule_html(
    matches: List[Match],
    teams: List[Team],
    output_path: str,
    num_rounds: int = 23,
    courts: int = 15,
    html_passcode: str | None = None,
    start_time_hhmm: str = DEFAULT_START_TIME_HHMM,
    round_minutes: int = DEFAULT_ROUND_MINUTES,
    *,
    include_members: bool = True,
) -> None:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    header_rounds = [f"R{r}" for r in range(1, num_rounds + 1)]
    team_lookup = {t.name: t for t in teams}

    # Match count summary (per team) for human checks.
    match_counts_by_team: Counter[str] = Counter()
    for m in matches:
        if m.team1 and m.team1.name:
            match_counts_by_team[m.team1.name] += 1
        if m.team2 and m.team2.name:
            match_counts_by_team[m.team2.name] += 1

    team_names = [t.name for t in teams if t.name]
    match_counts = [int(match_counts_by_team.get(name, 0)) for name in team_names]
    match_count_min = min(match_counts) if match_counts else 0
    match_count_max = max(match_counts) if match_counts else 0
    match_count_uniform = bool(match_counts) and (match_count_min == match_count_max)

    def club_root(group_name: str | None) -> str:
        if not group_name:
            return ""
        return group_name[:-1] if group_name[-1] in "ABC" else group_name

    rt_lookup: dict[tuple[int, str], tuple[int, str]] = {}
    opp_lookup: dict[tuple[int, str], str] = {}
    opp_members_lookup: dict[tuple[int, str], str] = {}
    for m in matches:
        stamp = m.start_time.strftime("%H:%M")
        rt_lookup[(m.round_num, m.team1.name)] = (m.court, stamp)
        rt_lookup[(m.round_num, m.team2.name)] = (m.court, stamp)
        opp_lookup[(m.round_num, m.team1.name)] = m.team2.name
        opp_lookup[(m.round_num, m.team2.name)] = m.team1.name
        opp_members_lookup[(m.round_num, m.team1.name)] = m.team2.members
        opp_members_lookup[(m.round_num, m.team2.name)] = m.team1.members

    # Back-to-back (consecutive rounds) detection.
    rounds_by_team: dict[str, set[int]] = defaultdict(set)
    for (r, team_name) in rt_lookup.keys():
        if team_name:
            rounds_by_team[team_name].add(int(r))
    # We list only the *start* of a consecutive streak (so the 2nd match doesn't look like a start).
    back_to_back_starts: set[tuple[str, int]] = set()
    for team_name, rset in rounds_by_team.items():
        for r in rset:
            if (r + 1) in rset and (r - 1) not in rset:
                back_to_back_starts.add((team_name, r))

    def cell_value(team_name: str, rnd: int) -> str:
        slot = rt_lookup.get((rnd, team_name))
        if not slot:
            return ""
        opp = opp_lookup.get((rnd, team_name))
        opp_members = opp_members_lookup.get((rnd, team_name)) if include_members else ""
        if opp and opp_members:
            return f"C{slot[0]} {escape(opp)} / {escape(opp_members)}<br><small>{slot[1]}</small>"
        if opp:
            return f"C{slot[0]} {escape(opp)}<br><small>{slot[1]}</small>"
        return f"C{slot[0]}<br><small>{slot[1]}</small>"

    def alphabetical_team_key(team: Team) -> str:
        return team.name.casefold() if isinstance(team.name, str) else ""

    # Build 対戦表 rows
    matrix_headers = ["試合", "開始", "終了"] + [f"コート{c}" for c in range(1, courts + 1)]
    match_rows: list[dict[str, Any]] = []
    for round_num in range(1, num_rounds + 1):
        any_match = next((m for m in matches if m.round_num == round_num), None)
        if any_match:
            round_start = any_match.start_time.strftime("%H:%M")
        else:
            base = _base_datetime_from_hhmm(start_time_hhmm)
            round_start = (base + timedelta(minutes=int(round_minutes) * (round_num - 1))).strftime("%H:%M")
        start_dt = datetime.strptime(round_start, "%H:%M")
        round_end = (start_dt + timedelta(minutes=int(round_minutes))).strftime("%H:%M")
        row: list[tuple[str, bool]] = [(f"第{round_num}試合", False), (round_start, False), (round_end, False)]
        clubs_in_round: set[str] = set()
        teams_in_round: set[str] = set()
        for court in range(1, courts + 1):
            match = next((m for m in matches if m.round_num == round_num and m.court == court), None)
            if match:
                # iPhone等でヘッダー(コート番号)が見えなくなることがあるため、セル側にもコート番号を入れておく
                if include_members and (match.team1.members or match.team2.members):
                    t1m = escape(match.team1.members or "")
                    t2m = escape(match.team2.members or "")
                    row.append((
                        f"C{court} {escape(match.team1.name)} vs {escape(match.team2.name)}"
                        f"<br><small>{t1m} / {t2m}</small>",
                        True,
                    ))
                else:
                    row.append((f"C{court} {match.team1.name} vs {match.team2.name}", False))
                clubs_in_round.add(match.team1.group)
                clubs_in_round.add(match.team2.group)
                teams_in_round.add(match.team1.name)
                teams_in_round.add(match.team2.name)
            else:
                row.append(("", False))
        match_rows.append({
            "cells": row,
            "meta": {
                "clubs": sorted(c for c in clubs_in_round if c),
                "club_roots": sorted({root for root in (club_root(c) for c in clubs_in_round if c) if root}),
                "teams": sorted(t for t in teams_in_round if t),
            },
        })

    short_entries: list[dict[str, Any]] = []
    for match in matches:
        time_str = match.start_time.strftime("%H:%M")
        for cur, opp in ((match.team1, match.team2), (match.team2, match.team1)):
            club = team_lookup.get(cur.name).group if cur.name in team_lookup else ""
            short_entries.append({
                "round": match.round_num,
                "court": match.court,
                "time": time_str,
                "pair": cur.name,
                "members": cur.members,
                "opp": opp.name,
                "opp_members": opp.members,
                "club": club,
                "club_root": club_root(club),
            })

    def normalize_name(value: str) -> str:
        return value.casefold() if isinstance(value, str) else ""

    short_headers = ["試合", "コート", "時間", "ペア名"]
    if include_members:
        short_headers += ["選手名"]
    short_headers += ["相手ペア名"]
    if include_members:
        short_headers += ["相手選手名"]
    short_rows_team = [
        {
            "cells": [
                (f"第{entry['round']}試合", False),
                (str(entry["court"]), False),
                (entry["time"], False),
                (entry["pair"], False),
                *([(entry["members"], False)] if include_members else []),
                (entry["opp"], False),
                *([(entry["opp_members"], False)] if include_members else []),
            ],
            "meta": {
                "team": entry["pair"],
                "club": entry["club"],
                "club_root": entry["club_root"],
            },
        }
        for entry in sorted(short_entries, key=lambda e: (normalize_name(e["pair"]), e["round"], e["court"]))
    ]

    short_rows_round = [
        {
            "cells": [
                (f"第{entry['round']}試合", False),
                (str(entry["court"]), False),
                (entry["time"], False),
                (entry["pair"], False),
                *([(entry["members"], False)] if include_members else []),
                (entry["opp"], False),
                *([(entry["opp_members"], False)] if include_members else []),
            ],
            "meta": {
                "team": entry["pair"],
                "club": entry["club"],
                "club_root": entry["club_root"],
            },
        }
        for entry in sorted(short_entries, key=lambda e: (e["round"], e["court"], normalize_name(e["pair"])) )
    ]

    personal_headers = ["ペア名"]
    if include_members:
        personal_headers += ["選手名"]
    personal_headers += [*header_rounds]
    personal_rows: list[dict[str, Any]] = []
    for team in sorted(teams, key=alphabetical_team_key):
        row: list[tuple[str, bool]] = [
            (team.name, False),
            *([(team.members, False)] if include_members else []),
        ]
        for rnd in range(1, num_rounds + 1):
            val = cell_value(team.name, rnd)
            row.append((val, True))
        personal_rows.append({
            "cells": row,
            "meta": {
                "team": team.name,
                "club": team.group,
                "club_root": club_root(team.group),
            },
        })

    team_filter_options: list[tuple[str, str]] = []
    for team in sorted(teams, key=alphabetical_team_key):
        if include_members and team.members:
            label = f"{team.name}（{team.members}）"
        else:
            label = team.name
        team_filter_options.append((team.name, label))
    keyword_candidates = sorted({t.name for t in teams if t.name}, key=lambda n: n.casefold())
    club_options = sorted({t.group for t in teams if t.group}, key=lambda g: g.casefold())
    club_root_options = sorted({root for root in (club_root(t.group) for t in teams if t.group) if root}, key=lambda g: g.casefold())

    # For client-side highlighting when filtering by club/area.
    club_to_teams: dict[str, list[str]] = {}
    root_to_teams: dict[str, list[str]] = {}
    for t in teams:
        if not t.name:
            continue
        if t.group:
            club_to_teams.setdefault(t.group, []).append(t.name)
            root = club_root(t.group)
            if root:
                root_to_teams.setdefault(root, []).append(t.name)

    def escape_attr(value: str) -> str:
        return escape(value, quote=True)

    # Pair count summary (club-root x level)
    summary_counts: dict[str, dict[str, int]] = {}
    for t in teams:
        root = club_root(t.group) or ""
        lvl = (t.level or "").strip().upper()
        bucket = summary_counts.setdefault(root, {"A": 0, "B": 0, "C": 0, "?": 0, "total": 0})
        if lvl in ("A", "B", "C"):
            bucket[lvl] += 1
        else:
            bucket["?"] += 1
        bucket["total"] += 1

    show_unknown = any(v.get("?", 0) > 0 for v in summary_counts.values())
    summary_headers = ["エリア", "A", "B", "C"] + (["不明"] if show_unknown else []) + ["合計"]
    summary_rows: list[list[str]] = []
    grand = {"A": 0, "B": 0, "C": 0, "?": 0, "total": 0}
    for root in sorted(summary_counts.keys(), key=lambda s: (s or "").casefold()):
        label = root or "(未分類)"
        c = summary_counts[root]
        grand["A"] += int(c.get("A", 0) or 0)
        grand["B"] += int(c.get("B", 0) or 0)
        grand["C"] += int(c.get("C", 0) or 0)
        grand["?"] += int(c.get("?", 0) or 0)
        grand["total"] += int(c.get("total", 0) or 0)
        row = [label, str(c.get("A", 0)), str(c.get("B", 0)), str(c.get("C", 0))]
        if show_unknown:
            row.append(str(c.get("?", 0)))
        row.append(str(c.get("total", 0)))
        summary_rows.append(row)

    # Grand total row
    total_row = ["合計", str(grand["A"]), str(grand["B"]), str(grand["C"])]
    if show_unknown:
        total_row.append(str(grand["?"]))
    total_row.append(str(grand["total"]))
    summary_rows.append(total_row)

    def render_summary_table(fh) -> None:
        fh.write("<section class='table-block' id='pair-summary-section'>")
        fh.write("<h2>ペア数サマリー</h2>")
        fh.write("<div class='wrap wrap-pair-summary'><table id='pair-summary'><thead><tr>")
        for idx, h in enumerate(summary_headers):
            cls = " class='summary-total-col'" if idx == (len(summary_headers) - 1) else ""
            fh.write(f"<th{cls}>{escape(h)}</th>")
        fh.write("</tr></thead><tbody>")
        for row in summary_rows:
            is_total = bool(row) and row[0] == "合計"
            tr_cls = " class='summary-total-row'" if is_total else ""
            fh.write(f"<tr{tr_cls}>")
            for idx, value in enumerate(row):
                td_classes: list[str] = []
                if idx == (len(row) - 1):
                    td_classes.append("summary-total-col")
                if is_total:
                    td_classes.append("summary-total")
                td_cls = f" class='{' '.join(td_classes)}'" if td_classes else ""
                fh.write(f"<td{td_cls}>{escape(value)}</td>")
            fh.write("</tr>")
        fh.write("</tbody></table></div></section>")

    def render_match_count_summary(fh) -> None:
        fh.write("<section class='table-block' id='match-count-summary-section'>")
        fh.write("<h2>試合数サマリー</h2>")
        fh.write("<div style='font-size: 13px; line-height: 1.45;'>")
        fh.write(
            f"<div><span class='team-meta'>総試合数</span>: {len(matches)}（対戦数）</div>"
        )
        if match_count_uniform:
            fh.write(
                f"<div><span class='team-meta'>各ペア試合数（実績）</span>: 全{len(team_names)}ペア {match_count_min}試合</div>"
            )
        else:
            fh.write(
                f"<div><span class='team-meta'>各ペア試合数（実績）</span>: min={match_count_min} / max={match_count_max}（全{len(team_names)}ペア）</div>"
            )
        fh.write(
            "<div><small>注: 1試合は2ペアが出場するため、ペア別試合数を合計すると総試合数の2倍になります。</small></div>"
        )
        fh.write("</div>")
        fh.write("</section>")

    def render_back_to_back_table(fh) -> None:
        # List matches where either side *starts* a back-to-back sequence.
        headers = ["試合", "コート", "時刻", "対戦", "連戦メモ"]
        rows: list[dict[str, Any]] = []
        for m in matches:
            t1 = m.team1.name
            t2 = m.team2.name
            r = int(m.round_num)
            flag1 = (t1, r) in back_to_back_starts
            flag2 = (t2, r) in back_to_back_starts
            if not (flag1 or flag2):
                continue

            memo_parts: list[str] = []
            for team_name, is_start in ((t1, flag1), (t2, flag2)):
                if not is_start:
                    continue
                nxt = rt_lookup.get((r + 1, team_name))
                nxt_opp = opp_lookup.get((r + 1, team_name))
                if nxt and nxt_opp:
                    memo_parts.append(
                        f"{team_name}は次のR{r+1}も C{nxt[0]} vs {nxt_opp}（{nxt[1]}）で連戦"
                    )
                elif nxt:
                    memo_parts.append(
                        f"{team_name}は次のR{r+1}も C{nxt[0]}（{nxt[1]}）で連戦"
                    )
                else:
                    memo_parts.append(
                        f"{team_name}は次のR{r+1}も試合で連戦"
                    )
            memo = " / ".join(memo_parts)

            if include_members and (m.team1.members or m.team2.members):
                t1m = escape(m.team1.members or "")
                t2m = escape(m.team2.members or "")
                versus = (
                    f"{escape(t1)} vs {escape(t2)}"
                    f"<br><small>{t1m} / {t2m}</small>"
                )
                versus_cell = (versus, True)
            else:
                versus_cell = (f"{t1} vs {t2}", False)

            club1 = team_lookup.get(t1).group if t1 in team_lookup and team_lookup.get(t1) else ""
            club2 = team_lookup.get(t2).group if t2 in team_lookup and team_lookup.get(t2) else ""
            rows.append(
                {
                    "cells": [
                        (f"R{r}", False),
                        (f"C{m.court}", False),
                        (m.start_time.strftime("%H:%M"), False),
                        versus_cell,
                        (memo, False),
                    ],
                    "meta": {
                        "teams": [t1, t2],
                        "clubs": [c for c in (club1, club2) if c],
                        "club_roots": sorted({root for root in (club_root(club1), club_root(club2)) if root}),
                    },
                }
            )

        fh.write("<section class='table-block' id='back-to-back-notes-section'>")
        fh.write("<h2>連戦が絡む試合（要確認）</h2>")
        fh.write(
            "<div style='font-size: 13px; line-height: 1.45; margin-bottom: 8px;'>"
            "<div><small>ここは『連戦の開始試合』だけを抽出しています（2試合目は重複表示しません）。連戦のペアは直後の審判が厳しいことがあります。負けた側が連続で審判になるケースもあるので、当事者同士で事前に把握して調整してください。</small></div>"
            "</div>"
        )
        fh.write("</section>")

        if not rows:
            # No consecutive-round matches.
            return

        render_table(
            fh,
            "back-to-back",
            "連戦が絡む試合（一覧）",
            headers,
            rows,
            enable_team_filter=True,
            team_options=team_filter_options,
            enable_club_filter=True,
            club_options=club_options,
            enable_club_root_filter=True,
            club_root_options=club_root_options,
            keyword_options=keyword_candidates,
            sticky_columns={0: ["round-col"], 1: ["court-col"]},
        )

    def render_table(
        fh,
        section_id: str,
        title: str,
        headers: List[str],
        rows: List[dict[str, Any]],
        *,
        enable_team_filter: bool = False,
        team_options: List[tuple[str, str]] | None = None,
        enable_club_filter: bool = False,
        club_options: List[str] | None = None,
        enable_club_root_filter: bool = False,
        club_root_options: List[str] | None = None,
        keyword_options: List[str] | None = None,
        sticky_columns: Dict[int, List[str]] | None = None,
    ) -> None:
        sticky_map = sticky_columns or {}
        fh.write(f"<section class='table-block' id='{escape(section_id)}-section'>")
        fh.write(f"<h2>{escape(title)}</h2>")
        fh.write(f"<div class='filter-bar' data-table='{escape(section_id)}'>")
        datalist_id = f"{section_id}-keywords"
        fh.write(f"<label>検索:<input type='search' placeholder='キーワード' data-role='search' list='{escape(datalist_id)}'></label>")
        fh.write("<label>列:<select data-role='column'><option value='all'>全列</option>")
        for idx, header in enumerate(headers):
            fh.write(f"<option value='{idx}'>{escape(header)}</option>")
        fh.write("</select></label>")
        if enable_team_filter and team_options:
            fh.write("<label>ペア:<select data-role='team'><option value='all'>全ペア</option>")
            for value, label in team_options:
                fh.write(f"<option value='{escape_attr(value)}'>{escape(label)}</option>")
            fh.write("</select></label>")
        if enable_club_filter and club_options:
            fh.write("<label>クラブ:<select data-role='club'><option value='all'>全クラブ</option>")
            for club in club_options:
                fh.write(f"<option value='{escape_attr(club)}'>{escape(club)}</option>")
            fh.write("</select></label>")
        if enable_club_root_filter and club_root_options:
            fh.write("<label>エリア:<select data-role='club-root'><option value='all'>全エリア</option>")
            for root in club_root_options:
                fh.write(f"<option value='{escape_attr(root)}'>{escape(root)}</option>")
            fh.write("</select></label>")
        fh.write("</div>")
        if keyword_options:
            fh.write(f"<datalist id='{escape(datalist_id)}'>")
            for option in keyword_options:
                fh.write(f"<option value='{escape(option)}'></option>")
            fh.write("</datalist>")
        fh.write("<div class='wrap wrap-" + escape(section_id) + "'><table id='" + escape(section_id) + "'><thead><tr>")
        for idx, header in enumerate(headers):
            classes: list[str] = []
            if idx in sticky_map:
                classes.append("sticky-col")
                classes.extend(sticky_map[idx])
            class_attr = f" class='{' '.join(classes)}'" if classes else ""
            fh.write(f"<th{class_attr}>{escape(header)}</th>")
        fh.write("</tr></thead><tbody>")
        for row in rows:
            is_dict = isinstance(row, dict)
            meta = row.get("meta", {}) if is_dict else {}
            attrs: list[tuple[str, str]] = []
            team_value = meta.get("team")
            if team_value:
                attrs.append(("data-team", team_value))
            teams_value = meta.get("teams")
            if teams_value:
                joined_teams = "|".join(sorted({t for t in teams_value if t}))
                if joined_teams:
                    attrs.append(("data-teams", joined_teams))
            club_value = meta.get("club")
            if club_value:
                attrs.append(("data-club", club_value))
            clubs_value = meta.get("clubs")
            if clubs_value:
                joined = "|".join(sorted({c for c in clubs_value if c}))
                if joined:
                    attrs.append(("data-clubs", joined))
            club_root_value = meta.get("club_root")
            if club_root_value:
                attrs.append(("data-club-root", club_root_value))
            club_roots_value = meta.get("club_roots")
            if club_roots_value:
                joined_roots = "|".join(sorted({c for c in club_roots_value if c}))
                if joined_roots:
                    attrs.append(("data-club-roots", joined_roots))
            attr_str = " ".join(f"{name}='{escape_attr(val)}'" for name, val in attrs)
            if attr_str:
                fh.write(f"<tr {attr_str}>")
            else:
                fh.write("<tr>")
            cells = row.get("cells", []) if is_dict else row
            for idx, (value, raw) in enumerate(cells):
                classes: list[str] = []
                if idx in sticky_map:
                    classes.append("sticky-col")
                    classes.extend(sticky_map[idx])
                class_attr = f" class='{' '.join(classes)}'" if classes else ""
                if raw:
                    fh.write(f"<td{class_attr}>{value}</td>")
                else:
                    fh.write(f"<td{class_attr}>{escape(str(value) if value is not None else '')}</td>")
            fh.write("</tr>")
        fh.write("</tbody></table></div></section>")

    pass_hash = hashlib.sha256(html_passcode.encode("utf-8")).hexdigest() if html_passcode else ""

    with path.open("w", encoding="utf-8") as fh:
        fh.write("<!DOCTYPE html><html lang='ja'><head><meta charset='utf-8'>")
        fh.write("<meta name='viewport' content='width=device-width, initial-scale=1'>")
        fh.write("<title>対戦表＋個人スケジュール</title>")
        fh.write(
            """
<style>
:root {
    --match-col-width: 95px;
    --short-round-width: 90px;
    --short-court-width: 52px;
    --short-member-width: 160px;
    --personal-team-width: 180px;
    --personal-member-width: 220px;
    --personal-round-width: 120px;
}
body {font-family: system-ui, sans-serif; margin: 16px; color: #111;}
h1 {margin-bottom: 8px;}
h2 {margin: 32px 0 8px;}
table {border-collapse: collapse; width: 100%; font-size: 12px; min-width: 620px;}
th, td {border: 1px solid #ddd; padding: 4px 6px; vertical-align: top; background-clip: padding-box;}
th {position: -webkit-sticky; position: sticky; top: 0; background: #f7f7f7; z-index: 4;}
th.sticky-col {z-index: 6;}
td.sticky-col {z-index: 2;}
tbody tr:nth-child(odd) {background: #fafafa;}
tbody tr:nth-child(odd) td.sticky-col {background: #fdfdfd;}
tbody tr:nth-child(even) td.sticky-col {background: #fff;}
small {color: #666;}
.team-meta {font-weight: 600;}
.wrap {overflow-x: auto; -webkit-overflow-scrolling: touch; max-width: 100%; box-shadow: inset 0 0 0 1px #f0f0f0; border-radius: 4px; position: relative; isolation: isolate;}
.wrap-personal-schedule {overflow: auto; max-height: 72vh;}
.wrap table {width: max-content;}
.filter-bar {display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 8px;}
.filter-bar label {display: flex; align-items: center; gap: 4px; font-size: 13px;}
input[type='search'] {padding: 4px 6px; font-size: 13px; border: 1px solid #bbb; border-radius: 4px;}

/* Pair summary: highlight totals */
#pair-summary td.summary-total-col, #pair-summary th.summary-total-col {font-weight: 700; background: #f7f7f7;}
#pair-summary tr.summary-total-row td {font-weight: 700; background: #f7f7f7;}
select {padding: 4px; font-size: 13px; border: 1px solid #bbb; border-radius: 4px;}
section.table-block {margin-bottom: 40px;}
datalist option {font-size: 12px;}
tr.matched td {background: #fff8d5;}
td.cell-hit {background: #ffe08a !important; box-shadow: inset 0 0 0 2px rgba(0,0,0,0.18);}
mark {background: #ffe08a; padding: 0 2px; border-radius: 2px;}
#lock-overlay.lock-overlay {position: fixed; inset: 0; background: rgba(0,0,0,0.55); display: flex; align-items: center; justify-content: center; z-index: 9999; padding: 16px;}
.lock-card {background: #fff; color: #111; border-radius: 10px; padding: 16px; max-width: 460px; width: 100%; box-shadow: 0 10px 30px rgba(0,0,0,0.3);}
.lock-title {font-weight: 700; font-size: 16px; margin-bottom: 8px;}
.lock-note {font-size: 12px; color: #555; margin-bottom: 12px; line-height: 1.4;}
.lock-label {display: block; font-size: 13px; margin-bottom: 10px;}
.lock-label input {width: 100%; margin-top: 6px; padding: 10px; border: 1px solid #bbb; border-radius: 6px; font-size: 14px;}
.lock-actions {display: flex; align-items: center; gap: 10px;}
.lock-actions button {padding: 10px 12px; border: 1px solid #333; background: #111; color: #fff; border-radius: 6px; font-size: 14px;}
.lock-error {font-size: 12px; color: #b00020;}
body.locked > :not(#lock-overlay) {filter: blur(2px); pointer-events: none; user-select: none;}
#match-matrix th.match-col,
#short-team th.round-col, #short-round th.round-col,
#short-team th.court-col, #short-round th.court-col,
#personal-schedule th.team-col, #personal-schedule th.member-col {
    background: #f0f0f0;
}
.sticky-col {position: -webkit-sticky; position: sticky; box-shadow: inset -1px 0 0 rgba(0,0,0,0.08);}
td.sticky-col {background: #fff;}
#short-team, #short-round {table-layout: fixed;}
#match-matrix {table-layout: auto;}
#personal-schedule {table-layout: fixed;}
#match-matrix td:not(.sticky-col), #short-team td:not(.sticky-col), #short-round td:not(.sticky-col), #personal-schedule td:not(.sticky-col) {white-space: normal; word-break: break-word; overflow-wrap: break-word;}
#match-matrix .match-col, #short-team .round-col, #short-round .round-col, #short-team .court-col, #short-round .court-col {white-space: nowrap;}
#match-matrix .match-col {left: 0; min-width: var(--match-col-width); max-width: var(--match-col-width);}
#short-team .round-col, #short-round .round-col {left: 0; min-width: var(--short-round-width); max-width: var(--short-round-width);}
#short-team .court-col, #short-round .court-col {left: var(--short-round-width); min-width: var(--short-court-width); max-width: var(--short-court-width);}
#personal-schedule .team-col {left: 0; width: var(--personal-team-width); min-width: var(--personal-team-width); max-width: var(--personal-team-width);}
#personal-schedule .member-col {left: var(--personal-team-width); width: var(--personal-member-width); min-width: var(--personal-member-width); max-width: var(--personal-member-width);}
#personal-schedule .team-col, #personal-schedule .member-col {white-space: normal; word-break: break-word; overflow-wrap: anywhere; line-height: 1.15;}
#personal-schedule th:not(.sticky-col), #personal-schedule td:not(.sticky-col) {min-width: var(--personal-round-width); max-width: var(--personal-round-width);}
#match-matrix td small {display: block; color: #333; line-height: 1.1;}

/*
 iOS Safari: position: sticky inside overflow scrolling can become unstable during pinch-zoom
 when -webkit-overflow-scrolling: touch is enabled. Prefer stability over momentum.
*/
@supports (-webkit-touch-callout: none) {
    .wrap { -webkit-overflow-scrolling: auto; }
}

@media (max-width: 768px) {
    :root {
        --match-col-width: 84px;
        --short-round-width: 52px;
        --short-court-width: 28px;
        --short-member-width: 72px;
        --personal-team-width: 52px;
        --personal-member-width: 84px;
        --personal-round-width: 44px;
    }
    body {margin: 12px;}
    table {font-size: 11px; min-width: 420px;}
    th, td {padding: 3px 4px;}
    .filter-bar {gap: 8px;}
    #personal-schedule td.team-col, #personal-schedule td.member-col,
    #personal-schedule th.team-col, #personal-schedule th.member-col {
        padding: 2px 3px;
        font-size: 11px;
        line-height: 1.1;
    }

    /* iPhone: keep columns compact but readable; horizontal scrolling remains available */
    #match-matrix th:not(.sticky-col), #match-matrix td:not(.sticky-col) {min-width: 80px;}

    /* Short list: let labels wrap so we can shrink columns more */
    #short-team .round-col, #short-round .round-col {white-space: normal;}
    #short-team .court-col, #short-round .court-col {white-space: normal;}
    #short-team th.round-col, #short-round th.round-col,
    #short-team th.court-col, #short-round th.court-col {line-height: 1.05; padding: 2px 3px;}
    #short-team td.round-col, #short-round td.round-col,
    #short-team td.court-col, #short-round td.court-col {line-height: 1.05; padding: 2px 3px;}

    /* Short list: member-name columns should NOT decide the table width */
    #short-team th:nth-child(5), #short-team td:nth-child(5),
    #short-team th:nth-child(7), #short-team td:nth-child(7),
    #short-round th:nth-child(5), #short-round td:nth-child(5),
    #short-round th:nth-child(7), #short-round td:nth-child(7) {
        width: var(--short-member-width);
        min-width: var(--short-member-width);
        max-width: var(--short-member-width);
        white-space: normal;
        word-break: break-word;
        overflow-wrap: anywhere;
        line-height: 1.05;
        padding: 2px 3px;
    }

    /* Pair summary: keep the 'エリア' column compact on iPhone */
    .wrap-pair-summary table { width: 100%; min-width: 0; }
    #pair-summary { width: 100%; min-width: 0; }
    #pair-summary th, #pair-summary td { padding: 2px 3px; }
    #pair-summary th:first-child, #pair-summary td:first-child {
        width: 4.2em;
        min-width: 4.2em;
        max-width: 4.2em;
        white-space: normal;
        word-break: break-word;
        overflow-wrap: anywhere;
        line-height: 1.05;
    }
}

@media print {
    @page { size: A4; margin: 8mm; }
    body { margin: 0; }
    h1 { margin: 0 0 6mm; }
    .filter-bar, datalist, #lock-overlay { display: none !important; }

    /* Print only the compact (short) lists */
    #match-matrix-section, #personal-schedule-section { display: none !important; }

    /* Remove scroll containers / sticky behavior */
    .wrap { overflow: visible !important; max-height: none !important; box-shadow: none !important; }
    .wrap table { width: 100% !important; }
    table { min-width: 0 !important; }
    th, td { font-size: 9pt; padding: 2mm 2mm; }
    th { position: static !important; }
    .sticky-col { position: static !important; box-shadow: none !important; }
    th.sticky-col, td.sticky-col { z-index: auto !important; }

    /* Improve page breaks */
    section.table-block { break-inside: avoid; page-break-inside: avoid; }
    tr { break-inside: avoid; page-break-inside: avoid; }
}
</style>
        """
    )
        fh.write("</head><body")
        if pass_hash:
            fh.write(" class='locked'")
        fh.write(">")
        fh.write("<h1>対戦一覧＆個人スケジュール</h1>")

        if pass_hash:
            fh.write(
                """
<div id='lock-overlay' class='lock-overlay'>
  <div class='lock-card'>
    <div class='lock-title'>パスコードが必要です</div>
    <div class='lock-note'>注意: これは“完全な暗号化”ではなく簡易ロックです（URL/HTMLを知っている人が解析すれば閲覧可能）。</div>
    <label class='lock-label'>パスコード <input id='lock-pass' type='password' inputmode='text' autocomplete='off' /></label>
    <div class='lock-actions'>
      <button id='lock-open' type='button'>開く</button>
      <span id='lock-error' class='lock-error' aria-live='polite'></span>
    </div>
  </div>
</div>
                """
            )

        render_match_count_summary(fh)
        render_summary_table(fh)
        render_table(
            fh,
            "match-matrix",
            "対戦表",
            matrix_headers,
            match_rows,
            enable_team_filter=True,
            team_options=team_filter_options,
            enable_club_filter=True,
            club_options=club_options,
            enable_club_root_filter=True,
            club_root_options=club_root_options,
            keyword_options=keyword_candidates,
            sticky_columns={0: ["match-col"]},
        )
        render_table(
            fh,
            "personal-schedule",
            "個人スケジュール表",
            personal_headers,
            personal_rows,
            enable_team_filter=True,
            team_options=team_filter_options,
            enable_club_filter=True,
            club_options=club_options,
            enable_club_root_filter=True,
            club_root_options=club_root_options,
            keyword_options=keyword_candidates,
            sticky_columns=({0: ["team-col"], 1: ["member-col"]} if include_members else {0: ["team-col"]}),
        )
        render_table(
            fh,
            "short-team",
            "対戦一覧短縮（チーム順）",
            short_headers,
            short_rows_team,
            enable_team_filter=True,
            team_options=team_filter_options,
            enable_club_filter=True,
            club_options=club_options,
            enable_club_root_filter=True,
            club_root_options=club_root_options,
            keyword_options=keyword_candidates,
            sticky_columns={0: ["round-col"], 1: ["court-col"]},
        )
        render_table(
            fh,
            "short-round",
            "対戦一覧短縮（試合順）",
            short_headers,
            short_rows_round,
            enable_team_filter=True,
            team_options=team_filter_options,
            enable_club_filter=True,
            club_options=club_options,
            enable_club_root_filter=True,
            club_root_options=club_root_options,
            keyword_options=keyword_candidates,
            sticky_columns={0: ["round-col"], 1: ["court-col"]},
        )

        # Put notes and extracted list at the bottom.
        render_back_to_back_table(fh)

        fh.write(
            r"""<script>
const HTML_PASS_HASH=""" + repr(pass_hash) + r""";
const CLUB_TEAMS=""" + json.dumps(club_to_teams, ensure_ascii=False) + r""";
const ROOT_TEAMS=""" + json.dumps(root_to_teams, ensure_ascii=False) + r""";
async function sha256Hex(text){
    const data=new TextEncoder().encode(text);
    const digest=await crypto.subtle.digest('SHA-256',data);
    const bytes=Array.from(new Uint8Array(digest));
    return bytes.map(b=>b.toString(16).padStart(2,'0')).join('');
}
function unlock(){
    document.body.classList.remove('locked');
    const overlay=document.getElementById('lock-overlay');
    if(overlay){overlay.remove();}
}
async function initLock(){
    if(!HTML_PASS_HASH){return;}
    const ok=localStorage.getItem('schedule_unlock_hash')===HTML_PASS_HASH;
    if(ok){unlock();return;}
    const input=document.getElementById('lock-pass');
    const btn=document.getElementById('lock-open');
    const err=document.getElementById('lock-error');
    const attempt=async ()=>{
        const val=(input&&input.value?input.value:'').trim();
        if(!val){err.textContent='パスコードを入力してください';return;}
        const h=await sha256Hex(val);
        if(h===HTML_PASS_HASH){
            localStorage.setItem('schedule_unlock_hash',HTML_PASS_HASH);
            unlock();
        }else{
            err.textContent='パスコードが違います';
        }
    };
    if(btn){btn.addEventListener('click',attempt);}
    if(input){input.addEventListener('keydown',e=>{if(e.key==='Enter'){attempt();}});}
}
initLock();
const escapeRegex=(str)=>str.replace(/[-\/\\^$*+?.()|[\]{}]/g,'\\$&');
document.querySelectorAll('.filter-bar').forEach(bar=>{
    const tableId=bar.dataset.table;
    const table=document.getElementById(tableId);
    if(!table){return;}
    const search=bar.querySelector('[data-role="search"]');
    const column=bar.querySelector('[data-role="column"]');
    const team=bar.querySelector('[data-role="team"]');
    const club=bar.querySelector('[data-role="club"]');
    const area=bar.querySelector('[data-role="club-root"]');
    const rows=Array.from(table.querySelectorAll('tbody tr'));

    rows.forEach(row=>{
        row.querySelectorAll('td').forEach(cell=>{
            if(cell.dataset.rawHtml===undefined){
                cell.dataset.rawHtml=cell.innerHTML;
            }
        });
    });

    const resetRow=row=>{
        row.classList.remove('matched');
        row.querySelectorAll('td').forEach(cell=>cell.classList.remove('cell-hit'));
        row.querySelectorAll('td').forEach(cell=>{
            if(cell.dataset.rawHtml!==undefined){
                cell.innerHTML=cell.dataset.rawHtml;
            }else{
                cell.dataset.rawHtml=cell.innerHTML;
            }
        });
    };

    const isNameChar=(ch)=>{
        if(!ch){return false;}
        // Only letters/numbers count as "name characters" for boundary checks.
        // This keeps B1 from matching B11 (digit boundary) while allowing matches
        // next to separators like '/', '-', '_' that may appear in the table text.
        return /[\p{L}\p{N}]/u.test(ch);
    };
    const containsExactName=(text,name)=>{
        if(!text||!name){return false;}
        let start=0;
        while(true){
            const idx=text.indexOf(name,start);
            if(idx===-1){return false;}
            const before=idx>0?text[idx-1]:'';
            const after=(idx+name.length)<text.length?text[idx+name.length]:'';
            if(!isNameChar(before)&&!isNameChar(after)){
                return true;
            }
            start=idx+1;
        }
    };
    const applyCellHits=(row,names)=>{
        if(!names||!names.length){return false;}
        let hit=false;
        const cells=Array.from(row.querySelectorAll('td'));
        cells.forEach(cell=>{
            // Use innerText so <br> becomes a separator; textContent would
            // concatenate pair name and member names, breaking boundary checks.
            const text=(cell.innerText||cell.textContent||'');
            for(const name of names){
                if(name && containsExactName(text,name)){
                    cell.classList.add('cell-hit');
                    hit=true;
                    break;
                }
            }
        });
        return hit;
    };


    const highlightRowText=(row,term,colValue)=>{
        if(!term){
            return false;
        }
        const targets=colValue==='all'?Array.from(row.children):[row.children[parseInt(colValue,10)]].filter(Boolean);
        const pattern=new RegExp(escapeRegex(term),'gi');
        let found=false;
        targets.forEach(cell=>{
            if(!cell){
                return;
            }
            const raw=cell.dataset.rawHtml!==undefined?cell.dataset.rawHtml:cell.innerHTML;
            const updated=raw.replace(pattern,match=>{
                found=true;
                return `<mark>${match}</mark>`;
            });
            cell.innerHTML=updated;
        });
        return found;
    };

    const filter=()=>{
        const rawTerm=(search&&search.value?search.value:'').trim();
        const term=rawTerm.toLowerCase();
        const col=column.value;
        const teamVal=team?team.value:'all';
        const clubVal=club?club.value:'all';
        const areaVal=area?area.value:'all';
        const narrowedBySelect=teamVal!=='all'||clubVal!=='all'||areaVal!=='all';

        rows.forEach(row=>{
            resetRow(row);
            let visible=true;
            if(teamVal!=='all'){
                const single=(row.dataset.team||'');
                const many=(row.dataset.teams||'');
                const list=many?many.split('|'):[];
                visible=visible&&(single===teamVal||list.includes(teamVal));
            }
            if(clubVal!=='all'){
                const clubsAttr=row.dataset.clubs||row.dataset.club||'';
                const clubList=clubsAttr?clubsAttr.split('|'):[];
                visible=visible&&clubList.includes(clubVal);
            }
            if(areaVal!=='all'){
                const rootAttr=row.dataset.clubRoots||row.dataset.clubRoot||'';
                const rootList=rootAttr?rootAttr.split('|'):[];
                visible=visible&&rootList.includes(areaVal);
            }
            let textMatch=true;
            if(term){
                let hay='';
                if(col==='all'){
                    hay=row.textContent.toLowerCase();
                }else{
                    const cell=row.children[parseInt(col,10)];
                    hay=cell?cell.textContent.toLowerCase():'';
                }
                textMatch=hay.includes(term);
                visible=visible&&textMatch;
            }
            row.style.display=visible?'':'none';
            let highlighted=false;
            if(visible&&rawTerm){
                highlighted=highlightRowText(row,rawTerm,col);
            }
            // When filtering by pair/club/area, highlight the relevant *cells* so it's
            // easy to spot where the match is in wide tables.
            if(visible&&!rawTerm){
                let names=[];
                if(teamVal!=='all'){
                    names=[teamVal];
                }else if(clubVal!=='all'){
                    names=CLUB_TEAMS[clubVal]||[];
                }else if(areaVal!=='all'){
                    names=ROOT_TEAMS[areaVal]||[];
                }
                highlighted=applyCellHits(row,names)||highlighted;
            }
            if(!rawTerm&&narrowedBySelect&&visible){
                row.classList.add('matched');
            }else if(highlighted){
                row.classList.add('matched');
            }
        });
    };

    if(search){
        search.addEventListener('input',filter);
    }
    column.addEventListener('change',filter);
    if(team){
        team.addEventListener('change',filter);
    }
    if(club){
        club.addEventListener('change',filter);
    }
    if(area){
        area.addEventListener('change',filter);
    }
});
</script>"""
        )
        fh.write("</body></html>")

def write_score_sheets_html(
    matches: List[Match],
    teams: List[Team],
    output_path: str,
    *,
    per_page: int = 10,
    columns: int = 2,
    include_members: bool = True,
    round_minutes: int = DEFAULT_ROUND_MINUTES,
    title: str = "得点記入表",
) -> None:
    """Write printable score sheets (one per match) as a static HTML.

    Layout is intentionally close to the provided template image.
    """

    if per_page <= 0:
        raise ValueError("per_page must be positive")
    if columns <= 0:
        raise ValueError("columns must be positive")
    if round_minutes <= 0:
        raise ValueError("round_minutes must be positive")

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    team_lookup = {t.name: t for t in teams}
    # Sort by court first so cut sheets can be grouped per court easily.
    ms = sorted(matches, key=lambda m: (m.court, m.round_num))

    def members_for(name: str, fallback: str) -> str:
        if not include_members:
            return ""
        t = team_lookup.get(name)
        return (t.members if (t and t.members) else fallback) or ""

    def fmt_time(dt: datetime) -> str:
        return dt.strftime("%H:%M")

    def fmt_end(dt: datetime) -> str:
        return (dt + timedelta(minutes=int(round_minutes))).strftime("%H:%M")

    def esc(s: str) -> str:
        return escape(s) if s else ""

    def card_html(m: Match) -> str:
        t1 = m.team1.name
        t2 = m.team2.name
        m1 = members_for(t1, m.team1.members)
        m2 = members_for(t2, m.team2.members)
        start = fmt_time(m.start_time)
        end = fmt_end(m.start_time)
        match_label = f"第{m.round_num}試合"

        parts: list[str] = [
            "<div class='card'>",
            "<table class='sheet'>",
            "<tr>",
            f"<th class='meta meta-top match'>{match_label}</th>",
            f"<th class='court-head' colspan='2'>コート <span class='court-no'>{m.court}</span></th>",
            "</tr>",
            "<tr>",
            f"<th class='meta'>時間</th>",
            f"<td class='pair'>{esc(t1)}</td>",
            f"<td class='pair'>{esc(t2)}</td>",
            "</tr>",
            "<tr>",
            f"<td class='time'>{start}-{end}</td>",
            f"<td class='members'>{esc(m1) if include_members else ''}</td>",
            f"<td class='members'>{esc(m2) if include_members else ''}</td>",
            "</tr>",
        ]

        parts.extend(
            [
                "<tr class='score-row'>",
                "<th class='meta'>得点</th>",
                "<td class='blank blank-score'></td>",
                "<td class='blank blank-score'></td>",
                "</tr>",
                "<tr class='sign-row'>",
                "<th class='meta'>サイン(審判/勝者)</th>",
                "<td class='blank blank-sign'></td>",
                "<td class='blank blank-sign'></td>",
                "</tr>",
                "</table>",
                "</div>",
            ]
        )
        return "".join(parts)

    pages: list[list[Match]] = [ms[i : i + per_page] for i in range(0, len(ms), per_page)]
    rows = int(math.ceil(per_page / columns))

    with path.open("w", encoding="utf-8") as fh:
        fh.write("<!DOCTYPE html><html lang='ja'><head><meta charset='utf-8'>")
        fh.write("<meta name='viewport' content='width=device-width, initial-scale=1'>")
        fh.write(f"<title>{escape(title)}</title>")
        fh.write(
            """
<style>
body {font-family: system-ui, sans-serif; margin: 12px; color: #111;}
h1 {margin: 0 0 10px; font-size: 18px;}
.note {margin: 0 0 12px; font-size: 12px; color: #333;}

.page {page-break-after: always;}
.grid {display: grid; grid-template-columns: repeat(var(--cols), 1fr); grid-template-rows: repeat(var(--rows), 1fr); gap: 5mm; align-items: stretch;}

.card {border: 2px solid #111; border-radius: 4px; padding: 0; break-inside: avoid; height: 100%; display: flex;}

table.sheet {border-collapse: collapse; width: 100%; table-layout: fixed; height: 100%; flex: 1;}
table.sheet th, table.sheet td {border: 2px solid #111; padding: 2.8mm 2.3mm; vertical-align: middle;}

th.meta {width: 28%; font-size: 16px; font-weight: 900; text-align: left;}
th.meta-top {vertical-align: top;}

th.court-head {text-align: center; font-weight: 900; font-size: 18px; padding: 2.2mm;}
.court-no {font-size: 30px; font-weight: 900; margin-left: 2mm;}

th.match {font-size: 20px; font-weight: 900; text-align: center;}
td.time {font-size: 20px; font-weight: 900; text-align: center;}
td.pair {font-size: 22px; font-weight: 900; text-align: center; line-height: 1.1;}
td.members {font-size: 18px; font-weight: 700; text-align: center; word-break: break-word; overflow-wrap: anywhere;}

td.blank {background: #fff;}
td.blank-score {min-height: 18mm; height: auto;}
td.blank-sign {min-height: 15mm; height: auto;}

@media print {
    @page { size: A4; margin: 7mm; }
  body {margin: 0;}
  .page-title {display: none;}
  .note {display: none;}
    /* Make 8-up/10-up actually consume the full printable height */
    .page {height: calc(297mm - 14mm); --gap: 4mm;}
    .grid {height: 100%; gap: var(--gap);}
    .card {height: calc((100% - ((var(--rows) - 1) * var(--gap))) / var(--rows));}

    /* Scale blank boxes with card height so bottom whitespace turns into writable area */
    .score-row td.blank-score {height: calc(((100% - 0mm) * 0.28));}
    .sign-row td.blank-sign {height: calc(((100% - 0mm) * 0.22));}
}
</style>
            """
        )
        fh.write(f"</head><body class='pp-{int(per_page)} cols-{int(columns)}'>")
        fh.write(f"<h1 class='page-title'>{escape(title)}</h1>")
        fh.write(
            "<p class='note'><strong>印刷の倍率（Scale）を確認/調整してください（例: 100%）。</strong><br>各試合1枚。印刷してカットして使えます。</p>"
        )

        for page in pages:
            fh.write(f"<div class='page' style='--cols:{int(columns)}; --rows:{int(rows)}'>")
            fh.write("<div class='grid'>")
            for m in page:
                fh.write(card_html(m))
            fh.write("</div></div>")

        fh.write("</body></html>")


def write_wall_cards_html(
    matches: List[Match],
    teams: List[Team],
    output_path: str,
    *,
    columns: int = 4,
    include_members: bool = True,
    round_minutes: int = DEFAULT_ROUND_MINUTES,
    title: str = "壁貼り（カード式・得点欄付き）",
) -> None:
    """Write a wall-posting HTML using match cards.

    - 1試合=1カード
    - 得点欄は「1試合につき1行」（例: ____ : ____）
    - 印刷しやすいように多段組（columns=4/5など）

    This is intended to be saved as PDF (Ctrl+P → PDF) and posted on the wall.
    """

    if columns <= 0:
        raise ValueError("columns must be positive")
    if round_minutes <= 0:
        raise ValueError("round_minutes must be positive")

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    team_lookup = {t.name: t for t in teams}

    # Keep a stable order: time (round) then court.
    ms = sorted(matches, key=lambda m: (m.round_num, m.court))

    def members_for(name: str, fallback: str) -> str:
        if not include_members:
            return ""
        t = team_lookup.get(name)
        return (t.members if (t and t.members) else fallback) or ""

    def fmt_time(dt: datetime) -> str:
        return dt.strftime("%H:%M")

    def fmt_end(dt: datetime) -> str:
        return (dt + timedelta(minutes=int(round_minutes))).strftime("%H:%M")

    def esc(s: str) -> str:
        return escape(s) if s else ""

    # --- Lightweight team marker (SVG chip) for quick visual scanning ---
    import re
    import colorsys

    def _team_group_key(team_name: str) -> str:
        s = (team_name or "").strip()
        if not s:
            return ""
        m = re.match(r"^([^0-9A-Za-z]+)", s)
        if m:
            key = m.group(1).strip()
            if key:
                return key
        m2 = re.match(r"^(.+?)(?=\d)", s)
        if m2:
            key = m2.group(1).strip()
            if key:
                return key
        return s

    def _rgb_to_hex(rgb: tuple[int, int, int]) -> str:
        r, g, b = rgb
        r = max(0, min(255, int(r)))
        g = max(0, min(255, int(g)))
        b = max(0, min(255, int(b)))
        return f"#{r:02x}{g:02x}{b:02x}"

    def _hsl_to_hex(h: float, s: float, l: float) -> str:
        r, g, b = colorsys.hls_to_rgb((h % 360.0) / 360.0, l / 100.0, s / 100.0)
        return _rgb_to_hex((int(round(r * 255)), int(round(g * 255)), int(round(b * 255))))

    team_names: list[str] = sorted(
        {
            t
            for m in matches
            for t in (
                (m.team1.name if m.team1 else ""),
                (m.team2.name if m.team2 else ""),
            )
            if t
        }
    )
    group_by_name: dict[str, str] = {name: _team_group_key(name) for name in team_names}
    groups: list[str] = sorted({g for g in group_by_name.values() if g})
    group_index: dict[str, int] = {g: i for i, g in enumerate(groups)}
    shapes = ["circle", "square", "triangle", "diamond", "hex", "star"]

    _VIVID_BASE_COLORS: list[str] = [
        "#e6194B",
        "#3cb44b",
        "#4363d8",
        "#f58231",
        "#911eb4",
        "#42d4f4",
        "#ffe119",
        "#f032e6",
        "#bfef45",
        "#469990",
        "#9A6324",
        "#800000",
        "#000075",
        "#808000",
        "#a9a9a9",
    ]

    base_by_group: dict[str, str] = {}
    for i, g in enumerate(groups):
        if i < len(_VIVID_BASE_COLORS):
            base_by_group[g] = _VIVID_BASE_COLORS[i]
        else:
            h = (i * 137.508) % 360.0
            base_by_group[g] = _hsl_to_hex(h, s=92.0, l=42.0)

    def _chip_svg(base_hex: str, shape: str) -> str:
        fill = esc(base_hex)
        if shape == "circle":
            body = "<circle cx='7' cy='7' r='5.4' />"
        elif shape == "square":
            body = "<rect x='2' y='2' width='10' height='10' rx='1.6' ry='1.6' />"
        elif shape == "triangle":
            body = "<polygon points='7,1.8 12.4,11.8 1.6,11.8' />"
        elif shape == "diamond":
            body = "<polygon points='7,1.6 12.4,7 7,12.4 1.6,7' />"
        elif shape == "hex":
            body = "<polygon points='7,1.4 12.2,4.3 12.2,9.7 7,12.6 1.8,9.7 1.8,4.3' />"
        else:
            body = "<polygon points='7,1.2 8.7,5.3 13.1,5.6 9.7,8.4 10.8,12.7 7,10.4 3.2,12.7 4.3,8.4 0.9,5.6 5.3,5.3' />"
        return (
            "<svg class='chip' width='14' height='14' viewBox='0 0 14 14' aria-hidden='true' focusable='false'>"
            f"<g fill='{fill}' stroke='#111' stroke-width='1.2'>{body}</g>"
            "</svg>"
        )

    def team_block(team: Team) -> str:
        name = team.name or ""
        g = _team_group_key(name)
        base_hex = base_by_group.get(g, "#777777")
        shape = shapes[group_index.get(g, 0) % len(shapes)]
        chip = _chip_svg(base_hex, shape)
        members = members_for(name, team.members)
        members_html = f"<div class='members'>{esc(members)}</div>" if (include_members and members) else ""
        # Border color prints reliably (even when background graphics is OFF).
        return (
            f"<div class='team' style='border-left-color:{esc(base_hex)}'>"
            f"{chip}<div class='name'>{esc(name)}</div>{members_html}</div>"
        )

    def card_html(m: Match) -> str:
        start = fmt_time(m.start_time)
        end = fmt_end(m.start_time)
        meta = f"第{m.round_num}試合 / {start}-{end} / コート{m.court}"
        return (
            "<article class='card'>"
            f"<header class='meta'>{esc(meta)}</header>"
            "<div class='teams'>"
            f"{team_block(m.team1)}"
            f"{team_block(m.team2)}"
            "</div>"
            "<div class='score'>"
            "<span class='label'>得点</span>"
            "<span class='blank'></span><span class='sep'>:</span><span class='blank'></span>"
            "</div>"
            "</article>"
        )

    with path.open("w", encoding="utf-8") as fh:
        fh.write("<!DOCTYPE html><html lang='ja'><head><meta charset='utf-8'>")
        fh.write("<meta name='viewport' content='width=device-width, initial-scale=1'>")
        fh.write(f"<title>{escape(title)}</title>")
        fh.write(
            """
<style>
body {font-family: system-ui, sans-serif; margin: 12px; color: #111;}
h1 {margin: 0 0 10px; font-size: 18px;}
.note {margin: 0 0 12px; font-size: 12px; color: #333;}

.columns {column-count: var(--cols); column-gap: 10mm;}
.card {break-inside: avoid; border: 2px solid #111; border-radius: 6px; padding: 6mm 6mm 5mm; margin: 0 0 8mm;}
.meta {font-size: 12.5px; font-weight: 800; margin: 0 0 5mm;}

.teams {display: grid; gap: 4mm;}
.team {border-left: 8px solid #777; padding-left: 4mm; display: grid; grid-template-columns: 16px 1fr; column-gap: 3mm; align-items: start;}
.chip {margin-top: 1px;}
.name {font-size: 18px; font-weight: 900; line-height: 1.15; word-break: break-word; overflow-wrap: anywhere;}
.members {grid-column: 2 / 3; font-size: 14px; font-weight: 700; color: #222; margin-top: 1mm; word-break: break-word; overflow-wrap: anywhere;}

.score {margin-top: 6mm; display: flex; align-items: baseline; gap: 4mm;}
.score .label {font-size: 13px; font-weight: 900;}
.score .blank {flex: 1 1 auto; border-bottom: 3px solid #111; min-width: 24mm; height: 8mm;}
.score .sep {font-size: 16px; font-weight: 900;}

@media print {
  @page { size: A4; margin: 7mm; }
  body { margin: 0; -webkit-print-color-adjust: exact; print-color-adjust: exact; }
  h1, .note { display: none; }
  .columns { column-gap: 7mm; }
  .card { margin: 0 0 6mm; }
}
</style>
            """
        )
        fh.write(f"</head><body style='--cols:{int(columns)}'>")
        fh.write(f"<h1>{escape(title)}</h1>")
        fh.write(
            "<p class='note'><strong>印刷の倍率（Scale）を確認/調整してください（例: 100%）。</strong><br>"
            "カード式（縦長）なのでPDF化して壁貼りしやすいです。各試合につき得点欄は1行です。</p>"
        )
        fh.write("<div class='columns'>")
        for m in ms:
            fh.write(card_html(m))
        fh.write("</div>")
        fh.write("</body></html>")


def write_wall_schedule_html(
    matches: List[Match],
    output_path: str,
    num_rounds: int,
    courts: int,
    *,
    start_time_hhmm: str = DEFAULT_START_TIME_HHMM,
    round_minutes: int = DEFAULT_ROUND_MINUTES,
    courts_per_page: int = 1,
    team_color_rules: list[tuple[str, str]] | None = None,
    auto_team_colors: bool = True,
    cell_background: bool = True,
) -> None:
    """Generate a print-friendly (wall-posting) HTML.

    - 1ページあたり1〜4コート（画面表示の面付け）
    - 印刷も『1ページあたり1〜4コート』（courts_per_pageに従う）
    - No JavaScript required
    """

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    mc_lookup: dict[tuple[int, int], Match] = {(m.round_num, m.court): m for m in matches}

    def round_start_hhmm(round_num: int) -> str:
        any_match = next((m for m in matches if m.round_num == round_num), None)
        if any_match:
            return any_match.start_time.strftime("%H:%M")
        base = _base_datetime_from_hhmm(start_time_hhmm)
        return (base + timedelta(minutes=int(round_minutes) * (round_num - 1))).strftime("%H:%M")

    rules: list[tuple[str, str]] = list(team_color_rules or [])

    def _hex_to_rgb(hex_color: str) -> tuple[int, int, int]:
        s = hex_color.strip()
        if s.startswith("#"):
            s = s[1:]
        if len(s) != 6:
            raise ValueError(f"bad hex color: {hex_color}")
        return int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16)

    def _rgb_to_hex(rgb: tuple[int, int, int]) -> str:
        r, g, b = rgb
        r = max(0, min(255, int(r)))
        g = max(0, min(255, int(g)))
        b = max(0, min(255, int(b)))
        return f"#{r:02x}{g:02x}{b:02x}"

    def _blend(a: tuple[int, int, int], b: tuple[int, int, int], t: float) -> tuple[int, int, int]:
        # t=0 -> a, t=1 -> b
        return (
            int(round(a[0] + (b[0] - a[0]) * t)),
            int(round(a[1] + (b[1] - a[1]) * t)),
            int(round(a[2] + (b[2] - a[2]) * t)),
        )

    def _hsl_to_hex(h: float, s: float, l: float) -> str:
        import colorsys

        # colorsys uses HLS
        r, g, b = colorsys.hls_to_rgb((h % 360.0) / 360.0, l / 100.0, s / 100.0)
        return _rgb_to_hex((int(round(r * 255)), int(round(g * 255)), int(round(b * 255))))

    import re

    def _team_group_key(team_name: str) -> str:
        """Group teams by prefix so 上海A1/上海A2 share the same color.

        Rule (simple + robust for JP/CJK names):
        - Take all leading characters up to the first ASCII letter/digit.
        - If that becomes empty, fall back to the substring before first digit.
        - If still empty, use the full name.
        """

        s = (team_name or "").strip()
        if not s:
            return ""
        m = re.match(r"^([^0-9A-Za-z]+)", s)
        if m:
            key = m.group(1).strip()
            if key:
                return key
        m2 = re.match(r"^(.+?)(?=\d)", s)
        if m2:
            key = m2.group(1).strip()
            if key:
                return key
        return s

    # Build stable color assignment per team *group*.
    # - Manual rules override (keyword match)
    # - Otherwise, auto assigns distinct hues in stable group order.
    team_names: list[str] = sorted(
        {
            t
            for m in matches
            for t in (
                (m.team1.name if m.team1 else ""),
                (m.team2.name if m.team2 else ""),
            )
            if t
        }
    )
    group_by_name: dict[str, str] = {name: _team_group_key(name) for name in team_names}
    groups: list[str] = sorted({g for g in group_by_name.values() if g})

    # Stable shape assignment per group (in addition to color).
    # Even if colors look similar, shape makes it immediately distinguishable.
    group_index: dict[str, int] = {g: i for i, g in enumerate(groups)}
    shapes = ["circle", "square", "triangle", "diamond", "hex", "star"]

    base_by_group: dict[str, str] = {}
    bg_by_group: dict[str, str] = {}
    white = (255, 255, 255)

    # --- Color palette helpers (wall-friendly, high-contrast) ---
    # For "遠くから見てパッと違う" を優先して、まずは鮮やかな定番の distinct palette を使う。
    # グループ数が多い場合は HSL(ゴールデンアングル)で追加生成する。
    _VIVID_BASE_COLORS: list[str] = [
        "#e6194B",  # red
        "#3cb44b",  # green
        "#4363d8",  # blue
        "#f58231",  # orange
        "#911eb4",  # purple
        "#42d4f4",  # cyan
        "#ffe119",  # yellow
        "#f032e6",  # magenta
        "#bfef45",  # lime
        "#469990",  # teal
        "#9A6324",  # brown
        "#800000",  # maroon
        "#000075",  # navy
        "#aaffc3",  # mint
        "#ffd8b1",  # apricot
        "#dcbeff",  # lavender
        "#808000",  # olive
        "#fabed4",  # pink
        "#fffac8",  # beige
        "#a9a9a9",  # gray
    ]

    def _pick_distinct_group_colors(n: int) -> list[tuple[str, str]]:
        if n <= 0:
            return []

        out: list[tuple[str, str]] = []
        for i in range(n):
            if i < len(_VIVID_BASE_COLORS):
                base_hex = _VIVID_BASE_COLORS[i]
            else:
                # Fallback: vivid HSL using golden-angle spacing (deterministic).
                h = (i * 137.508) % 360.0
                base_hex = _hsl_to_hex(h, s=92.0, l=42.0)

            # Background: not too pale so it still reads at a distance.
            bg_hex = _rgb_to_hex(_blend(_hex_to_rgb(base_hex), white, 0.55))
            out.append((base_hex, bg_hex))

        return out

    # Apply manual rules (first match wins by rule order) to groups.
    for g in groups:
        for keyword, base_hex in rules:
            if keyword and (keyword in g):
                base_by_group[g] = base_hex
                bg_by_group[g] = _rgb_to_hex(_blend(_hex_to_rgb(base_hex), white, 0.55))
                break

    # Also allow keyword match against full team names (if user targets e.g. "蛇口")
    for name in team_names:
        g = group_by_name.get(name, "")
        if not g or g in base_by_group:
            continue
        for keyword, base_hex in rules:
            if keyword and (keyword in name):
                base_by_group[g] = base_hex
                bg_by_group[g] = _rgb_to_hex(_blend(_hex_to_rgb(base_hex), white, 0.55))
                break

    # Auto colors (unique-ish per team)
    if auto_team_colors:
        need = sum(1 for g in groups if g not in base_by_group)
        palette = _pick_distinct_group_colors(need)
        pi = 0
        for g in groups:
            if g in base_by_group:
                continue
            base_hex, bg_hex = palette[pi]
            pi += 1
            base_by_group[g] = base_hex
            bg_by_group[g] = bg_hex

    def _chip_svg(base_hex: str, shape: str) -> str:
        # Inline SVG prints reliably (even when background graphics is OFF).
        # Keep a black stroke so it stays visible in B/W printing.
        fill = escape(base_hex)
        if shape == "circle":
            body = "<circle cx='7' cy='7' r='5.4' />"
        elif shape == "square":
            body = "<rect x='2' y='2' width='10' height='10' rx='1.6' ry='1.6' />"
        elif shape == "triangle":
            body = "<polygon points='7,1.8 12.4,11.8 1.6,11.8' />"
        elif shape == "diamond":
            body = "<polygon points='7,1.6 12.4,7 7,12.4 1.6,7' />"
        elif shape == "hex":
            body = "<polygon points='7,1.4 12.2,4.3 12.2,9.7 7,12.6 1.8,9.7 1.8,4.3' />"
        else:  # star
            body = "<polygon points='7,1.2 8.7,5.3 13.1,5.6 9.7,8.4 10.8,12.7 7,10.4 3.2,12.7 4.3,8.4 0.9,5.6 5.3,5.3' />"
        return (
            "<svg class='chip' width='14' height='14' viewBox='0 0 14 14' aria-hidden='true' focusable='false'>"
            f"<g fill='{fill}' stroke='#111' stroke-width='1.2'>{body}</g>"
            "</svg>"
        )

    def render_team_cell(team: Team | None) -> str:
        if not team:
            return ""
        name = team.name or ""
        g = _team_group_key(name)
        base_hex = base_by_group.get(g, "#777777")
        shape = shapes[group_index.get(g, 0) % len(shapes)]
        chip = _chip_svg(base_hex, shape)

        member_html = f"<small>{escape(team.members)}</small>" if team.members else ""
        # Score entry is under member names (one number per team) so the wall table stays narrow.
        score_html = "<div class='score-entry'><span class='score-label'>得点</span><span class='score-blank'></span></div>"
        return (
            "<div class='team-cell'>"
            f"{chip}"
            "<div class='team-text'>"
            f"<div class='team-name'>{escape(name)}</div>"
            f"{member_html}"
            f"{score_html}"
            "</div>"
            "</div>"
        )

    cpp = int(courts_per_page)
    if cpp <= 0:
        cpp = 1
    if cpp > 4:
        cpp = 4

    with path.open("w", encoding="utf-8") as fh:
        fh.write("<!DOCTYPE html><html lang='ja'><head><meta charset='utf-8'>")
        fh.write("<meta name='viewport' content='width=device-width, initial-scale=1'>")
        fh.write("<title>壁貼り用（コート別）</title>")
        fh.write(
            """
<style>
body {font-family: system-ui, sans-serif; margin: 12px; color: #111;}
h1 {margin: 0 0 10px; font-size: 18px;}
h2 {margin: 18px 0 8px; font-size: 16px;}
.note {font-size: 12px; color: #444; margin: 0 0 14px;}
.page {margin-bottom: 18px;}
.grid {display: grid; gap: 10px; align-items: start;}
.cols-1 {grid-template-columns: 1fr;}
.cols-2 {grid-template-columns: 1fr 1fr;}
.cols-3 {grid-template-columns: 1fr 1fr 1fr;}
.cols-4 {grid-template-columns: 1fr 1fr 1fr 1fr;}

/* Scrolling wrapper so sticky columns work reliably */
.court-wrap {overflow: auto; -webkit-overflow-scrolling: touch; max-width: 100%;}
table {border-collapse: collapse; width: 100%; min-width: 520px; table-layout: fixed;}
th, td {border: 1px solid #bbb; padding: 6px 8px; vertical-align: top;}
th {background: #f2f2f2; text-align: left;}
td.round, td.time {white-space: nowrap;}
td.team {font-weight: 600;}
td.team small {font-weight: 400; color: #333;}
tr:nth-child(even) td {background: #fafafa;}

/* Team cell layout + chip */
.team-cell {display: flex; gap: 6px; align-items: flex-start;}
.team-text {min-width: 0;}
.chip {flex: 0 0 auto; margin-top: 1px;}
.team-name {font-weight: 800;}

/* Score entry inside team cell */
.score-entry {margin-top: 3px; display: flex; gap: 6px; align-items: baseline;}
.score-label {font-size: 11px; font-weight: 900; white-space: nowrap; color: #111;}
.score-blank {flex: 1 1 auto; min-width: 26px; border-bottom: 2px solid #111; height: 12px;}

/* Column sizing + sticky left columns (screen) */
:root { --wall-col-round: 84px; --wall-col-time: 62px; }
th.round-col {min-width: var(--wall-col-round); width: var(--wall-col-round);}
th.time-col {min-width: var(--wall-col-time); width: var(--wall-col-time);}

th.round-col, td.round {position: sticky; left: 0; z-index: 2; background: #f2f2f2;}
th.time-col, td.time {position: sticky; left: var(--wall-col-round); z-index: 2; background: #f2f2f2;}
td.round, td.time {background: #fff;}
tr:nth-child(even) td.round, tr:nth-child(even) td.time {background: #fafafa;}

/* (score entry moved under member names) */

/* Thicker borders to clearly separate courts */
th.group-start, td.group-start {border-left: 3px solid #111;}
th.group-end, td.group-end {border-right: 3px solid #111;}
thead tr:first-child th.court-head {border-top: 3px solid #111;}
tr.last-round td.group-start, tr.last-round td.group-mid, tr.last-round td.group-end {border-bottom: 3px solid #111;}

td.group-mid { /* marker class for last-row bottom border */ }

th.court-head {text-align: center; background: #e9e9e9;}

@media print {
    /* 縦で1枚に収めやすくする */
    @page { size: A4; margin: 7mm; }
    body { -webkit-print-color-adjust: exact; print-color-adjust: exact; }
  h1, .note { display: none; }
    /* Print one '.page' block at a time (1〜3コート/枚) */
    .page { margin: 0; break-after: page; page-break-after: always; }
    .page:last-of-type { break-after: auto; page-break-after: auto; }
    .grid { display: grid; gap: 6mm; }
    .court-wrap { overflow: visible !important; }
                /* Make the table consume the full printable height; distribute rows by round count */
                .page { height: calc(297mm - 14mm); }
                section.court-group { height: 100%; }
                .court-wrap { height: 100%; }
                table { height: 100%; }
                :root { --thead-row-h: 9mm; }
                thead tr { height: var(--thead-row-h); }
                tbody tr { height: calc((100% - (2 * var(--thead-row-h))) / var(--rounds)); }
                tbody td { vertical-align: middle; }
        /* Default (3コート/枚相当の密度) */
        table { min-width: 0 !important; font-size: 9pt; line-height: 1.15; }
        th, td { padding: 1.3mm 1.6mm; }
        td.team small { font-size: 8pt; }
        .score-label { font-size: 8pt; }
        .score-blank { height: 7mm; border-bottom-width: 2px; }

        /* 1〜2コート/枚は縦が余りやすいので、字を大きくして読みやすさ優先 */
        body.cpp-2 { --thead-row-h: 10mm; }
        body.cpp-2 table { font-size: 11.5pt; line-height: 1.18; }
        body.cpp-2 th, body.cpp-2 td { padding: 2.0mm 2.1mm; }
        body.cpp-2 td.team small { font-size: 10.25pt; }
        body.cpp-2 .score-label { font-size: 9.5pt; }
        body.cpp-2 .score-blank { height: 8.5mm; }

        body.cpp-1 { --thead-row-h: 11mm; }
        body.cpp-1 table { font-size: 13pt; line-height: 1.2; }
        body.cpp-1 th, body.cpp-1 td { padding: 2.4mm 2.4mm; }
        body.cpp-1 td.team small { font-size: 11.5pt; }
        body.cpp-1 .score-label { font-size: 10.5pt; }
        body.cpp-1 .score-blank { height: 9mm; }
    /* Sticky is not needed for print */
    th.round-col, td.round, th.time-col, td.time { position: static !important; }
  h2 { margin-top: 0; }
    section.court-group { break-inside: avoid; page-break-inside: avoid; }
    table, tr, td, th { break-inside: avoid; page-break-inside: avoid; }
}
</style>
            """
        )
        fh.write(f"</head><body class='cpp-{int(cpp)}'>")
        fh.write("<h1>壁貼り用（コート別）</h1>")
        note = "印刷は Ctrl+P。1枚に表示するコート数は --wall-courts-per-page で調整できます。"
        fh.write(
            "<p class='note'><strong>印刷の倍率（Scale）を確認/調整してください（例: 100%）。</strong><br>"
            + escape(note)
            + "</p>"
        )

        court_list = list(range(1, courts + 1))
        for page_start in range(0, len(court_list), cpp):
            chunk = court_list[page_start : page_start + cpp]
            fh.write(f"<div class='page' style='--rounds:{int(num_rounds)}'>")
            fh.write("<section class='court-group'>")
            if len(chunk) == 1:
                fh.write(f"<h2>コート{chunk[0]}</h2>")
            else:
                fh.write(f"<h2>コート{chunk[0]}〜{chunk[-1]}</h2>")

            fh.write("<div class='court-wrap'>")
            fh.write("<table><thead>")

            # Header row 1: shared left columns + per-court group headers
            fh.write("<tr>")
            fh.write("<th class='round-col' rowspan='2'>試合</th>")
            fh.write("<th class='time-col' rowspan='2'>開始</th>")
            for court in chunk:
                fh.write(f"<th class='court-head group-start group-end' colspan='2'>コート{court}</th>")
            fh.write("</tr>")

            # Header row 2: per-court columns
            fh.write("<tr>")
            for _court in chunk:
                fh.write("<th class='group-start'>チーム1</th>")
                fh.write("<th class='group-end'>チーム2</th>")
            fh.write("</tr>")

            fh.write("</thead><tbody>")
            for round_num in range(1, num_rounds + 1):
                tr_class = "last-round" if round_num == num_rounds else ""
                fh.write(f"<tr class='{tr_class}'>")
                fh.write(f"<td class='round'>第{round_num}試合</td>")
                fh.write(f"<td class='time'>{escape(round_start_hhmm(round_num))}</td>")

                for court in chunk:
                    m = mc_lookup.get((round_num, court))
                    t1 = m.team1 if m else None
                    t2 = m.team2 if m else None
                    t1_html = render_team_cell(t1) if t1 else ""
                    t2_html = render_team_cell(t2) if t2 else ""
                    t1_group = _team_group_key(t1.name) if t1 and t1.name else ""
                    t2_group = _team_group_key(t2.name) if t2 and t2.name else ""
                    t1_bg = bg_by_group.get(t1_group, "") if t1_group else ""
                    t2_bg = bg_by_group.get(t2_group, "") if t2_group else ""
                    t1_base = base_by_group.get(t1_group, "") if t1_group else ""
                    t2_base = base_by_group.get(t2_group, "") if t2_group else ""

                    def _cell_style(bg: str, base: str) -> str:
                        parts: list[str] = []
                        if cell_background and bg:
                            parts.append(f"background-color:{bg}")
                        # Strong visual marker that prints even when background is disabled.
                        if base:
                            parts.append(
                                f"box-shadow: inset 16px 0 0 0 {base}, inset 0 0 0 3px {base}"
                            )
                        if not parts:
                            return ""
                        return " style='" + "; ".join(parts) + ";'"

                    t1_style = _cell_style(escape(t1_bg), escape(t1_base))
                    t2_style = _cell_style(escape(t2_bg), escape(t2_base))
                    fh.write(f"<td class='team group-start group-mid'{t1_style}>{t1_html}</td>")
                    fh.write(f"<td class='team group-end group-mid'{t2_style}>{t2_html}</td>")

                fh.write("</tr>")
            fh.write("</tbody></table></div></section>")
            fh.write("</div>")

        fh.write("</body></html>")

import typer

app = typer.Typer()


_WALL_COLOR_ALIASES: dict[str, str] = {
    "red": "#d60000",
    "r": "#d60000",
    "blue": "#0057d6",
    "b": "#0057d6",
    "green": "#0a8f2a",
    "g": "#0a8f2a",
    "orange": "#d66a00",
    "o": "#d66a00",
    "purple": "#6b2bd6",
    "p": "#6b2bd6",
    "gray": "#666666",
    "grey": "#666666",
    "yellow": "#b59a00",
    "y": "#b59a00",
}


def _parse_wall_team_color_rules(values: list[str]) -> list[tuple[str, str]]:
    """Parse repeated --wall-team-color values like '蛇口:red'.

    - Keyword match: substring match against team name.
    - Priority: first match wins (order of options).
    - Color is one of: red/blue/green/orange/purple/gray/yellow
    """

    rules: list[tuple[str, str]] = []
    for raw in values:
        if raw is None:
            continue
        s = str(raw).strip()
        if not s:
            continue
        sep = ":" if ":" in s else ("：" if "：" in s else "")
        if not sep:
            raise typer.BadParameter("--wall-team-color は 'キーワード:色' 形式で指定してください。例: 蛇口:red")
        keyword, color_raw = s.split(sep, 1)
        keyword = keyword.strip()
        color_key = color_raw.strip().lower()
        if not keyword:
            raise typer.BadParameter("--wall-team-color のキーワードが空です")
        base_hex: str | None = None
        if color_key in _WALL_COLOR_ALIASES:
            base_hex = _WALL_COLOR_ALIASES[color_key]
        else:
            # Accept hex like #RRGGBB or RRGGBB
            hex_candidate = color_key
            if hex_candidate.startswith("#"):
                hex_candidate = hex_candidate[1:]
            if len(hex_candidate) == 6 and all(c in "0123456789abcdef" for c in hex_candidate):
                base_hex = f"#{hex_candidate}"
        if not base_hex:
            allowed = ", ".join(sorted(_WALL_COLOR_ALIASES.keys())) + ", #RRGGBB"
            raise typer.BadParameter(f"不明な色: {color_raw}（利用可能: {allowed}）")
        rules.append((keyword, base_hex))
    return rules


def _coerce_hhmm(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return None
        # Accept ranges like '13:29-13:42' (take start).
        for sep in ("-", "－", "–", "—", "〜", "～"):
            if sep in v:
                v = v.split(sep, 1)[0].strip()
                break
        # accept HH:MM or H:MM
        parts = v.split(":")
        if len(parts) == 2:
            try:
                h = int(parts[0])
                m = int(parts[1])
                if 0 <= h <= 23 and 0 <= m <= 59:
                    return f"{h:02d}:{m:02d}"
            except Exception:
                return None
        return None
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    # openpyxl can return datetime.time
    try:
        from datetime import time as _time

        if isinstance(value, _time):
            return f"{value.hour:02d}:{value.minute:02d}"
    except Exception:
        pass
    return None


def load_schedule_from_xlsx(schedule_xlsx: str, *, fallback_start_time_hhmm: str, fallback_round_minutes: int) -> tuple[list[Match], list[Team], int, int]:
    """Load matches/teams from an Excel exported by this tool, including manual edits.

    Expected sheets:
    - 対戦表: round rows with per-court team1/team2 cells
    - ペア一覧: team metadata (members/level/group)
    """

    wb = openpyxl.load_workbook(schedule_xlsx, data_only=True)
    if "対戦表" not in wb.sheetnames:
        raise ValueError("Excelに '対戦表' シートが見つかりません")
    ws = wb["対戦表"]

    teams_by_name: dict[str, Team] = {}

    def split_team_cell(value: Any) -> tuple[str, str]:
        """Parse a team cell that may contain 'TeamName\nMember1, Member2'."""
        if value is None:
            return "", ""
        raw = str(value).replace("\r", "\n").strip()
        if not raw:
            return "", ""
        lines = [ln.strip() for ln in raw.split("\n") if ln.strip()]
        if not lines:
            return "", ""
        name = lines[0]
        members = " ".join(lines[1:]).strip() if len(lines) > 1 else ""
        return name, members
    if "ペア一覧" in wb.sheetnames:
        ws_pairs = wb["ペア一覧"]
        # header: ペア名, 選手名, レベル, グループ, 試合数
        for row in ws_pairs.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name = str(row[0]).strip() if row[0] is not None else ""
            if not name:
                continue
            members = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            level = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            group = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            teams_by_name[name] = Team(name=name, members=members, level=level, group=group)

    def get_team(name_raw: Any, *, members_hint: str = "") -> Team:
        name = str(name_raw).strip() if name_raw is not None else ""
        if not name:
            raise ValueError("対戦表に空のチーム名が含まれています")
        if name in teams_by_name:
            team = teams_by_name[name]
        else:
            team = Team(name=name, members="", level="", group="")
            teams_by_name[name] = team

        # Prefer the 'ペア一覧' sheet as the source of truth for member names.
        # Only use embedded members text from the match table when we don't have
        # members in the pair list (e.g. legacy exports without a reliable pair list).
        if members_hint and not team.members:
            team.members = members_hint
        return team

    max_col = ws.max_column
    if max_col < 5:
        raise ValueError("対戦表の列数が少なすぎます（想定: '試合','開始','終了', 以降コート毎に2列）")
    courts = (max_col - 3) // 2
    if courts <= 0:
        raise ValueError("対戦表からコート数を推定できません")

    matches: list[Match] = []
    detected_rounds: set[int] = set()

    def parse_round_number(value: Any) -> int | None:
        if value is None:
            return None
        if isinstance(value, int):
            return int(value)
        if isinstance(value, float):
            # openpyxl may return numeric cells as float
            if value.is_integer():
                return int(value)
            return None
        if isinstance(value, str):
            s = value.strip()
            if not s:
                return None
            # Accept plain digits or strings like '第12試合'
            try:
                return int(s)
            except Exception:
                pass
            import re

            m = re.search(r"\d+", s)
            if not m:
                return None
            try:
                return int(m.group(0))
            except Exception:
                return None
        return None

    for row_idx in range(2, ws.max_row + 1):
        round_val = ws.cell(row=row_idx, column=1).value
        if round_val is None or str(round_val).strip() == "":
            continue

        # The exported Excel can contain footer/summary rows (e.g. '表示試合数').
        # Treat only rows with a parsable round number as match rows.
        round_num = parse_round_number(round_val)
        if round_num is None:
            continue
        detected_rounds.add(round_num)

        start_raw = ws.cell(row=row_idx, column=2).value
        start_hhmm = _coerce_hhmm(start_raw)
        if not start_hhmm:
            base = _base_datetime_from_hhmm(fallback_start_time_hhmm)
            start_dt = base + timedelta(minutes=int(fallback_round_minutes) * (round_num - 1))
        else:
            start_dt = _base_datetime_from_hhmm(start_hhmm)

        for court in range(1, courts + 1):
            col_team1 = 3 + (court - 1) * 2 + 1
            col_team2 = col_team1 + 1
            t1 = ws.cell(row=row_idx, column=col_team1).value
            t2 = ws.cell(row=row_idx, column=col_team2).value
            if t1 is None and t2 is None:
                continue
            t1n, t1m = split_team_cell(t1)
            t2n, t2m = split_team_cell(t2)
            t1s = t1n
            t2s = t2n
            if not t1s and not t2s:
                continue
            if not t1s or not t2s:
                raise ValueError(f"対戦表に片側だけチーム名があります: round={round_num} court={court}")
            team1 = get_team(t1s, members_hint=t1m)
            team2 = get_team(t2s, members_hint=t2m)
            matches.append(Match(round_num=round_num, court=court, team1=team1, team2=team2, start_time=start_dt))

    num_rounds = max(detected_rounds) if detected_rounds else 0
    teams = list(teams_by_name.values())
    return matches, teams, num_rounds, courts


def load_schedule_from_short_list_xlsx(
    workbook_path: str,
    *,
    sheet_name: str = "対戦一覧_短縮",
    fallback_start_time_hhmm: str = DEFAULT_START_TIME_HHMM,
    fallback_round_minutes: int = DEFAULT_ROUND_MINUTES,
) -> tuple[list[Match], list[Team], int, int]:
    """Load matches from a 'short list' sheet in an external workbook (xlsm/xlsx).

    Expected columns (typical):
      試合, コート, 時間, ペア名, 選手名, 相手ペア名, 相手選手名

    We infer rounds by grouping matches with the same start time (in appearance order).
    """

    # openpyxl emits a noisy warning when a workbook has a print-area defined name
    # that it can't re-apply exactly; it's safe for our use (we only read cell values).
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message=r"Print area cannot be set to Defined name:.*",
            category=UserWarning,
        )
        wb = openpyxl.load_workbook(workbook_path, data_only=True)

    preferred_sheet_names = [
        sheet_name,
        "対戦一覧_短縮",
        "対戦一覧短縮（試合順）",
        "対戦一覧短縮（チーム順）",
        "全対戦リスト",
    ]
    ws = None
    for s in preferred_sheet_names:
        if s in wb.sheetnames:
            ws = wb[s]
            break
    if ws is None:
        raise ValueError(
            "入力Excelに短縮対戦リストが見つかりません。"
            f"探した候補: {preferred_sheet_names} / 実際のシート: {wb.sheetnames}"
        )

    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    def _norm(v: Any) -> str:
        return str(v).strip() if v is not None else ""

    def _find_col(pred) -> int | None:
        for idx, h in enumerate(header_row):
            if pred(_norm(h)):
                return idx
        return None

    def _parse_match_no(value: Any) -> int | None:
        if value is None:
            return None
        # Typical formats: '第04', '第4', 4
        s = str(value).strip()
        if not s:
            return None
        s = s.replace("試合", "").replace("第", "").replace("回", "").strip()
        # keep only leading digits
        digits = "".join(ch for ch in s if ch.isdigit())
        if not digits:
            return None
        try:
            n = int(digits)
        except Exception:
            return None
        return n if n > 0 else None

    col_match_no = _find_col(lambda s: s == "試合" or "試合" in s)
    col_court = _find_col(lambda s: "コート" in s)  # required
    col_time = _find_col(lambda s: ("時間" in s) or (s == "開始") or ("開始" in s and "終了" not in s))
    col_pair = _find_col(
        lambda s: (
            s == "ペア名"
            or ("ペア名" in s and "相手" not in s)
            or ("チーム" in s and "相手" not in s and ("クラス" in s or "+" in s))
            or s == "チーム＋クラス"
        )
    )
    col_members = _find_col(lambda s: (s == "選手名" or s == "氏名" or ("選手名" in s and "相手" not in s) or ("氏名" in s and "相手" not in s)))
    col_opp_pair = _find_col(
        lambda s: (
            ("相手" in s and "ペア名" in s)
            or s == "相手ペア名"
            or s == "相手チーム＋クラス"
            or ("相手" in s and "チーム" in s and ("クラス" in s or "+" in s))
        )
    )
    col_opp_members = _find_col(lambda s: ("相手" in s and ("選手名" in s or "氏名" in s)) or s in ("相手選手名", "相手氏名"))

    if col_court is None or col_pair is None or col_opp_pair is None:
        raise ValueError(
            "短縮対戦リストの必須列が見つかりません（最低限: コート, ペア名, 相手ペア名）。"
            f"ヘッダー: {header_row}"
        )

    teams_by_name: dict[str, Team] = {}

    def _get_team(name: str, members: str) -> Team:
        nm = name.strip()
        if nm in teams_by_name:
            t = teams_by_name[nm]
        else:
            t = Team(name=nm, members="", level="", group="")
            teams_by_name[nm] = t
        if members and not t.members:
            t.members = members
        return t

    last_time_hhmm: str | None = None
    max_court = 0
    matches: list[Match] = []
    times_seen: list[str] = []
    rounds_seen: list[int] = []
    seen_match_keys: set[tuple[str, int, frozenset[str]]] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        court_raw = row[col_court] if col_court < len(row) else None
        if court_raw is None or str(court_raw).strip() == "":
            continue
        try:
            court = int(float(court_raw))
        except Exception:
            # skip rows that don't look like matches
            continue
        if court <= 0:
            continue

        match_no_raw = row[col_match_no] if (col_match_no is not None and col_match_no < len(row)) else None
        match_no = _parse_match_no(match_no_raw)

        time_raw = row[col_time] if (col_time is not None and col_time < len(row)) else None
        time_hhmm = _coerce_hhmm(time_raw)
        if not time_hhmm:
            # Some workbooks omit time in repeated rows; carry forward.
            if last_time_hhmm:
                time_hhmm = last_time_hhmm
            else:
                time_hhmm = fallback_start_time_hhmm
        last_time_hhmm = time_hhmm

        pair_name = _norm(row[col_pair] if col_pair < len(row) else "")
        opp_pair_name = _norm(row[col_opp_pair] if col_opp_pair < len(row) else "")
        if not pair_name or not opp_pair_name:
            continue

        members = _norm(row[col_members] if (col_members is not None and col_members < len(row)) else "")
        opp_members = _norm(row[col_opp_members] if (col_opp_members is not None and col_opp_members < len(row)) else "")

        start_dt = _base_datetime_from_hhmm(time_hhmm)

        team1 = _get_team(pair_name, members)
        team2 = _get_team(opp_pair_name, opp_members)

        key = (time_hhmm, court, frozenset({team1.name, team2.name}))
        if key in seen_match_keys:
            # Likely a "both perspectives" list (two rows per match). Avoid double-counting.
            continue
        seen_match_keys.add(key)

        matches.append(Match(round_num=match_no or 0, court=court, team1=team1, team2=team2, start_time=start_dt))
        max_court = max(max_court, court)
        times_seen.append(time_hhmm)
        if match_no:
            rounds_seen.append(match_no)

    teams = list(teams_by_name.values())
    # Infer level/group from name when the summary workbook doesn't carry them.
    for t in teams:
        if not t.level:
            levels = [c for c in t.name if c in "ABC"]
            t.level = levels[0] if levels else ""
        if not t.group:
            # Keep the same style as create_team_list.py: strip trailing digits.
            try:
                t.group = t.name.rstrip("0123456789")
            except Exception:
                t.group = t.group or ""
    if rounds_seen:
        # Prefer explicit match numbers if present (most accurate for ordering).
        num_rounds = max(rounds_seen)
    else:
        # Fallback: assign round numbers by chronological time order.
        unique_times = sorted(set(times_seen), key=_base_datetime_from_hhmm)
        time_to_round: dict[str, int] = {t: i + 1 for i, t in enumerate(unique_times)}
        for m in matches:
            m.round_num = time_to_round[m.start_time.strftime("%H:%M")]
        num_rounds = len(unique_times)

    matches.sort(key=lambda m: (m.round_num, m.court))
    courts = max_court
    refresh_team_stats(teams, matches)
    return matches, teams, num_rounds, courts


def write_matches_into_summary_sheet_grid(
    wb: openpyxl.Workbook,
    matches: list[Match],
    *,
    target_sheet_name: str = "集計表",
    header_row: int = 2,
    first_match_row: int = 3,
    rows_per_match: int = 3,
    match_no_col: int = 1,
    time_col: int = 2,
    first_court_col: int = 3,
    clear_scores: bool = True,
) -> tuple[int, int]:
    """Fill a macro-style summary sheet grid (3 rows per match).

    Typical layout in the summary workbook:
    - Row `first_match_row`: match no, start time, and pair names per court (2 cols per court)
    - Row +1: formulas (e.g., end time, VLOOKUP for members)
    - Row +2: score entry row

    This function only writes the pair-name rows + (optionally) clears score cells.
    It deliberately avoids overwriting formula rows.

    Returns (num_rounds_written, courts_detected).
    """

    if target_sheet_name not in wb.sheetnames:
        raise ValueError(f"target sheet not found: {target_sheet_name}")
    ws = wb[target_sheet_name]
    if not matches:
        raise ValueError("matches is empty")

    # Infer court count from header row (e.g., 1,2,3... at every other column)
    courts = 0
    col = int(first_court_col)
    while True:
        v = ws.cell(row=int(header_row), column=col).value
        if v is None or str(v).strip() == "":
            break
        try:
            n = int(float(v))
        except Exception:
            break
        courts = max(courts, n)
        col += 2
    if courts <= 0:
        # Fallback to observed max court
        courts = max(int(m.court) for m in matches)

    # Prefer explicit round count from matches.
    num_rounds = max(int(m.round_num) for m in matches)
    by_slot: dict[tuple[int, int], Match] = {(int(m.round_num), int(m.court)): m for m in matches}

    for rnd in range(1, num_rounds + 1):
        base_row = int(first_match_row) + (rnd - 1) * int(rows_per_match)

        ws.cell(row=base_row, column=int(match_no_col)).value = rnd

        # Start time: take the earliest match time in this round.
        round_matches = [by_slot[(rnd, c)] for c in range(1, courts + 1) if (rnd, c) in by_slot]
        if round_matches:
            start_dt = min(m.start_time for m in round_matches)
            try:
                ws.cell(row=base_row, column=int(time_col)).value = start_dt.time()
            except Exception:
                ws.cell(row=base_row, column=int(time_col)).value = start_dt
        else:
            ws.cell(row=base_row, column=int(time_col)).value = None

        for court in range(1, courts + 1):
            c1 = int(first_court_col) + (court - 1) * 2
            c2 = c1 + 1
            m = by_slot.get((rnd, court))
            if m:
                ws.cell(row=base_row, column=c1).value = m.team1.name
                ws.cell(row=base_row, column=c2).value = m.team2.name
            else:
                ws.cell(row=base_row, column=c1).value = None
                ws.cell(row=base_row, column=c2).value = None

            if clear_scores:
                score_row = base_row + 2
                ws.cell(row=score_row, column=c1).value = None
                ws.cell(row=score_row, column=c2).value = None

    return num_rounds, courts


@app.command()
def sync_pairs_from_team_list(
    schedule_file: str = typer.Option(..., help="更新対象のスケジュールExcel（'ペア一覧' を含むこと）"),
    team_list_file: str = typer.Option(..., help="最新のチーム一覧Excel（チームリスト.xlsx 等。氏名/レベル/グループのマスター）"),
    output_file: str = typer.Option("", help="出力Excel。空なら schedule_file の末尾に _pairsynced を付ける"),
):
    """Sync the 'ペア一覧' sheet in a schedule workbook from a team list workbook.

    Use this when member names change frequently and you want a single source of truth
    (e.g. 集計表 → create_team_list.py でチームリスト.xlsxを作る → このコマンドでスケジュールへ反映)。

    Notes:
    - Only updates rows that already exist in 'ペア一覧' (by ペア名 match).
    - Does not touch '対戦表' matchups.
    """

    in_path = Path(schedule_file)
    if not in_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {schedule_file}")
    tl_path = Path(team_list_file)
    if not tl_path.exists():
        raise typer.BadParameter(f"チーム一覧ファイルが見つかりません: {team_list_file}")

    out_path = Path(output_file) if output_file else in_path.with_name(f"{in_path.stem}_pairsynced{in_path.suffix}")

    # Load the latest teams from the team list workbook
    teams = load_teams(str(tl_path))
    if not teams:
        raise typer.BadParameter(f"チーム一覧が空です: {team_list_file}")
    latest_by_name: dict[str, Team] = {t.name: t for t in teams if t.name}

    try:
        wb = openpyxl.load_workbook(str(in_path))
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {schedule_file}"
        )
    if "ペア一覧" not in wb.sheetnames:
        raise typer.BadParameter("Excelに 'ペア一覧' シートが見つかりません")

    ws = wb["ペア一覧"]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    def _find_col(*candidates: str) -> int | None:
        for idx, v in enumerate(header, start=1):
            if isinstance(v, str) and v.strip() in candidates:
                return idx
        return None

    col_name = _find_col("ペア名") or 1
    col_members = _find_col("選手名", "氏名") or 2
    col_level = _find_col("レベル") or 3
    col_group = _find_col("グループ") or 4

    updated = 0
    missing_in_team_list: set[str] = set()

    for row in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=col_name).value
        name = str(name_cell).strip() if name_cell is not None else ""
        if not name:
            continue
        src = latest_by_name.get(name)
        if not src:
            missing_in_team_list.add(name)
            continue

        ws.cell(row=row, column=col_members).value = src.members
        ws.cell(row=row, column=col_level).value = src.level
        ws.cell(row=row, column=col_group).value = src.group
        updated += 1

    wb.save(str(out_path))
    wb.close()

    print(f"ペア一覧を同期しました: {out_path}")
    print(f"  更新行数: {updated} / チーム一覧件数: {len(latest_by_name)}")
    if missing_in_team_list:
        # keep output short
        sample = sorted(list(missing_in_team_list))[:8]
        suffix = " ..." if len(missing_in_team_list) > len(sample) else ""
        print(f"  注意: チーム一覧に無いペア名がペア一覧にあります: {', '.join(sample)}{suffix}")


@app.command()
def export_from_summary(
    input_file: str = typer.Option(..., help="集計表（.xlsx/.xlsm）。'対戦一覧_短縮' シートを含むこと"),
    sheet_name: str = typer.Option("対戦一覧_短縮", help="読み込む短縮対戦リストのシート名"),
    output_file: str = typer.Option("", help="出力Excel。空なら input_file と同名で _from_summary.xlsx"),
    html_passcode: str = typer.Option("", help="HTMLの簡易ロック用パスコード（注意: 完全な暗号化ではありません）"),
    include_members: bool = typer.Option(True, help="HTMLに選手名（氏名）を含める（短縮シートに氏名がある前提）"),
    wall_html: bool = typer.Option(False, help="壁貼り用（コート別）HTMLも出力する（印刷前提・JS不要）"),
    wall_courts_per_page: int = typer.Option(2, help="壁貼り用HTMLの1ページあたりコート数（1〜3推奨）。標準は2"),
    wall_team_color: list[str] = typer.Option(
        [],
        "--wall-team-color",
        help="壁貼り用のチーム色分け（複数可）。例: --wall-team-color '蛇口:red' --wall-team-color '上海:blue'",
    ),
    wall_auto_colors: bool = typer.Option(
        True,
        "--wall-auto-colors/--no-wall-auto-colors",
        help="壁貼り用HTMLで、未指定チームも自動で色付けする（デフォルトON）",
    ),
    wall_cell_background: bool = typer.Option(
        False,
        "--wall-cell-background/--no-wall-cell-background",
        help="壁貼り用HTMLのチームセル背景色を塗る（デフォルトOFF）。背景印刷がOFFになりがちなのでOFF推奨",
    ),
    allow_court_gaps: bool = typer.Option(True, help="空きコートを許容する（集計表由来ではON推奨）"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="1試合の時間（分）。終了時刻の計算に使用"),
    excel_include_members: bool = typer.Option(False, help="Excelの『対戦表』セルに選手名（氏名）を改行で表示する"),
    excel_members_below: bool = typer.Option(False, help="Excelの『対戦表』を2行構成にする（上=ペア名、下=選手名）"),
    excel_members_vlookup: bool = typer.Option(False, help="『選手名』行をペア一覧からVLOOKUPで自動表示する（ペア名セルをキー）。※Excelで開いて計算が必要"),
    start_time: str = typer.Option(DEFAULT_START_TIME_HHMM, help="開始時刻 (HH:MM) ※時間列が欠けている場合の補助"),
):
    """Export schedule Excel/HTML from an existing summary workbook.

    This is for workflows where the "source of truth" is an external 集計表 (xlsm)
    that already contains a match list sheet (対戦一覧_短縮).
    """

    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")
    try:
        _parse_hhmm(start_time)
    except ValueError as e:
        raise typer.BadParameter(str(e))

    in_path = Path(input_file)
    if not in_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {input_file}")

    out_path = Path(output_file) if output_file else in_path.with_name(f"{in_path.stem}_from_summary.xlsx")

    try:
        matches, teams, num_rounds, courts = load_schedule_from_short_list_xlsx(
            str(in_path),
            sheet_name=str(sheet_name),
            fallback_start_time_hhmm=start_time,
            fallback_round_minutes=int(round_minutes),
        )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {input_file}"
        )

    if num_rounds <= 0 or courts <= 0 or not matches:
        raise RuntimeError("短縮対戦リストから試合を読み取れませんでした（列名やシート名を確認してください）")

    # Keep original times from the summary sheet (do not override by uniform round minutes).
    write_to_excel_like_summary(
        matches,
        teams,
        str(out_path),
        bool(allow_court_gaps),
        num_rounds,
        courts,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        excel_include_members=bool(excel_include_members),
        excel_members_below=bool(excel_members_below),
        excel_members_vlookup=bool(excel_members_vlookup),
        normalize_round_times=False,
    )

    html_output_path = out_path.with_suffix('.html')
    write_personal_schedule_html(
        matches,
        teams,
        str(html_output_path),
        num_rounds=num_rounds,
        courts=courts,
        html_passcode=html_passcode or None,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        include_members=bool(include_members),
    )
    if wall_html:
        wall_rules = _parse_wall_team_color_rules(list(wall_team_color))
        wall_output_path = out_path.with_name(f"{out_path.stem}_wall.html")
        write_wall_schedule_html(
            matches,
            str(wall_output_path),
            num_rounds=num_rounds,
            courts=courts,
            start_time_hhmm=start_time,
            round_minutes=int(round_minutes),
            courts_per_page=int(wall_courts_per_page),
            team_color_rules=wall_rules,
            auto_team_colors=bool(wall_auto_colors),
            cell_background=bool(wall_cell_background),
        )

    print(f"Excel出力: {out_path}")
    print(f"HTML一式(対戦表+短縮+個人): {html_output_path}")
    if wall_html:
        print(f"壁貼り用HTML(コート別): {wall_output_path}")
    print(f"読込元: {in_path} / シート: {sheet_name}")
    print(f"総試合数: {len(matches)} / ラウンド数: {num_rounds} / コート数: {courts}")


@app.command()
def fill_summary_grid(
    input_file: str = typer.Option(..., help="集計表（.xlsx/.xlsm）。'対戦一覧_短縮' から '集計表' の対戦表入力欄に流し込む"),
    sheet_name: str = typer.Option("対戦一覧_短縮", help="読み込む短縮対戦リストのシート名"),
    target_sheet: str = typer.Option("集計表", help="流し込み先シート名（3行=1試合の入力シート）"),
    output_file: str = typer.Option("", help="出力ファイル。空なら input_file と同名で _filled を付ける"),
    clear_scores: bool = typer.Option(True, help="流し込み時に得点欄をクリアする"),
    start_time: str = typer.Option(DEFAULT_START_TIME_HHMM, help="開始時刻 (HH:MM) ※時間列が欠けている場合の補助"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="1試合の時間（分）。時間列が欠けている場合の補助"),
):
    """Fill the summary workbook's match-input grid from 対戦一覧_短縮.

    This automates the "23ラウンド分コピペ" work: pair names are written into the
    macro-friendly grid on the target sheet while preserving formulas/macros.
    """

    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")
    try:
        _parse_hhmm(start_time)
    except ValueError as e:
        raise typer.BadParameter(str(e))

    in_path = Path(input_file)
    if not in_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {input_file}")

    out_path = Path(output_file) if output_file else in_path.with_name(f"{in_path.stem}_filled{in_path.suffix}")

    try:
        matches, _teams, _num_rounds, _courts = load_schedule_from_short_list_xlsx(
            str(in_path),
            sheet_name=str(sheet_name),
            fallback_start_time_hhmm=start_time,
            fallback_round_minutes=int(round_minutes),
        )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {input_file}"
        )

    if not matches:
        raise RuntimeError("短縮対戦リストから試合を読み取れませんでした（列名やシート名を確認してください）")

    try:
        with warnings.catch_warnings():
            # openpyxl emits a noisy warning when a workbook has a print-area defined name
            # that it can't re-apply exactly; it's safe for our use (we only write cell values).
            warnings.filterwarnings("ignore", message=r"Print area cannot be set to Defined name:.*", category=UserWarning)
            wb = openpyxl.load_workbook(
                str(in_path),
                data_only=False,
                keep_vba=(in_path.suffix.lower() == ".xlsm"),
            )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {input_file}"
        )

    num_rounds_written, courts_detected = write_matches_into_summary_sheet_grid(
        wb,
        matches,
        target_sheet_name=str(target_sheet),
        clear_scores=bool(clear_scores),
    )

    wb.save(str(out_path))
    wb.close()

    print(f"流し込み完了: {out_path}")
    print(f"  対象シート: {target_sheet}")
    print(f"  ラウンド数: {num_rounds_written} / コート数: {courts_detected} / 得点クリア: {clear_scores}")


@app.command()
def fill_summary_grid_from_xlsx(
    schedule_file: str = typer.Option(..., help="（手で編集した後の）スケジュールExcel。'対戦表' を含むこと"),
    summary_file: str = typer.Option(..., help="流し込み先の集計表（.xlsx/.xlsm）。マクロ/数式は保持"),
    target_sheet: str = typer.Option("集計表", help="流し込み先シート名（3行=1試合の入力シート）"),
    output_file: str = typer.Option("", help="出力ファイル。空なら summary_file と同名で _filled_from_xlsx を付ける"),
    clear_scores: bool = typer.Option(True, help="流し込み時に得点欄をクリアする"),
    start_time: str = typer.Option(DEFAULT_START_TIME_HHMM, help="開始時刻 (HH:MM) ※対戦表に開始が無い場合の補助"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="1試合の時間（分）。開始が無い場合の補助"),
):
    """Fill a summary workbook's match-input grid from an edited schedule Excel.

    This matches the common workflow:
      生成した schedule.xlsx を手で微修正 → 集計表(xlsm)の入力グリッドへ流し込み →
      集計表マクロで 対戦一覧_短縮 を再生成 → Pythonで最終配布物(HTML/Excel)を出力。
    """

    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")
    try:
        _parse_hhmm(start_time)
    except ValueError as e:
        raise typer.BadParameter(str(e))

    sched_path = Path(schedule_file)
    if not sched_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {schedule_file}")
    summ_path = Path(summary_file)
    if not summ_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {summary_file}")

    out_path = Path(output_file) if output_file else summ_path.with_name(f"{summ_path.stem}_filled_from_xlsx{summ_path.suffix}")

    try:
        matches, _teams, _num_rounds, _courts = load_schedule_from_xlsx(
            str(sched_path),
            fallback_start_time_hhmm=str(start_time),
            fallback_round_minutes=int(round_minutes),
        )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {schedule_file}"
        )

    if not matches:
        raise RuntimeError("スケジュールExcelから試合を読み取れませんでした（'対戦表' を確認してください）")

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message=r"Print area cannot be set to Defined name:.*", category=UserWarning)
            wb = openpyxl.load_workbook(
                str(summ_path),
                data_only=False,
                keep_vba=(summ_path.suffix.lower() == ".xlsm"),
            )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {summary_file}"
        )

    num_rounds_written, courts_detected = write_matches_into_summary_sheet_grid(
        wb,
        matches,
        target_sheet_name=str(target_sheet),
        clear_scores=bool(clear_scores),
    )

    wb.save(str(out_path))
    wb.close()

    print(f"流し込み完了: {out_path}")
    print(f"  参照スケジュール: {sched_path}")
    print(f"  対象集計表: {summ_path} / シート: {target_sheet}")
    print(f"  ラウンド数: {num_rounds_written} / コート数: {courts_detected} / 得点クリア: {clear_scores}")


@app.command()
def template(
    output_file: str = typer.Option("チームリスト_テンプレ.xlsx", help="ヘッダーのみのチーム一覧テンプレートExcelを出力"),
    sheet_name: str = typer.Option(TEAM_LIST_TEMPLATE_SHEET_NAME, help="作成するシート名"),
):
    data = build_team_list_template_bytes(sheet_name=sheet_name)
    Path(output_file).write_bytes(data)
    print(f"テンプレート出力: {output_file} (sheet='{sheet_name}')")


@app.command()
def sample_xlsx(
    output_file: str = typer.Option("チームリスト_サンプル.xlsx", help="ダミーデータ入りのサンプルチーム一覧Excelを出力"),
    sheet_name: str = typer.Option(TEAM_LIST_SAMPLE_SHEET_NAME, help="作成するシート名"),
):
    data = build_team_list_sample_bytes(sheet_name=sheet_name)
    Path(output_file).write_bytes(data)
    print(f"サンプル出力: {output_file} (sheet='{sheet_name}')")

@app.command()
def generate_schedule(
    input_file: str = typer.Option("チームリスト.xlsx", help="チーム一覧Excel"),
    output_file: str = typer.Option("graph_schedule.xlsx", help="出力ファイル"),
    num_rounds: int = 23,
    courts: int = 15,
    diversity_attempts: int = typer.Option(1, help="分散最大化の試行回数"),
    graph_mode: bool = typer.Option(True, help="グラフ構築モードを使用 (必須条件安定化)"),
    allow_court_gaps: bool = typer.Option(False, help="途中ラウンドの空きコートを許容するか（審判運用のため通常はOFF推奨）"),
    max_consecutive: int = typer.Option(2, help="最大連戦数（2推奨）。満たせない場合は自動で3に緩和（2 or 3）"),
    relax_max_consecutive: bool = typer.Option(
        True,
        "--relax-max-consecutive/--no-relax-max-consecutive",
        help="連戦上限を満たせない場合に自動緩和する（max_consecutive=2 のとき 3 に緩和）。デフォルトON",
    ),
    matches_per_team: int = typer.Option(0, help="各ペアの試合数。0で自動（全員同数を最優先）。例: 6"),
    html_passcode: str = typer.Option("", help="HTMLの簡易ロック用パスコード（注意: 完全な暗号化ではありません）"),
    include_members: bool = typer.Option(True, help="HTMLに選手名（氏名）を含める（対戦表/短縮/個人に反映）"),
    wall_html: bool = typer.Option(False, help="壁貼り用（コート別）HTMLも出力する（印刷前提・JS不要）"),
    wall_courts_per_page: int = typer.Option(2, help="壁貼り用HTMLの1ページあたりコート数（1〜3推奨）。標準は2"),
    wall_team_color: list[str] = typer.Option(
        [],
        "--wall-team-color",
        help="壁貼り用のチーム色分け（複数可）。例: --wall-team-color '蛇口:red' --wall-team-color '上海:blue'",
    ),
    wall_auto_colors: bool = typer.Option(
        True,
        "--wall-auto-colors/--no-wall-auto-colors",
        help="壁貼り用HTMLで、未指定チームも自動で色付けする（デフォルトON）",
    ),
    wall_cell_background: bool = typer.Option(
        False,
        "--wall-cell-background/--no-wall-cell-background",
        help="壁貼り用HTMLのチームセル背景色を塗る（デフォルトOFF）。背景印刷がOFFになりがちなのでOFF推奨",
    ),
    excel_include_members: bool = typer.Option(False, help="Excelの『対戦表』セルに選手名（氏名）を改行で表示する（手修正しやすい。配布する場合は注意）"),
    excel_members_below: bool = typer.Option(True, help="Excelの『対戦表』を2行構成にする（上=ペア名、下=選手名）"),
    excel_members_vlookup: bool = typer.Option(True, help="『選手名』行をペア一覧からVLOOKUPで自動表示する（ペア名セルをキー）。※Excelで開いて計算が必要"),
    start_time: str = typer.Option(DEFAULT_START_TIME_HHMM, help="開始時刻 (HH:MM)"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="1ラウンドの時間（分）"),
):
    if not graph_mode:
        raise typer.BadParameter("現在は graph_mode=True のみをサポートします")
    global TARGET_MATCHES_PER_TEAM

    probe_teams = load_teams(input_file)
    if matches_per_team < 0:
        raise typer.BadParameter("matches_per_team は 0 以上を指定してください")
    auto_mode = matches_per_team == 0
    if auto_mode:
        initial_target = compute_auto_matches_per_team(len(probe_teams), num_rounds, courts)
    else:
        initial_target = matches_per_team

    if max_consecutive not in (2, 3):
        raise typer.BadParameter("max_consecutive は 2 または 3 を指定してください")

    capacity = num_rounds * courts
    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")
    try:
        _parse_hhmm(start_time)
    except ValueError as e:
        raise typer.BadParameter(str(e))

    best_matches: List[Match] | None = None
    best_teams: List[Team] | None = None
    best_score: int = -1
    tried_targets: Set[int] = set()
    target_candidates = [initial_target] if not auto_mode else list(range(initial_target, 0, -1))
    for target in target_candidates:
        # Ensure total matches is an integer.
        if (len(probe_teams) % 2 == 1) and (target % 2 == 1):
            target -= 1
        if target <= 0 or target in tried_targets:
            continue
        tried_targets.add(target)

        expected_matches_total = expected_total_matches(len(probe_teams), target)
        if (len(probe_teams) * target) % 2 != 0:
            if auto_mode:
                continue
            raise typer.BadParameter(
                f"不正な組合せ: ペア数={len(probe_teams)} と 目標={target} では総試合数が整数になりません（ペア数が奇数の時は試合数は偶数が必要）"
            )
        if expected_matches_total > capacity:
            if auto_mode:
                continue
            raise typer.BadParameter(
                f"容量不足: courts*num_rounds={capacity} 試合枠に対し、必要試合数={expected_matches_total}（ペア数={len(probe_teams)}, 目標={target}試合/ペア）"
            )

        TARGET_MATCHES_PER_TEAM = target
        if target != initial_target:
            print(f"自動調整: 目標試合数/ペアを {target} に下げて再試行します")
        print(f"目標試合数/ペア: {TARGET_MATCHES_PER_TEAM}（自動={auto_mode}） / 必要試合数 {expected_matches_total} / 容量 {capacity}")

        best_matches = None
        best_teams = None
        best_score = -1
        for attempt in range(diversity_attempts):
            teams = load_teams(input_file)
            # まずグラフ方式を試行
            graph_ok = False
            try:
                matches = schedule_matches_graph(teams, num_rounds, courts, seed=attempt)
                refresh_team_stats(teams, matches)
                if all(t.matches == TARGET_MATCHES_PER_TEAM for t in teams) and len(matches) == expected_total_matches(len(teams), TARGET_MATCHES_PER_TEAM):
                    graph_ok = True
            except Exception as e:
                print(f"グラフ試行 {attempt} 失敗: {e}")
            if not graph_ok:
                # フォールバック: ヒューリスティック
                print(f"フォールバックヒューリスティック使用 (試行 {attempt})")
                teams = load_teams(input_file)
                matches = schedule_matches_heuristic(teams, num_rounds, courts, seed=attempt)
                refresh_team_stats(teams, matches)
            # 縦方向分散 + 帯再配置後処理
            if all(t.matches == TARGET_MATCHES_PER_TEAM for t in teams) and len(matches) == expected_total_matches(len(teams), TARGET_MATCHES_PER_TEAM):
                matches = rebalance_vertical_distribution(matches, teams, num_rounds, courts)
                matches = enforce_segments_and_quotas(matches, teams, num_rounds, courts)
                matches = balanced_round_reassignment(matches, num_rounds, courts)
                matches = boost_group_diversity(matches, teams)
                matches = reduce_back_to_back(matches, num_rounds, courts)
                matches = tighten_level_bands(matches, num_rounds, courts)
                # コート衝突修復
                collisions = detect_collisions(matches)
                if collisions:
                    print(f"衝突スロット数: {len(collisions)} → 修復試行")
                    matches = repair_collisions(matches, num_rounds, courts)
                    print(f"After repair, max round: {max(m.round_num for m in matches)}, min round: {min(m.round_num for m in matches)}")
                    print(f"After repair, max court: {max(m.court for m in matches)}, min court: {min(m.court for m in matches)}")
                    after = detect_collisions(matches)
                    if after:
                        print(f"修復後も残る衝突: {len(after)}")
                # Ensure we never exceed physical court capacity even if collision repair is imperfect.
                matches = normalize_round_capacity(matches, num_rounds, courts)
                matches = ensure_round_one_full(matches, num_rounds, courts)
                if not allow_court_gaps:
                    matches = eliminate_mid_session_court_gaps(matches, num_rounds, courts)
                matches = compact_court_usage(matches, num_rounds, courts)

                def _max_team_streak(ms: List[Match]) -> int:
                    rounds_map: Dict[str, List[int]] = defaultdict(list)
                    for m in ms:
                        rounds_map[m.team1.name].append(m.round_num)
                        rounds_map[m.team2.name].append(m.round_num)
                    best = 0
                    for rs in rounds_map.values():
                        if not rs:
                            continue
                        sr = sorted(rs)
                        cur = 1
                        mx = 1
                        for i in range(1, len(sr)):
                            if sr[i] == sr[i - 1] + 1:
                                cur += 1
                                mx = max(mx, cur)
                            else:
                                cur = 1
                        best = max(best, mx)
                    return best

                def _consecutive_optimize(ms: List[Match], limit: int) -> List[Match]:
                    for _ in range(3):
                        if _max_team_streak(ms) <= limit:
                            break
                        ms = reduce_max_consecutive_streak(ms, num_rounds, courts, max_consecutive=limit)
                        ms = normalize_round_capacity(ms, num_rounds, courts)
                        ms = ensure_round_one_full(ms, num_rounds, courts)
                        if not allow_court_gaps:
                            ms = eliminate_mid_session_court_gaps(ms, num_rounds, courts)
                        ms = compact_court_usage(ms, num_rounds, courts)
                    return ms

                matches = _consecutive_optimize(matches, limit=max_consecutive)
                if max_consecutive == 2 and _max_team_streak(matches) > 2:
                    if relax_max_consecutive:
                        print("連戦上限2が満たせないため、連戦上限3に緩和します")
                        matches = _consecutive_optimize(matches, limit=3)
            # 条件確認
            if any(t.matches != TARGET_MATCHES_PER_TEAM for t in teams):
                print(f"試行 {attempt}: 未達ペアあり -> スキップ")
                continue
            if len(matches) != expected_total_matches(len(teams), TARGET_MATCHES_PER_TEAM):
                print(f"試行 {attempt}: 総試合数 {len(matches)} 不一致")
                continue
            if (not relax_max_consecutive) and _max_team_streak(matches) > max_consecutive:
                print(
                    f"試行 {attempt}: 連戦上限{max_consecutive}を満たせません (最大連戦={_max_team_streak(matches)}) -> スキップ"
                )
                continue
            score = compute_diversity_score(teams)
            if score > best_score:
                best_score = score
                best_matches = matches
                best_teams = teams

        if best_matches is not None:
            break

    if best_matches is None:
        raise RuntimeError("全試行失敗: 条件を満たすスケジュールを構築できませんでした (グラフ+ヒューリスティック)")

    print(f"最大連戦: {_max_team_streak(best_matches)}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_path = Path(output_file)
    if base_path.suffix:
        stamped_name = f"{base_path.stem}_{timestamp}{base_path.suffix}"
    else:
        stamped_name = f"{base_path.name}_{timestamp}"
    final_output_path = base_path.with_name(stamped_name)

    refresh_team_stats(best_teams, best_matches)
    apply_round_times(best_matches, start_time_hhmm=start_time, round_minutes=int(round_minutes))
    write_to_excel_like_summary(
        best_matches,
        best_teams,
        str(final_output_path),
        allow_court_gaps,
        num_rounds,
        courts,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        excel_include_members=bool(excel_include_members),
        excel_members_below=bool(excel_members_below),
        excel_members_vlookup=bool(excel_members_vlookup),
    )
    html_output_path = final_output_path.with_suffix('.html')
    write_personal_schedule_html(
        best_matches,
        best_teams,
        str(html_output_path),
        num_rounds,
        courts,
        html_passcode=html_passcode or None,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        include_members=bool(include_members),
    )
    if wall_html:
        wall_rules = _parse_wall_team_color_rules(list(wall_team_color))
        wall_output_path = final_output_path.with_name(f"{final_output_path.stem}_wall.html")
        write_wall_schedule_html(
            best_matches,
            str(wall_output_path),
            num_rounds=num_rounds,
            courts=courts,
            start_time_hhmm=start_time,
            round_minutes=int(round_minutes),
            courts_per_page=int(wall_courts_per_page),
            team_color_rules=wall_rules,
            auto_team_colors=bool(wall_auto_colors),
            cell_background=bool(wall_cell_background),
        )
    print(f"スケジュール出力: {final_output_path}")
    print("  含まれる主要シート: 対戦表 / 個人スケジュール表 / 対戦一覧短縮（チーム順/試合順）")
    print(f"HTML一式(対戦表+短縮+個人): {html_output_path}")
    if wall_html:
        print(f"壁貼り用HTML(コート別): {wall_output_path}")
    print(f"総試合数: {len(best_matches)} / 期待 { expected_total_matches(len(best_teams), TARGET_MATCHES_PER_TEAM) }")
    print(f"分散スコア(総対戦グループ種類合計): {best_score}")
    for level in ['A','B','C']:
        level_matches = [m for m in best_matches if m.team1.level == level]
        print(f"レベル {level}: {len(level_matches)} 試合")
        lvl_teams = [t for t in best_teams if t.level == level]
        if lvl_teams:
            print(f"  チーム数 {len(lvl_teams)} / 平均試合数 {sum(t.matches for t in lvl_teams)/len(lvl_teams):.1f}")
        else:
            print("  チーム数 0")
    # 成功確認
    under = [t.name for t in best_teams if t.matches != TARGET_MATCHES_PER_TEAM]
    if under:
        print("警告: 未達ペア", under)
    else:
        print(f"全ペア{TARGET_MATCHES_PER_TEAM}試合達成 ✅")


@app.command()
def html_from_xlsx(
    input_file: str = typer.Option(..., help="（手で編集した後の）スケジュールExcelファイル。'対戦表' と 'ペア一覧' を含むこと"),
    output_file: str = typer.Option("", help="出力HTMLファイル。空なら input_file と同名で .html"),
    html_passcode: str = typer.Option("", help="HTMLの簡易ロック用パスコード（注意: 完全な暗号化ではありません）"),
    start_time: str = typer.Option(DEFAULT_START_TIME_HHMM, help="開始時刻 (HH:MM) ※Excelに開始が無い場合の補助"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="1ラウンドの時間（分） ※Excelに開始が無い場合の補助"),
    include_members: bool = typer.Option(False, help="HTMLに選手名（氏名）を含める。通常はOFF推奨"),
    wall_html: bool = typer.Option(False, help="壁貼り用（コート別）HTMLも出力する（印刷前提・JS不要）"),
    wall_courts_per_page: int = typer.Option(2, help="壁貼り用HTMLの1ページあたりコート数（1〜3推奨）。標準は2"),
    wall_team_color: list[str] = typer.Option(
        [],
        "--wall-team-color",
        help="壁貼り用のチーム色分け（複数可）。例: --wall-team-color '蛇口:red' --wall-team-color '上海:blue'",
    ),
    wall_auto_colors: bool = typer.Option(
        True,
        "--wall-auto-colors/--no-wall-auto-colors",
        help="壁貼り用HTMLで、未指定チームも自動で色付けする（デフォルトON）",
    ),
    wall_cell_background: bool = typer.Option(
        False,
        "--wall-cell-background/--no-wall-cell-background",
        help="壁貼り用HTMLのチームセル背景色を塗る（デフォルトOFF）。背景印刷がOFFになりがちなのでOFF推奨",
    ),
):
    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")
    try:
        _parse_hhmm(start_time)
    except ValueError as e:
        raise typer.BadParameter(str(e))

    in_path = Path(input_file)
    if not in_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {input_file}")
    out_path = Path(output_file) if output_file else in_path.with_suffix(".html")

    try:
        matches, teams, num_rounds, courts = load_schedule_from_xlsx(
            str(in_path),
            fallback_start_time_hhmm=start_time,
            fallback_round_minutes=int(round_minutes),
        )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {input_file}"
        )
    if num_rounds <= 0 or courts <= 0:
        raise RuntimeError("Excelからラウンド数/コート数を推定できませんでした")

    write_personal_schedule_html(
        matches,
        teams,
        str(out_path),
        num_rounds=num_rounds,
        courts=courts,
        html_passcode=html_passcode or None,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        include_members=bool(include_members),
    )
    if wall_html:
        wall_rules = _parse_wall_team_color_rules(list(wall_team_color))
        wall_out_path = out_path.with_name(f"{out_path.stem}_wall.html")
        write_wall_schedule_html(
            matches,
            str(wall_out_path),
            num_rounds=num_rounds,
            courts=courts,
            start_time_hhmm=start_time,
            round_minutes=int(round_minutes),
            courts_per_page=int(wall_courts_per_page),
            team_color_rules=wall_rules,
            auto_team_colors=bool(wall_auto_colors),
            cell_background=bool(wall_cell_background),
        )
    print(f"HTML出力: {out_path}")
    if wall_html:
        print(f"壁貼り用HTML(コート別): {wall_out_path}")


@app.command()
def score_sheets_from_summary(
    input_file: str = typer.Option(..., help="集計表（.xlsx/.xlsm）。'対戦一覧_短縮' シートを含むこと"),
    sheet_name: str = typer.Option("対戦一覧_短縮", help="読み込む短縮対戦リストのシート名"),
    output_file: str = typer.Option("", help="出力HTML。空なら input_file と同名で _score_sheets.html"),
    per_page: int = typer.Option(8, help="A4 1枚あたりの枚数（面付け数）。標準は8（半分カット運用向け）"),
    columns: int = typer.Option(2, help="面付けの列数。例: 2"),
    include_members: bool = typer.Option(True, help="得点記入表に選手名（氏名）も入れる"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="終了時刻の計算に使う（分）"),
):
    """Export printable per-match score sheets from a summary workbook."""

    if per_page <= 0:
        raise typer.BadParameter("per_page は 1 以上を指定してください")
    if columns <= 0:
        raise typer.BadParameter("columns は 1 以上を指定してください")
    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")

    in_path = Path(input_file)
    if not in_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {input_file}")

    out_path = Path(output_file) if output_file else in_path.with_name(f"{in_path.stem}_score_sheets.html")

    try:
        matches, teams, _num_rounds, _courts = load_schedule_from_short_list_xlsx(
            str(in_path),
            sheet_name=str(sheet_name),
            fallback_start_time_hhmm=DEFAULT_START_TIME_HHMM,
            fallback_round_minutes=int(round_minutes),
        )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {input_file}"
        )

    if not matches:
        raise RuntimeError("短縮対戦リストから試合を読み取れませんでした（列名やシート名を確認してください）")

    write_score_sheets_html(
        matches,
        teams,
        str(out_path),
        per_page=int(per_page),
        columns=int(columns),
        include_members=bool(include_members),
        round_minutes=int(round_minutes),
    )
    print(f"得点記入表HTML出力: {out_path}")


@app.command()
def score_sheets_from_xlsx(
    input_file: str = typer.Option(..., help="（手で編集した後の）スケジュールExcelファイル。'対戦表' と 'ペア一覧' を含むこと"),
    output_file: str = typer.Option("", help="出力HTML。空なら input_file と同名で _score_sheets.html"),
    per_page: int = typer.Option(8, help="A4 1枚あたりの枚数（面付け数）。標準は8（半分カット運用向け）"),
    columns: int = typer.Option(2, help="面付けの列数。例: 2"),
    include_members: bool = typer.Option(True, help="得点記入表に選手名（氏名）も入れる"),
    start_time: str = typer.Option(DEFAULT_START_TIME_HHMM, help="開始時刻 (HH:MM) ※Excelに開始が無い場合の補助"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="終了時刻の計算に使う（分） ※Excelに開始が無い場合の補助"),
):
    """Export printable per-match score sheets from an exported schedule workbook."""

    if per_page <= 0:
        raise typer.BadParameter("per_page は 1 以上を指定してください")
    if columns <= 0:
        raise typer.BadParameter("columns は 1 以上を指定してください")
    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")
    try:
        _parse_hhmm(start_time)
    except ValueError as e:
        raise typer.BadParameter(str(e))

    in_path = Path(input_file)
    if not in_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {input_file}")

    out_path = Path(output_file) if output_file else in_path.with_name(f"{in_path.stem}_score_sheets.html")

    try:
        matches, teams, _num_rounds, _courts = load_schedule_from_xlsx(
            str(in_path),
            fallback_start_time_hhmm=str(start_time),
            fallback_round_minutes=int(round_minutes),
        )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {input_file}"
        )
    if not matches:
        raise RuntimeError("Excelから試合を読み取れませんでした")

    write_score_sheets_html(
        matches,
        teams,
        str(out_path),
        per_page=int(per_page),
        columns=int(columns),
        include_members=bool(include_members),
        round_minutes=int(round_minutes),
    )
    print(f"得点記入表HTML出力: {out_path}")


@app.command()
def xlsx_from_xlsx(
    input_file: str = typer.Option(..., help="（手で編集した後の）スケジュールExcelファイル。'対戦表' と 'ペア一覧' を含むこと"),
    output_file: str = typer.Option("", help="出力Excelファイル。空なら input_file の末尾に _rebuilt を付ける"),
    allow_court_gaps: bool = typer.Option(False, help="途中ラウンドの空きコートを許容するか（審判運用のため通常はOFF推奨）"),
    start_time: str = typer.Option(DEFAULT_START_TIME_HHMM, help="開始時刻 (HH:MM) ※Excelに開始が無い場合の補助"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="1ラウンドの時間（分） ※Excelに開始が無い場合の補助"),
    excel_include_members: bool = typer.Option(False, help="Excelの『対戦表』セルに選手名（氏名）を改行で表示する（手修正しやすい。配布する場合は注意）"),
    excel_members_below: bool = typer.Option(False, help="Excelの『対戦表』を2行構成にする（上=ペア名、下=選手名）"),
    excel_members_vlookup: bool = typer.Option(False, help="『選手名』行をペア一覧からVLOOKUPで自動表示する（ペア名セルをキー）。※Excelで開いて計算が必要"),
):
    """Rebuild an Excel workbook from a manually edited schedule.

    Use this when you edited the exported Excel (matchups or member names) and want
    the derived sheets (short lists / personal schedule) to be regenerated.
    """
    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")
    try:
        _parse_hhmm(start_time)
    except ValueError as e:
        raise typer.BadParameter(str(e))

    in_path = Path(input_file)
    if not in_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {input_file}")

    out_path = Path(output_file) if output_file else in_path.with_name(f"{in_path.stem}_rebuilt{in_path.suffix}")

    try:
        matches, teams, num_rounds, courts = load_schedule_from_xlsx(
            str(in_path),
            fallback_start_time_hhmm=start_time,
            fallback_round_minutes=int(round_minutes),
        )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {input_file}"
        )
    if num_rounds <= 0 or courts <= 0:
        raise RuntimeError("Excelからラウンド数/コート数を推定できませんでした")

    write_to_excel_like_summary(
        matches,
        teams,
        str(out_path),
        bool(allow_court_gaps),
        num_rounds,
        courts,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        excel_include_members=bool(excel_include_members),
        excel_members_below=bool(excel_members_below),
        excel_members_vlookup=bool(excel_members_vlookup),
    )
    print(f"Excel出力: {out_path}")


@app.command()
def release_from_summary(
    input_file: str = typer.Option(..., help="集計表（.xlsx/.xlsm）。'対戦一覧_短縮' シートを含むこと"),
    sheet_name: str = typer.Option("対戦一覧_短縮", help="読み込む短縮対戦リストのシート名"),
    output_file: str = typer.Option("", help="出力Excel。空なら input_file と同名で _from_summary.xlsx"),
    html_passcode: str = typer.Option("", help="HTMLの簡易ロック用パスコード（注意: 完全な暗号化ではありません）"),
    include_members_html: bool = typer.Option(True, help="HTML（個人/短縮）に選手名（氏名）を含める"),
    wall_courts_per_page: int = typer.Option(2, help="壁貼り用HTMLの1ページあたりコート数（標準2）"),
    wall_team_color: list[str] = typer.Option(
        [],
        "--wall-team-color",
        help="壁貼り用のチーム色分け（複数可）。例: --wall-team-color '蛇口:red' --wall-team-color '上海:blue'",
    ),
    wall_auto_colors: bool = typer.Option(True, "--wall-auto-colors/--no-wall-auto-colors", help="未指定も自動で色付け"),
    wall_cell_background: bool = typer.Option(
        False,
        "--wall-cell-background/--no-wall-cell-background",
        help="壁貼りの背景色。標準OFF（背景印刷がOFFになりがちなので）",
    ),
    score_sheets_per_page: int = typer.Option(8, help="得点記入表の面付け数（標準8）"),
    score_sheets_columns: int = typer.Option(2, help="得点記入表の列数（標準2）"),
    score_sheets_include_members: bool = typer.Option(True, help="得点記入表に選手名（氏名）も入れる"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="終了時刻の計算に使う（分）"),
    allow_court_gaps: bool = typer.Option(True, help="空きコートを許容する（集計表由来ではON推奨）"),
    excel_include_members: bool = typer.Option(False, help="Excelの『対戦表』セルに選手名（氏名）を改行で表示する"),
    excel_members_below: bool = typer.Option(True, help="Excelの『対戦表』を2行構成にする（上=ペア名、下=選手名）"),
    excel_members_vlookup: bool = typer.Option(True, help="『選手名』行をペア一覧からVLOOKUPで自動表示する（要Excel再計算）"),
    start_time: str = typer.Option(DEFAULT_START_TIME_HHMM, help="開始時刻 (HH:MM) ※時間列が欠けている場合の補助"),
):
    """集計表（短縮一覧）から、最終配布セットを一発で出力。

    出力:
      - Excel（_from_summary.xlsx）
      - HTML一式（_from_summary.html）
      - 壁貼りHTML（_from_summary_wall.html）
      - 得点記入表（_from_summary_score_sheets.html）
    """

    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")
    try:
        _parse_hhmm(start_time)
    except ValueError as e:
        raise typer.BadParameter(str(e))

    in_path = Path(input_file)
    if not in_path.exists():
        raise typer.BadParameter(f"入力ファイルが見つかりません: {input_file}")

    out_xlsx = Path(output_file) if output_file else in_path.with_name(f"{in_path.stem}_from_summary.xlsx")

    try:
        matches, teams, num_rounds, courts = load_schedule_from_short_list_xlsx(
            str(in_path),
            sheet_name=str(sheet_name),
            fallback_start_time_hhmm=start_time,
            fallback_round_minutes=int(round_minutes),
        )
    except PermissionError:
        raise typer.BadParameter(
            f"Excelファイルを開いたままの可能性があります。Excelを閉じてから再実行してください: {input_file}"
        )

    if num_rounds <= 0 or courts <= 0 or not matches:
        raise RuntimeError("短縮対戦リストから試合を読み取れませんでした（列名やシート名を確認してください）")

    write_to_excel_like_summary(
        matches,
        teams,
        str(out_xlsx),
        bool(allow_court_gaps),
        num_rounds,
        courts,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        excel_include_members=bool(excel_include_members),
        excel_members_below=bool(excel_members_below),
        excel_members_vlookup=bool(excel_members_vlookup),
        normalize_round_times=False,
    )

    out_html = out_xlsx.with_suffix(".html")
    write_personal_schedule_html(
        matches,
        teams,
        str(out_html),
        num_rounds=num_rounds,
        courts=courts,
        html_passcode=html_passcode or None,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        include_members=bool(include_members_html),
    )

    wall_rules = _parse_wall_team_color_rules(list(wall_team_color))
    out_wall = out_xlsx.with_name(f"{out_xlsx.stem}_wall.html")
    write_wall_schedule_html(
        matches,
        str(out_wall),
        num_rounds=num_rounds,
        courts=courts,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        courts_per_page=int(wall_courts_per_page),
        team_color_rules=wall_rules,
        auto_team_colors=bool(wall_auto_colors),
        cell_background=bool(wall_cell_background),
    )

    out_scores = out_xlsx.with_name(f"{out_xlsx.stem}_score_sheets.html")
    write_score_sheets_html(
        matches,
        teams,
        str(out_scores),
        per_page=int(score_sheets_per_page),
        columns=int(score_sheets_columns),
        include_members=bool(score_sheets_include_members),
        round_minutes=int(round_minutes),
    )

    print(f"Excel出力: {out_xlsx}")
    print(f"HTML一式(対戦表+短縮+個人): {out_html}")
    print(f"壁貼り用HTML(コート別): {out_wall}")
    print(f"得点記入表HTML出力: {out_scores}")
    print(f"読込元: {in_path} / シート: {sheet_name}")
    print(f"総試合数: {len(matches)} / ラウンド数: {num_rounds} / コート数: {courts}")


@app.command()
def release_from_team_list(
    input_file: str = typer.Option("チームリスト.xlsx", help="チーム一覧Excel"),
    output_file: str = typer.Option("graph_schedule.xlsx", help="出力ファイル（ベース名）。タイムスタンプ付きで保存"),
    num_rounds: int = 23,
    courts: int = 15,
    diversity_attempts: int = typer.Option(1, help="分散最大化の試行回数"),
    graph_mode: bool = typer.Option(True, help="グラフ構築モードを使用 (必須条件安定化)"),
    allow_court_gaps: bool = typer.Option(False, help="途中ラウンドの空きコートを許容するか（通常はOFF推奨）"),
    max_consecutive: int = typer.Option(2, help="最大連戦数（2推奨）。満たせない場合は自動で3に緩和（2 or 3）"),
    relax_max_consecutive: bool = typer.Option(
        False,
        "--relax-max-consecutive/--no-relax-max-consecutive",
        help="連戦上限を満たせない場合に自動緩和する（max_consecutive=2 のとき 3 に緩和）。標準はOFF（探せるだけ探す）",
    ),
    matches_per_team: int = typer.Option(0, help="各ペアの試合数。0で自動"),
    html_passcode: str = typer.Option("", help="HTMLの簡易ロック用パスコード（注意: 完全な暗号化ではありません）"),
    include_members_html: bool = typer.Option(True, help="HTMLに選手名（氏名）を含める"),
    wall_courts_per_page: int = typer.Option(2, help="壁貼り用HTMLの1ページあたりコート数（標準2）"),
    wall_team_color: list[str] = typer.Option(
        [],
        "--wall-team-color",
        help="壁貼り用のチーム色分け（複数可）。例: --wall-team-color '蛇口:red' --wall-team-color '上海:blue'",
    ),
    wall_auto_colors: bool = typer.Option(True, "--wall-auto-colors/--no-wall-auto-colors", help="未指定も自動で色付け"),
    wall_cell_background: bool = typer.Option(
        False,
        "--wall-cell-background/--no-wall-cell-background",
        help="壁貼りの背景色。標準OFF（背景印刷がOFFになりがちなので）",
    ),
    score_sheets_per_page: int = typer.Option(8, help="得点記入表の面付け数（標準8）"),
    score_sheets_columns: int = typer.Option(2, help="得点記入表の列数（標準2）"),
    score_sheets_include_members: bool = typer.Option(True, help="得点記入表に選手名（氏名）も入れる"),
    excel_include_members: bool = typer.Option(False, help="Excelの『対戦表』セルに選手名（氏名）を改行で表示する"),
    excel_members_below: bool = typer.Option(True, help="Excelの『対戦表』を2行構成にする（上=ペア名、下=選手名）"),
    excel_members_vlookup: bool = typer.Option(True, help="『選手名』行をペア一覧からVLOOKUPで自動表示する（要Excel再計算）"),
    start_time: str = typer.Option(DEFAULT_START_TIME_HHMM, help="開始時刻 (HH:MM)"),
    round_minutes: int = typer.Option(DEFAULT_ROUND_MINUTES, help="1ラウンドの時間（分）"),
):
    """チームリストからスケジュール生成し、最終配布セットを一発で出力。"""

    if not graph_mode:
        raise typer.BadParameter("現在は graph_mode=True のみをサポートします")
    if round_minutes <= 0:
        raise typer.BadParameter("round_minutes は 1 以上を指定してください")
    try:
        _parse_hhmm(start_time)
    except ValueError as e:
        raise typer.BadParameter(str(e))
    if max_consecutive not in (2, 3):
        raise typer.BadParameter("max_consecutive は 2 または 3 を指定してください")
    if matches_per_team < 0:
        raise typer.BadParameter("matches_per_team は 0 以上を指定してください")

    global TARGET_MATCHES_PER_TEAM

    probe_teams = load_teams(input_file)
    auto_mode = matches_per_team == 0
    initial_target = (
        compute_auto_matches_per_team(len(probe_teams), num_rounds, courts) if auto_mode else matches_per_team
    )

    capacity = num_rounds * courts

    best_matches: List[Match] | None = None
    best_teams: List[Team] | None = None
    best_score: int = -1
    best_key: tuple[int, int, int] | None = None
    best_streak_seen: int | None = None
    tried_targets: Set[int] = set()
    target_candidates = [initial_target] if not auto_mode else list(range(initial_target, 0, -1))

    for target in target_candidates:
        if (len(probe_teams) % 2 == 1) and (target % 2 == 1):
            target -= 1
        if target <= 0 or target in tried_targets:
            continue
        tried_targets.add(target)

        expected_matches_total = expected_total_matches(len(probe_teams), target)
        if (len(probe_teams) * target) % 2 != 0:
            if auto_mode:
                continue
            raise typer.BadParameter(
                f"不正な組合せ: ペア数={len(probe_teams)} と 目標={target} では総試合数が整数になりません（ペア数が奇数の時は試合数は偶数が必要）"
            )
        if expected_matches_total > capacity:
            if auto_mode:
                continue
            raise typer.BadParameter(
                f"容量不足: courts*num_rounds={capacity} 試合枠に対し、必要試合数={expected_matches_total}（ペア数={len(probe_teams)}, 目標={target}試合/ペア）"
            )

        TARGET_MATCHES_PER_TEAM = target
        if target != initial_target:
            print(f"自動調整: 目標試合数/ペアを {target} に下げて再試行します")
        print(
            f"目標試合数/ペア: {TARGET_MATCHES_PER_TEAM}（自動={auto_mode}） / 必要試合数 {expected_matches_total} / 容量 {capacity}"
        )

        best_matches = None
        best_teams = None
        best_score = -1
        best_key = None
        best_streak_seen = None

        for attempt in range(diversity_attempts):
            teams = load_teams(input_file)
            graph_ok = False
            try:
                matches = schedule_matches_graph(teams, num_rounds, courts, seed=attempt)
                refresh_team_stats(teams, matches)
                if all(t.matches == TARGET_MATCHES_PER_TEAM for t in teams) and len(matches) == expected_total_matches(
                    len(teams), TARGET_MATCHES_PER_TEAM
                ):
                    graph_ok = True
            except Exception as e:
                print(f"グラフ試行 {attempt} 失敗: {e}")
            if not graph_ok:
                print(f"フォールバックヒューリスティック使用 (試行 {attempt})")
                teams = load_teams(input_file)
                matches = schedule_matches_heuristic(teams, num_rounds, courts, seed=attempt)
                refresh_team_stats(teams, matches)

            if all(t.matches == TARGET_MATCHES_PER_TEAM for t in teams) and len(matches) == expected_total_matches(
                len(teams), TARGET_MATCHES_PER_TEAM
            ):
                matches = rebalance_vertical_distribution(matches, teams, num_rounds, courts)
                matches = enforce_segments_and_quotas(matches, teams, num_rounds, courts)
                matches = balanced_round_reassignment(matches, num_rounds, courts)
                matches = boost_group_diversity(matches, teams)
                matches = reduce_back_to_back(matches, num_rounds, courts)
                matches = tighten_level_bands(matches, num_rounds, courts)
                collisions = detect_collisions(matches)
                if collisions:
                    print(f"衝突スロット数: {len(collisions)} → 修復試行")
                    matches = repair_collisions(matches, num_rounds, courts)
                    after = detect_collisions(matches)
                    if after:
                        print(f"修復後も残る衝突: {len(after)}")
                matches = normalize_round_capacity(matches, num_rounds, courts)
                matches = ensure_round_one_full(matches, num_rounds, courts)
                if not allow_court_gaps:
                    matches = eliminate_mid_session_court_gaps(matches, num_rounds, courts)
                matches = compact_court_usage(matches, num_rounds, courts)

                def _consecutive_optimize(ms: List[Match], limit: int) -> List[Match]:
                    for _ in range(8):
                        if _max_team_streak(ms) <= limit:
                            break
                        ms = reduce_max_consecutive_streak(
                            ms,
                            num_rounds,
                            courts,
                            max_consecutive=limit,
                            max_iterations=(1600 if limit <= 2 else 800),
                        )
                        ms = normalize_round_capacity(ms, num_rounds, courts)
                        ms = ensure_round_one_full(ms, num_rounds, courts)
                        if not allow_court_gaps:
                            ms = eliminate_mid_session_court_gaps(ms, num_rounds, courts)
                        ms = compact_court_usage(ms, num_rounds, courts)
                    return ms

                matches = _consecutive_optimize(matches, limit=max_consecutive)
                if max_consecutive == 2 and _max_team_streak(matches) > 2:
                    if relax_max_consecutive:
                        print("連戦上限2が満たせないため、連戦上限3に緩和します")
                        matches = _consecutive_optimize(matches, limit=3)

            if any(t.matches != TARGET_MATCHES_PER_TEAM for t in teams):
                print(f"試行 {attempt}: 未達ペアあり -> スキップ")
                continue
            if len(matches) != expected_total_matches(len(teams), TARGET_MATCHES_PER_TEAM):
                print(f"試行 {attempt}: 総試合数 {len(matches)} 不一致")
                continue

            streak = _max_team_streak(matches)
            if best_streak_seen is None or streak < best_streak_seen:
                best_streak_seen = streak

            if (not relax_max_consecutive) and streak > max_consecutive:
                # Strict mode: keep searching for a schedule that truly satisfies the limit.
                continue

            score = compute_diversity_score(teams)
            consec_pairs = _count_consecutive_pairs(matches)
            # Prefer fewer consecutive streaks (avoid 3-in-a-row if possible), then fewer back-to-backs,
            # then maximize diversity.
            key = (int(streak), int(consec_pairs), int(-score))
            if best_key is None or key < best_key:
                best_key = key
                best_score = score
                best_matches = matches
                best_teams = teams

        if best_matches is not None:
            break

    if best_matches is None or best_teams is None:
        hint = ""
        if (not relax_max_consecutive) and max_consecutive == 2:
            hint = f"（補足: この探索では最大連戦の最良値={best_streak_seen}。diversity_attempts を増やして再試行してください）"
        raise RuntimeError(
            "全試行失敗: 条件を満たすスケジュールを構築できませんでした (グラフ+ヒューリスティック)" + hint
        )

    print(f"最大連戦: {_max_team_streak(best_matches)}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_path = Path(output_file)
    stamped_name = (
        f"{base_path.stem}_{timestamp}{base_path.suffix}" if base_path.suffix else f"{base_path.name}_{timestamp}"
    )
    out_xlsx = base_path.with_name(stamped_name)

    refresh_team_stats(best_teams, best_matches)
    apply_round_times(best_matches, start_time_hhmm=start_time, round_minutes=int(round_minutes))

    write_to_excel_like_summary(
        best_matches,
        best_teams,
        str(out_xlsx),
        allow_court_gaps,
        num_rounds,
        courts,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        excel_include_members=bool(excel_include_members),
        excel_members_below=bool(excel_members_below),
        excel_members_vlookup=bool(excel_members_vlookup),
    )

    out_html = out_xlsx.with_suffix(".html")
    write_personal_schedule_html(
        best_matches,
        best_teams,
        str(out_html),
        num_rounds=num_rounds,
        courts=courts,
        html_passcode=html_passcode or None,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        include_members=bool(include_members_html),
    )

    wall_rules = _parse_wall_team_color_rules(list(wall_team_color))
    out_wall = out_xlsx.with_name(f"{out_xlsx.stem}_wall.html")
    write_wall_schedule_html(
        best_matches,
        str(out_wall),
        num_rounds=num_rounds,
        courts=courts,
        start_time_hhmm=start_time,
        round_minutes=int(round_minutes),
        courts_per_page=int(wall_courts_per_page),
        team_color_rules=wall_rules,
        auto_team_colors=bool(wall_auto_colors),
        cell_background=bool(wall_cell_background),
    )

    out_scores = out_xlsx.with_name(f"{out_xlsx.stem}_score_sheets.html")
    write_score_sheets_html(
        best_matches,
        best_teams,
        str(out_scores),
        per_page=int(score_sheets_per_page),
        columns=int(score_sheets_columns),
        include_members=bool(score_sheets_include_members),
        round_minutes=int(round_minutes),
    )

    print(f"スケジュール出力: {out_xlsx}")
    print("  含まれる主要シート: 対戦表 / 個人スケジュール表 / 対戦一覧短縮（チーム順/試合順）")
    print(f"HTML一式(対戦表+短縮+個人): {out_html}")
    print(f"壁貼り用HTML(コート別): {out_wall}")
    print(f"得点記入表HTML出力: {out_scores}")
    print(f"総試合数: {len(best_matches)} / 期待 { expected_total_matches(len(best_teams), TARGET_MATCHES_PER_TEAM) }")
    print(f"分散スコア(総対戦グループ種類合計): {best_score}")

if __name__ == "__main__":
    app()