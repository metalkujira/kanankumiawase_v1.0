from openpyxl import load_workbook

wb = load_workbook('final_fixed.xlsx')
ws = wb['対戦表']
displayed = 0
for r in range(2, ws.max_row):
    for c in range(4, ws.max_column + 1, 2):
        v1 = ws.cell(row=r, column=c).value
        v2 = ws.cell(row=r, column=c + 1).value if c + 1 <= ws.max_column else None
        if v1 and v2:
            displayed += 1
print('Displayed matches counted:', displayed)
print('Expected (len matches cell value):', ws.cell(row=ws.max_row, column=4).value)
print('不足:', ws.cell(row=ws.max_row, column=6).value)