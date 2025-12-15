"""Create a scheduler input Excel (チームリスト.xlsx) from an existing workbook.

This script is intentionally generic so it can be shared publicly.

Usage (PowerShell):
  python create_team_list.py --input-file "source.xlsm" --output-file "チームリスト.xlsx"
"""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
import pandas as pd


def main() -> int:
    parser = argparse.ArgumentParser(description="Create チームリスト.xlsx from a source workbook")
    parser.add_argument(
        "--input-file",
        type=str,
        default="source.xlsm",
        help="Input workbook path (.xlsx/.xlsm). Default: source.xlsm",
    )
    parser.add_argument(
        "--output-file",
        type=str,
        default="チームリスト.xlsx",
        help="Output .xlsx path. Default: チームリスト.xlsx",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default="リスト",
        help="Sheet name to read. Default: リスト",
    )
    args = parser.parse_args()

    input_path = Path(args.input_file)
    output_path = Path(args.output_file)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    wb = openpyxl.load_workbook(str(input_path), read_only=True)
    if args.sheet not in wb.sheetnames:
        raise KeyError(f"Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")

    sheet = wb[args.sheet]
    data = list(sheet.values)
    if not data:
        raise ValueError("Sheet is empty")

    df = pd.DataFrame(data[1:], columns=data[0])
    valid_teams = df.dropna(subset=[df.columns[0]])

    wb_new = openpyxl.Workbook()
    ws = wb_new.active
    ws.title = "チームリスト"
    ws.append(["ペア名", "氏名", "優先対戦", "優先対戦相手"])

    for _, row in valid_teams.iterrows():
        name = row.get("ペア名 ↓値ばりで記入")
        members = row.get("氏名　↓値ばりで記入")
        if pd.isna(name):
            continue
        name = str(name)
        members = "" if pd.isna(members) else str(members)

        ws.append([name, members, "", ""])

    wb_new.save(str(output_path))
    print(f"Created: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())